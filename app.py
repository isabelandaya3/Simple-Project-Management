"""
LEB RFI/Submittal Tracker - Local Windows Application
=====================================================
A local-only tracker that monitors Outlook emails from Autodesk Construction Cloud
and provides a web dashboard for managing RFIs and Submittals.

Run: python app.py
Access: http://localhost:5000
"""

import os
import re
import json
import sqlite3
import hashlib
import secrets
import threading
import time
from datetime import datetime, timedelta
from pathlib import Path
from functools import wraps

from flask import Flask, request, jsonify, send_from_directory, session, redirect, url_for, render_template_string
import bcrypt

# Optional: dateutil for flexible date parsing
try:
    from dateutil import parser as date_parser
    HAS_DATEUTIL = True
except ImportError:
    HAS_DATEUTIL = False

# Optional: win32com for Outlook integration
try:
    import win32com.client
    import pythoncom
    HAS_WIN32COM = True
except ImportError:
    HAS_WIN32COM = False
    print("WARNING: pywin32 not installed. Email polling will be disabled.")
    print("Install with: pip install pywin32")

# Optional: Windows toast notifications
try:
    from winotify import Notification, audio
    HAS_WINOTIFY = True
except ImportError:
    HAS_WINOTIFY = False
    print("WARNING: winotify not installed. System notifications will be disabled.")
    print("Install with: pip install winotify")

# Optional: Airtable integration for offline form submission
try:
    from airtable_integration import get_airtable_form_url, sync_airtable_responses, load_airtable_config
    HAS_AIRTABLE = True
except ImportError:
    HAS_AIRTABLE = False
    print("INFO: Airtable integration not available.")

# Optional: openpyxl for Excel file updates
try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("INFO: openpyxl not installed. Excel tracker updates will be disabled.")
    print("Install with: pip install openpyxl")

# =============================================================================
# CONFIGURATION
# =============================================================================

BASE_DIR = Path(__file__).parent.absolute()
DATABASE_PATH = BASE_DIR / "tracker.db"
CONFIG_PATH = BASE_DIR / "config.json"

# Default configuration
DEFAULT_CONFIG = {
    "base_folder_path": r"\\sac-filsrv1\Projects\Structural-028\Projects\LEB\9.0_Const_Svcs",
    "outlook_folder": "Inbox",  # Can be "Inbox" or a subfolder like "LEB ACC"
    "poll_interval_minutes": 5,
    "server_port": 5000,
    "project_name": "LEB – Local Tracker",
    "rfi_tracker_excel_path": r"\\sac-filsrv1\Projects\Structural-028\Projects\LEB\9.0_Const_Svcs\LEB RFI Bulletin Tracker.xlsx"
}

def load_config():
    """Load configuration from config.json or create default."""
    if CONFIG_PATH.exists():
        with open(CONFIG_PATH, 'r') as f:
            config = json.load(f)
            # Merge with defaults for any missing keys
            for key, value in DEFAULT_CONFIG.items():
                if key not in config:
                    config[key] = value
            return config
    else:
        # Create default config file
        with open(CONFIG_PATH, 'w') as f:
            json.dump(DEFAULT_CONFIG, f, indent=2)
        return DEFAULT_CONFIG.copy()

CONFIG = load_config()

# =============================================================================
# REVIEW DUE DATE CONFIGURATION
# =============================================================================

# Priority-based minimum work days required from date_received to contractor due_date
PRIORITY_MIN_DAYS = {
    'High': 5,
    'Medium': 10,
    'Low': 15,
    None: 10  # Default to Medium if no priority set
}

# QCR requires at least this many work days before contractor due date
QCR_DAYS_BEFORE_DUE = 1

# QCR needs this many days to complete their review (interval between Reviewer due and QCR due)
QCR_REVIEW_DAYS = 2

# =============================================================================
# WORKDAY HELPER FUNCTIONS
# =============================================================================

def format_date_for_email(date_str):
    """Format a date string for display in emails (e.g., 'Wed, 1/19/26')."""
    if not date_str:
        return 'N/A'
    try:
        if isinstance(date_str, str):
            date_obj = datetime.strptime(date_str, '%Y-%m-%d')
        else:
            date_obj = date_str
        # Format: Wed, 1/19/26 (abbreviated day name, no leading zeros)
        day_name = date_obj.strftime('%a')  # Abbreviated day name
        return f"{day_name}, {date_obj.month}/{date_obj.day}/{date_obj.strftime('%y')}"
    except:
        return date_str or 'N/A'


# =============================================================================
# RFI BULLETIN TRACKER EXCEL UPDATE
# =============================================================================

def update_rfi_tracker_excel(item, action='close'):
    """
    Update the RFI Bulletin Tracker Excel file when an RFI is closed or reopened.
    
    Args:
        item: Dictionary containing item data (must have type='RFI')
        action: 'close' to add/update entry, 'reopen' to mark as reopened
    
    Returns:
        dict with 'success' and optional 'error' or 'message'
    """
    if not HAS_OPENPYXL:
        return {'success': False, 'error': 'openpyxl not installed'}
    
    # Only process RFIs
    if item.get('type') != 'RFI':
        return {'success': True, 'message': 'Not an RFI, skipping Excel update'}
    
    excel_path = CONFIG.get('rfi_tracker_excel_path')
    if not excel_path:
        return {'success': False, 'error': 'RFI tracker Excel path not configured'}
    
    excel_file = Path(excel_path)
    if not excel_file.exists():
        return {'success': False, 'error': f'Excel file not found: {excel_path}'}
    
    try:
        # Extract RFI number from identifier (e.g., "RFI #123" -> "123")
        identifier = item.get('identifier', '')
        rfi_number = ''
        rfi_match = re.search(r'#?(\d+)', identifier)
        if rfi_match:
            rfi_number = rfi_match.group(1)
        else:
            rfi_number = identifier  # Use as-is if no number found
        
        # Get title (name after numbers)
        title = item.get('title', '') or ''
        
        # Get the contractor's question - use source_subject or title as fallback
        # The question typically comes from the original email subject/body
        question = item.get('source_subject', '') or title
        
        # Get the final response
        response = item.get('final_response_text', '') or item.get('response_text', '') or ''
        
        # Load the workbook
        wb = load_workbook(excel_path)
        ws = wb.active  # Use the active sheet, or specify by name: wb['SheetName']
        
        # Find headers to determine columns (look in first few rows)
        header_row = None
        col_rfi_id = None
        col_title = None
        col_question = None
        col_response = None
        col_status = None
        
        for row_num in range(1, 6):  # Check first 5 rows for headers
            for col_num in range(1, 20):  # Check first 20 columns
                cell_value = str(ws.cell(row=row_num, column=col_num).value or '').strip().lower()
                if 'rfi' in cell_value and ('id' in cell_value or '#' in cell_value or 'number' in cell_value or cell_value == 'rfi'):
                    col_rfi_id = col_num
                    header_row = row_num
                elif cell_value in ('title', 'name', 'subject', 'description'):
                    col_title = col_num
                    header_row = row_num
                elif 'question' in cell_value or 'query' in cell_value:
                    col_question = col_num
                    header_row = row_num
                elif 'response' in cell_value or 'answer' in cell_value or 'reply' in cell_value:
                    col_response = col_num
                    header_row = row_num
                elif 'status' in cell_value:
                    col_status = col_num
                    header_row = row_num
        
        if not header_row:
            # Default to assuming row 1 is headers with standard column layout
            header_row = 1
            col_rfi_id = 1
            col_title = 2
            col_question = 3
            col_response = 4
            col_status = 5
        
        # Find existing row with this RFI number, or find first empty row
        existing_row = None
        first_empty_row = None
        
        for row_num in range(header_row + 1, ws.max_row + 2):
            cell_value = ws.cell(row=row_num, column=col_rfi_id or 1).value
            if cell_value is not None:
                # Check if this RFI already exists
                if str(cell_value).strip() == rfi_number:
                    existing_row = row_num
                    break
            else:
                # Found an empty row
                if first_empty_row is None:
                    first_empty_row = row_num
        
        if first_empty_row is None:
            first_empty_row = ws.max_row + 1
        
        target_row = existing_row or first_empty_row
        
        if action == 'close':
            # Add or update the RFI entry
            if col_rfi_id:
                ws.cell(row=target_row, column=col_rfi_id).value = rfi_number
            if col_title:
                ws.cell(row=target_row, column=col_title).value = title
            if col_question:
                ws.cell(row=target_row, column=col_question).value = question
            if col_response:
                ws.cell(row=target_row, column=col_response).value = response
            if col_status:
                ws.cell(row=target_row, column=col_status).value = 'Closed'
            
            action_msg = 'updated' if existing_row else 'added'
            
        elif action == 'reopen':
            # Mark the entry as reopened (if it exists)
            if existing_row and col_status:
                ws.cell(row=target_row, column=col_status).value = 'Reopened'
                action_msg = 'marked as reopened'
            else:
                action_msg = 'no existing entry to update'
        
        # Save the workbook
        wb.save(excel_path)
        wb.close()
        
        return {
            'success': True, 
            'message': f'RFI {rfi_number} {action_msg} in Excel tracker (row {target_row})'
        }
        
    except PermissionError:
        return {'success': False, 'error': 'Excel file is open by another user. Please close it and try again.'}
    except Exception as e:
        return {'success': False, 'error': f'Failed to update Excel: {str(e)}'}


def is_business_day(date):
    """Check if a date is a business day (Mon-Fri)."""
    return date.weekday() < 5  # 0=Monday, 4=Friday

def business_days_between(start_date, end_date):
    """
    Count business days (Mon-Fri) from start_date to end_date.
    Inclusive of start_date, exclusive of end_date.
    """
    if isinstance(start_date, str):
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if isinstance(end_date, str):
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    if hasattr(start_date, 'date'):
        start_date = start_date.date()
    if hasattr(end_date, 'date'):
        end_date = end_date.date()
    
    if start_date >= end_date:
        return 0
    
    count = 0
    current = start_date
    while current < end_date:
        if is_business_day(current):
            count += 1
        current += timedelta(days=1)
    return count

def subtract_business_days(end_date, n):
    """
    Return the date that is n business days before end_date.
    If n=3 and end_date is Friday, result is Tuesday (skipping Sat/Sun).
    """
    if isinstance(end_date, str):
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    if hasattr(end_date, 'date'):
        end_date = end_date.date()
    
    current = end_date
    days_subtracted = 0
    while days_subtracted < n:
        current -= timedelta(days=1)
        if is_business_day(current):
            days_subtracted += 1
    return current

def add_business_days(start_date, n):
    """
    Return the date that is n business days after start_date.
    """
    if isinstance(start_date, str):
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    if hasattr(start_date, 'date'):
        start_date = start_date.date()
    
    current = start_date
    days_added = 0
    while days_added < n:
        current += timedelta(days=1)
        if is_business_day(current):
            days_added += 1
    return current

def calculate_review_due_dates(date_received, contractor_due_date, priority):
    """
    Calculate internal due dates for Initial Reviewer and QCR.
    
    Returns dict with:
        - initial_reviewer_due_date
        - qcr_due_date
        - contractor_window_days
        - is_contractor_window_insufficient
        - required_days
    """
    if not date_received or not contractor_due_date:
        return {
            'initial_reviewer_due_date': None,
            'qcr_due_date': None,
            'contractor_window_days': None,
            'is_contractor_window_insufficient': False,
            'required_days': None
        }
    
    # Parse dates
    if isinstance(date_received, str):
        date_received = datetime.strptime(date_received, '%Y-%m-%d').date()
    if isinstance(contractor_due_date, str):
        contractor_due_date = datetime.strptime(contractor_due_date, '%Y-%m-%d').date()
    if hasattr(date_received, 'date'):
        date_received = date_received.date()
    if hasattr(contractor_due_date, 'date'):
        contractor_due_date = contractor_due_date.date()
    
    # Get required minimum days based on priority
    required_days = PRIORITY_MIN_DAYS.get(priority, PRIORITY_MIN_DAYS[None])
    
    # Calculate contractor window (business days available)
    contractor_window_days = business_days_between(date_received, contractor_due_date)
    
    # Check if window is insufficient
    is_insufficient = contractor_window_days < required_days
    
    # Calculate QCR due date (1 business day before contractor due date)
    qcr_due_date = subtract_business_days(contractor_due_date, QCR_DAYS_BEFORE_DUE)
    
    # Calculate Initial Reviewer due date
    # QCR needs QCR_REVIEW_DAYS (2) business days to review, so reviewer must submit 2 days before QCR due
    initial_reviewer_due_date = subtract_business_days(qcr_due_date, QCR_REVIEW_DAYS)
    
    # Ensure reviewer due date is not before date_received
    if initial_reviewer_due_date < date_received:
        initial_reviewer_due_date = date_received
    
    return {
        'initial_reviewer_due_date': initial_reviewer_due_date.strftime('%Y-%m-%d'),
        'qcr_due_date': qcr_due_date.strftime('%Y-%m-%d'),
        'contractor_window_days': contractor_window_days,
        'is_contractor_window_insufficient': is_insufficient,
        'required_days': required_days
    }

def get_due_date_status(due_date):
    """
    Get status color for a due date based on business days remaining.
    Returns: 'green', 'yellow', 'red', or None if no due date.
    """
    if not due_date:
        return None
    
    if isinstance(due_date, str):
        due_date = datetime.strptime(due_date, '%Y-%m-%d').date()
    if hasattr(due_date, 'date'):
        due_date = due_date.date()
    
    today = datetime.now().date()
    days_remaining = business_days_between(today, due_date)
    
    if due_date < today:
        return 'red'  # Overdue
    elif days_remaining <= 3:
        return 'yellow'  # 1-3 days remaining
    else:
        return 'green'  # More than 3 days

# =============================================================================
# FLASK APP SETUP
# =============================================================================

app = Flask(__name__, static_folder='static', template_folder='templates')

# Use a persistent secret key (stored in config or generate once)
SECRET_KEY_FILE = BASE_DIR / ".secret_key"
if SECRET_KEY_FILE.exists():
    app.secret_key = SECRET_KEY_FILE.read_text().strip()
else:
    app.secret_key = secrets.token_hex(32)
    SECRET_KEY_FILE.write_text(app.secret_key)

app.config['SESSION_TYPE'] = 'filesystem'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=24)
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SECURE'] = False  # Set True if using HTTPS

# =============================================================================
# EMAIL RETRY QUEUE - For handling failed email sends
# =============================================================================

# In-memory queue of pending emails to retry
PENDING_EMAILS = []
PENDING_EMAILS_LOCK = threading.Lock()
MAX_EMAIL_RETRIES = 3

def queue_pending_email(email_type, item_id, **kwargs):
    """Add an email to the retry queue."""
    with PENDING_EMAILS_LOCK:
        # Check if already queued
        for pending in PENDING_EMAILS:
            if pending['email_type'] == email_type and pending['item_id'] == item_id:
                return  # Already queued
        
        PENDING_EMAILS.append({
            'email_type': email_type,
            'item_id': item_id,
            'retries': 0,
            'last_attempt': None,
            'kwargs': kwargs
        })
        print(f"  [EmailQueue] Queued {email_type} for item {item_id}")

def send_email_with_retry(send_func, item_id, email_type, max_retries=3, **kwargs):
    """Try to send an email with retry logic. Queue for later if all retries fail."""
    import time as time_module
    
    for attempt in range(max_retries):
        try:
            result = send_func(item_id, **kwargs)
            if result.get('success'):
                return result
            else:
                error = result.get('error', 'Unknown error')
                # If it's an RPC/COM error, retry
                if 'remote procedure call' in str(error).lower() or '-2147' in str(error):
                    if attempt < max_retries - 1:
                        print(f"  [EmailRetry] Attempt {attempt + 1} failed for {email_type} item {item_id}: {error}. Retrying...")
                        time_module.sleep(2 ** attempt)  # Exponential backoff: 1s, 2s, 4s
                        continue
                return result  # Non-recoverable error
        except Exception as e:
            error_str = str(e)
            if 'remote procedure call' in error_str.lower() or '-2147' in error_str:
                if attempt < max_retries - 1:
                    print(f"  [EmailRetry] Attempt {attempt + 1} exception for {email_type} item {item_id}: {e}. Retrying...")
                    time_module.sleep(2 ** attempt)
                    continue
            return {'success': False, 'error': str(e)}
    
    # All retries failed, queue for later
    queue_pending_email(email_type, item_id, **kwargs)
    return {'success': False, 'error': 'All retries failed, queued for later', 'queued': True}

def process_pending_emails():
    """Process any pending emails in the queue. Called by folder watcher."""
    with PENDING_EMAILS_LOCK:
        if not PENDING_EMAILS:
            return
        
        to_remove = []
        for pending in PENDING_EMAILS:
            # Skip if attempted recently (wait at least 30 seconds between attempts)
            if pending['last_attempt']:
                elapsed = (datetime.now() - pending['last_attempt']).total_seconds()
                if elapsed < 30:
                    continue
            
            pending['last_attempt'] = datetime.now()
            pending['retries'] += 1
            
            email_type = pending['email_type']
            item_id = pending['item_id']
            kwargs = pending.get('kwargs', {})
            
            print(f"  [EmailQueue] Retrying {email_type} for item {item_id} (attempt {pending['retries']})")
            
            try:
                if email_type == 'multi_reviewer_qcr':
                    result = send_multi_reviewer_qcr_email(item_id)
                elif email_type == 'qcr_assignment':
                    result = send_qcr_assignment_email(item_id, **kwargs)
                elif email_type == 'multi_reviewer_sendback':
                    result = send_multi_reviewer_sendback_emails(item_id, kwargs.get('feedback', ''), kwargs.get('reviewer_ids'))
                else:
                    print(f"  [EmailQueue] Unknown email type: {email_type}")
                    to_remove.append(pending)
                    continue
                
                if result.get('success'):
                    print(f"  [EmailQueue] Successfully sent {email_type} for item {item_id}")
                    to_remove.append(pending)
                elif pending['retries'] >= MAX_EMAIL_RETRIES:
                    print(f"  [EmailQueue] Giving up on {email_type} for item {item_id} after {MAX_EMAIL_RETRIES} attempts")
                    to_remove.append(pending)
                    # Create a notification about the failed email
                    create_notification(
                        'email_failed',
                        f'⚠️ Email Failed: Item {item_id}',
                        f'Failed to send {email_type} email after {MAX_EMAIL_RETRIES} attempts. Please send manually.',
                        item_id=item_id
                    )
            except Exception as e:
                print(f"  [EmailQueue] Error retrying {email_type} for item {item_id}: {e}")
                if pending['retries'] >= MAX_EMAIL_RETRIES:
                    to_remove.append(pending)
        
        for pending in to_remove:
            PENDING_EMAILS.remove(pending)

# =============================================================================
# HTML TEMPLATES FOR MAGIC-LINK RESPONSE PAGES
# =============================================================================

BASE_PAGE_STYLE = '''
<style>
    * { box-sizing: border-box; margin: 0; padding: 0; }
    body {
        font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, Ubuntu, sans-serif;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        min-height: 100vh;
        display: flex;
        align-items: center;
        justify-content: center;
        padding: 20px;
    }
    .container {
        background: white;
        border-radius: 16px;
        box-shadow: 0 25px 50px rgba(0,0,0,0.15);
        padding: 40px;
        max-width: 700px;
        width: 100%;
    }
    h1 {
        color: #1a1a2e;
        font-size: 24px;
        margin-bottom: 8px;
    }
    .subtitle {
        color: #666;
        font-size: 14px;
        margin-bottom: 24px;
    }
    .info-box {
        background: #f8f9fa;
        border-radius: 8px;
        padding: 16px;
        margin-bottom: 24px;
    }
    .info-row {
        display: flex;
        margin-bottom: 8px;
    }
    .info-row:last-child { margin-bottom: 0; }
    .info-label {
        font-weight: 600;
        color: #444;
        width: 140px;
        flex-shrink: 0;
    }
    .info-value {
        color: #666;
    }
    .section-title {
        font-size: 16px;
        font-weight: 600;
        color: #1a1a2e;
        margin-bottom: 12px;
    }
    .form-group {
        margin-bottom: 20px;
    }
    label {
        display: block;
        font-weight: 500;
        color: #444;
        margin-bottom: 6px;
    }
    select, textarea {
        width: 100%;
        padding: 10px 12px;
        border: 1px solid #ddd;
        border-radius: 8px;
        font-size: 14px;
        font-family: inherit;
    }
    select:focus, textarea:focus {
        outline: none;
        border-color: #667eea;
        box-shadow: 0 0 0 3px rgba(102,126,234,0.1);
    }
    textarea { resize: vertical; min-height: 100px; }
    .file-list {
        background: #f8f9fa;
        border-radius: 8px;
        padding: 12px;
        max-height: 200px;
        overflow-y: auto;
    }
    .file-item {
        display: flex;
        align-items: center;
        padding: 8px;
        border-radius: 6px;
        margin-bottom: 4px;
    }
    .file-item:hover { background: #e9ecef; }
    .file-item input { margin-right: 10px; }
    .file-item label {
        margin-bottom: 0;
        font-weight: normal;
        color: #333;
        cursor: pointer;
    }
    .btn {
        display: inline-block;
        padding: 12px 24px;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        border: none;
        border-radius: 8px;
        font-size: 16px;
        font-weight: 600;
        cursor: pointer;
        transition: transform 0.2s, box-shadow 0.2s;
    }
    .btn:hover {
        transform: translateY(-2px);
        box-shadow: 0 5px 20px rgba(102,126,234,0.4);
    }
    .reviewer-notes-box {
        background: #fff3cd;
        border: 1px solid #ffc107;
        border-radius: 8px;
        padding: 16px;
        margin-bottom: 24px;
    }
    .reviewer-notes-box h3 {
        font-size: 14px;
        color: #856404;
        margin-bottom: 8px;
    }
    .reviewer-notes-box p {
        color: #856404;
        font-size: 14px;
        white-space: pre-wrap;
    }
    .error-container {
        text-align: center;
    }
    .error-icon {
        font-size: 64px;
        margin-bottom: 20px;
    }
    .success-container {
        text-align: center;
    }
    .success-icon {
        font-size: 64px;
        margin-bottom: 20px;
    }
</style>
'''

ERROR_PAGE_TEMPLATE = '''
<!DOCTYPE html>
<html>
<head>
    <title>Error - LEB Tracker</title>
    ''' + BASE_PAGE_STYLE + '''
</head>
<body>
    <div class="container error-container">
        <div class="error-icon">❌</div>
        <h1>Error</h1>
        <p style="color: #666; margin-top: 12px;">{{ error }}</p>
    </div>
</body>
</html>
'''

ALREADY_RESPONDED_TEMPLATE = '''
<!DOCTYPE html>
<html>
<head>
    <title>Already Submitted - LEB Tracker</title>
    ''' + BASE_PAGE_STYLE + '''
</head>
<body>
    <div class="container success-container">
        <div class="success-icon">✅</div>
        <h1>Already Submitted</h1>
        <p style="color: #666; margin-top: 12px;">
            This {{ 'review' if response_type == 'reviewer' else 'QC review' }} for 
            <strong>{{ item.type }} {{ item.identifier }}</strong> has already been submitted.
        </p>
    </div>
</body>
</html>
'''

SUCCESS_TEMPLATE = '''
<!DOCTYPE html>
<html>
<head>
    <title>Success - LEB Tracker</title>
    ''' + BASE_PAGE_STYLE + '''
</head>
<body>
    <div class="container success-container">
        <div class="success-icon">✅</div>
        <h1>{{ message }}</h1>
        <p style="color: #666; margin-top: 12px;">{{ details }}</p>
    </div>
</body>
</html>
'''

REVIEWER_RESPONSE_TEMPLATE = '''
<!DOCTYPE html>
<html>
<head>
    <title>Review Response - {{ item.type }} {{ item.identifier }}</title>
    ''' + BASE_PAGE_STYLE + '''
    <style>
        .version-badge {
            display: inline-block;
            background: #667eea;
            color: white;
            padding: 4px 12px;
            border-radius: 20px;
            font-size: 12px;
            font-weight: 600;
            margin-left: 10px;
        }
        .resubmit-notice {
            background: #fef3c7;
            border: 1px solid #f59e0b;
            border-radius: 8px;
            padding: 15px;
            margin-bottom: 20px;
        }
        .resubmit-notice h3 {
            margin: 0 0 8px 0;
            color: #92400e;
        }
        .previous-response {
            background: #f0f9ff;
            border: 1px solid #bae6fd;
            border-radius: 8px;
            padding: 15px;
            margin-bottom: 20px;
        }
        .previous-response h4 {
            margin: 0 0 10px 0;
            color: #0369a1;
        }
        .qcr-feedback {
            background: #fef2f2;
            border: 1px solid #fecaca;
            border-radius: 8px;
            padding: 15px;
            margin-bottom: 20px;
        }
        .qcr-feedback h4 {
            margin: 0 0 10px 0;
            color: #991b1b;
        }
        .version-history {
            font-size: 12px;
            color: #666;
            margin-top: 10px;
        }
        .closed-notice {
            background: #f3f4f6;
            border: 1px solid #d1d5db;
            border-radius: 8px;
            padding: 20px;
            text-align: center;
        }
        .closed-notice h3 {
            color: #6b7280;
            margin: 0 0 10px 0;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>Initial Review Response <span class="version-badge">v{{ version }}</span></h1>
        <p class="subtitle">{{ item.type }} {{ item.identifier }}</p>
        
        {% if is_closed %}
        <div class="closed-notice">
            <h3>🔒 This Item Has Been Closed</h3>
            <p>No further changes can be submitted. Contact the project administrator if this is unexpected.</p>
        </div>
        {% elif not can_submit %}
        <div class="closed-notice">
            <h3>⚠️ Submission Not Allowed</h3>
            <p>This item has been finalized in QC. Contact the project administrator if additional changes are required.</p>
        </div>
        {% else %}
        
        {% if is_resubmit and qcr_feedback %}
        <div class="qcr-feedback">
            <h4>↩️ QC Reviewer Requested Revisions</h4>
            <p><strong>Feedback on your v{{ version - 1 }} response:</strong></p>
            <div style="background: white; padding: 10px; border-radius: 4px; margin-top: 8px;">
                {{ qcr_feedback|replace('\n', '<br>')|safe }}
            </div>
        </div>
        {% endif %}
        
        {% if is_resubmit and previous_response %}
        <div class="previous-response">
            <h4>📄 Your Previous Response (v{{ version - 1 }})</h4>
            <p><strong>Category:</strong> {{ previous_response.category or 'N/A' }}</p>
            <p><strong>Notes:</strong></p>
            <div style="background: white; padding: 10px; border-radius: 4px; margin-top: 8px;">
                {{ (previous_response.text or 'No notes')|replace('\n', '<br>')|safe }}
            </div>
            {% if version_history %}
            <div class="version-history">
                <strong>Version History:</strong> {{ version_history }}
            </div>
            {% endif %}
        </div>
        {% endif %}
        
        {% if is_resubmit %}
        <div class="resubmit-notice">
            <h3>📝 Resubmitting Response</h3>
            <p>You are updating your response to version {{ version }}. The QC Reviewer will be notified of your changes.</p>
        </div>
        {% endif %}
        
        <div class="info-box">
            <div class="info-row">
                <span class="info-label">Title:</span>
                <span class="info-value">{{ item.title or 'N/A' }}</span>
            </div>
            <div class="info-row">
                <span class="info-label">Date Received:</span>
                <span class="info-value">{{ item.date_received or 'N/A' }}</span>
            </div>
            <div class="info-row">
                <span class="info-label">Initial Review Due Date:</span>
                <span class="info-value">{{ item.initial_reviewer_due_date or 'N/A' }}</span>
            </div>
            <div class="info-row">
                <span class="info-label">Folder:</span>
                <span class="info-value">{{ item.folder_link or 'N/A' }}</span>
            </div>
        </div>
        
        <form method="POST">
            <input type="hidden" name="token" value="{{ token }}">
            
            <div class="form-group">
                <label for="response_category">Response Category *</label>
                <select name="response_category" id="response_category" required>
                    <option value="">-- Select --</option>
                    <option value="Approved" {% if previous_response and previous_response.category == 'Approved' %}selected{% endif %}>Approved</option>
                    <option value="Approved as Noted" {% if previous_response and previous_response.category == 'Approved as Noted' %}selected{% endif %}>Approved as Noted</option>
                    <option value="For Record Only" {% if previous_response and previous_response.category == 'For Record Only' %}selected{% endif %}>For Record Only</option>
                    <option value="Rejected" {% if previous_response and previous_response.category == 'Rejected' %}selected{% endif %}>Rejected</option>
                    <option value="Revise and Resubmit" {% if previous_response and previous_response.category == 'Revise and Resubmit' %}selected{% endif %}>Revise and Resubmit</option>
                </select>
            </div>
            
            <div class="form-group">
                <p class="section-title">Select Files to Include in Response</p>
                <div class="file-list">
                    {% if files %}
                        {% for file in files %}
                        <div class="file-item">
                            <input type="checkbox" name="selected_files" value="{{ file }}" id="file_{{ loop.index }}" {% if previous_files and file in previous_files %}checked{% endif %}>
                            <label for="file_{{ loop.index }}">{{ file }}</label>
                        </div>
                        {% endfor %}
                    {% else %}
                        <p style="color: #666; padding: 8px;">No files found in folder. Please add your response documents first.</p>
                    {% endif %}
                </div>
            </div>
            
            <div class="form-group">
                <label for="notes">Description</label>
                <textarea name="notes" id="notes" placeholder="Provide response description for the official record...">{{ previous_response.text if previous_response else '' }}</textarea>
            </div>
            
            <div class="form-group">
                <label for="internal_notes">Internal Notes <span style="font-weight: normal; color: #888;">(not shared externally)</span></label>
                <textarea name="internal_notes" id="internal_notes" placeholder="Add any internal notes for the QC Reviewer...">{{ item.reviewer_internal_notes or '' }}</textarea>
            </div>
            
            <button type="submit" class="btn">{% if is_resubmit %}Submit Revision (v{{ version }}){% else %}Submit Review{% endif %}</button>
        </form>
        {% endif %}
    </div>
</body>
</html>
'''

QCR_RESPONSE_TEMPLATE = '''
<!DOCTYPE html>
<html>
<head>
    <title>QC Review - {{ item.type }} {{ item.identifier }}</title>
    ''' + BASE_PAGE_STYLE + '''
    <style>
        .reviewer-response-box {
            background: #f0fdf4;
            border: 1px solid #86efac;
            border-radius: 8px;
            padding: 15px;
            margin: 20px 0;
        }
        .reviewer-response-box h3 {
            margin: 0 0 12px 0;
            color: #166534;
        }
        .action-group {
            background: #fef3c7;
            border: 1px solid #fbbf24;
            border-radius: 8px;
            padding: 15px;
            margin: 20px 0;
        }
        .action-group h3 {
            margin: 0 0 12px 0;
            color: #92400e;
        }
        .radio-group {
            display: flex;
            flex-direction: column;
            gap: 10px;
        }
        .radio-option {
            display: flex;
            align-items: flex-start;
            gap: 10px;
            padding: 10px;
            background: white;
            border-radius: 6px;
            border: 1px solid #e5e7eb;
            cursor: pointer;
        }
        .radio-option:hover {
            border-color: #667eea;
        }
        .radio-option input[type="radio"] {
            margin-top: 3px;
        }
        .radio-option-content {
            flex: 1;
        }
        .radio-option-label {
            font-weight: 600;
            color: #333;
        }
        .radio-option-desc {
            font-size: 12px;
            color: #666;
            margin-top: 4px;
        }
        .response-text-container {
            margin-top: 15px;
        }
        .response-text-label {
            font-weight: 600;
            margin-bottom: 8px;
            color: #333;
        }
        .response-text-readonly {
            background: #f9fafb;
            border: 1px solid #e5e7eb;
            border-radius: 6px;
            padding: 12px;
            white-space: pre-wrap;
            color: #666;
            min-height: 80px;
        }
        #response_text_area {
            display: none;
        }
        .send-back-warning {
            display: none;
            background: #fef2f2;
            border: 1px solid #fecaca;
            border-radius: 8px;
            padding: 12px;
            margin-top: 15px;
            color: #991b1b;
        }
        .version-badge {
            display: inline-block;
            background: #667eea;
            color: white;
            padding: 4px 12px;
            border-radius: 20px;
            font-size: 12px;
            font-weight: 600;
            margin-left: 10px;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>QC Review <span class="version-badge">v{{ version }}</span></h1>
        <p class="subtitle">{{ item.type }} {{ item.identifier }}</p>
        
        {% if version_history %}
        <div style="background: #f0f9ff; border: 1px solid #bae6fd; border-radius: 8px; padding: 12px; margin-bottom: 15px; font-size: 13px;">
            <strong>📋 Reviewer Response Version History:</strong><br>
            Current: <strong>v{{ version }}</strong> ({{ item.reviewer_response_at[:16]|replace('T', ' ') if item.reviewer_response_at else 'N/A' }})<br>
            {% for v in version_history %}
            Previous: v{{ v.version }} ({{ v.submitted_at[:16]|replace('T', ' ') }}) - {{ v.response_category or 'N/A' }}{% if not loop.last %}<br>{% endif %}
            {% endfor %}
        </div>
        {% endif %}
        
        <div class="info-box">
            <div class="info-row">
                <span class="info-label">Title:</span>
                <span class="info-value">{{ item.title or 'N/A' }}</span>
            </div>
            <div class="info-row">
                <span class="info-label">Date Received:</span>
                <span class="info-value">{{ item.date_received or 'N/A' }}</span>
            </div>
            <div class="info-row">
                <span class="info-label">Priority:</span>
                <span class="info-value">{{ item.priority or 'Normal' }}</span>
            </div>
            <div class="info-row">
                <span class="info-label">Initial Reviewer:</span>
                <span class="info-value">{{ item.reviewer_name or 'N/A' }}</span>
            </div>
            <div class="info-row">
                <span class="info-label">QC Reviewer:</span>
                <span class="info-value">{{ item.qcr_name or 'N/A' }}</span>
            </div>
            <div class="info-row">
                <span class="info-label">QC Due Date:</span>
                <span class="info-value" style="color: #d97706;">{{ item.qcr_due_date or 'N/A' }}</span>
            </div>
            <div class="info-row">
                <span class="info-label">Contractor Due Date:</span>
                <span class="info-value">{{ item.due_date or 'N/A' }}</span>
            </div>
            <div class="info-row">
                <span class="info-label">Folder:</span>
                <span class="info-value">{{ item.folder_link or 'N/A' }}</span>
            </div>
        </div>
        
        <!-- Reviewer's Response Section -->
        <div class="reviewer-response-box">
            <h3>📝 Initial Reviewer's Submitted Response</h3>
            <div class="info-row">
                <span class="info-label">Category:</span>
                <span class="info-value" style="font-weight: 600; color: #166534;">{{ item.reviewer_response_category or 'Not specified' }}</span>
            </div>
            <div class="info-row">
                <span class="info-label">Selected Files:</span>
                <span class="info-value">{{ reviewer_files|join('; ') if reviewer_files else 'None selected' }}</span>
            </div>
            {% if item.reviewer_notes %}
            <div class="info-row" style="flex-direction: column; align-items: flex-start;">
                <span class="info-label">Reviewer's Description:</span>
                <div class="response-text-readonly" style="margin-top: 8px; width: 100%;">{{ item.reviewer_notes }}</div>
            </div>
            {% endif %}
            {% if item.reviewer_internal_notes %}
            <div style="margin-top: 12px; padding: 10px; background: #fff8e6; border: 1px solid #ffd966; border-radius: 6px;">
                <span class="info-label" style="color: #b7791f;">🔒 Reviewer's Internal Notes (Team Only):</span>
                <div style="margin-top: 6px; color: #744210; white-space: pre-wrap;">{{ item.reviewer_internal_notes }}</div>
            </div>
            {% endif %}
            {% if item.reviewer_response_text %}
            <div class="info-row" style="flex-direction: column; align-items: flex-start;">
                <span class="info-label">Reviewer's Response Text:</span>
                <div class="response-text-readonly" id="reviewer_original_text" style="margin-top: 8px; width: 100%;">{{ item.reviewer_response_text }}</div>
            </div>
            {% endif %}
        </div>
        
        <form method="POST" id="qcr-form">
            <input type="hidden" name="token" value="{{ token }}">
            
            <!-- QC Action Selection -->
            <div class="action-group">
                <h3>🎯 Your QC Decision</h3>
                <div class="radio-group">
                    <label class="radio-option">
                        <input type="radio" name="qc_action" value="Approve" required>
                        <div class="radio-option-content">
                            <div class="radio-option-label">✅ Approve</div>
                            <div class="radio-option-desc">Accept the reviewer's response as final. No changes needed.</div>
                        </div>
                    </label>
                    <label class="radio-option">
                        <input type="radio" name="qc_action" value="Modify" required>
                        <div class="radio-option-content">
                            <div class="radio-option-label">✏️ Modify</div>
                            <div class="radio-option-desc">Make adjustments to the response. You can tweak or revise the text.</div>
                        </div>
                    </label>
                    <label class="radio-option">
                        <input type="radio" name="qc_action" value="Send Back" required>
                        <div class="radio-option-content">
                            <div class="radio-option-label">↩️ Send Back to Reviewer</div>
                            <div class="radio-option-desc">Return to the reviewer for revisions. You'll specify what needs to change.</div>
                        </div>
                    </label>
                </div>
            </div>
            
            <div class="send-back-warning" id="send-back-warning">
                ⚠️ <strong>Sending Back:</strong> The item will be returned to the Initial Reviewer for revision. They will receive an email with your notes explaining what changes are needed.
            </div>
            
            <!-- Response Text Handling (shown only for Approve/Modify) -->
            <div class="form-group" id="response-mode-group" style="display: none;">
                <label style="font-weight: 600; margin-bottom: 12px; display: block;">📄 Response Text Handling</label>
                <div class="radio-group">
                    <label class="radio-option">
                        <input type="radio" name="response_mode" value="Keep" checked>
                        <div class="radio-option-content">
                            <div class="radio-option-label">Keep as is</div>
                            <div class="radio-option-desc">Use the reviewer's original response text without changes.</div>
                        </div>
                    </label>
                    <label class="radio-option">
                        <input type="radio" name="response_mode" value="Tweak">
                        <div class="radio-option-content">
                            <div class="radio-option-label">Tweak</div>
                            <div class="radio-option-desc">Start with the reviewer's text and make minor edits.</div>
                        </div>
                    </label>
                    <label class="radio-option">
                        <input type="radio" name="response_mode" value="Revise">
                        <div class="radio-option-content">
                            <div class="radio-option-label">Revise</div>
                            <div class="radio-option-desc">Write a completely new response from scratch.</div>
                        </div>
                    </label>
                </div>
                
                <div class="response-text-container" id="response-text-container">
                    <div class="response-text-label">Description (Final Response Text): <span id="notes-required-hint" style="display: none; color: #dc2626;">* Required for Send Back</span></div>
                    <div class="response-text-readonly" id="response_text_readonly">{{ item.reviewer_response_text or item.reviewer_notes or '' }}</div>
                    <textarea name="response_text" id="response_text_area" placeholder="Enter your response text...">{{ item.reviewer_response_text or item.reviewer_notes or '' }}</textarea>
                </div>
            </div>
            
            <!-- Final Response Category (shown only for Approve/Modify) -->
            <div class="form-group" id="category-group" style="display: none;">
                <label for="response_category">Final Response Category *</label>
                <select name="response_category" id="response_category">
                    <option value="">-- Select --</option>
                    <option value="Approved" {% if item.reviewer_response_category == 'Approved' %}selected{% endif %}>Approved</option>
                    <option value="Approved as Noted" {% if item.reviewer_response_category == 'Approved as Noted' %}selected{% endif %}>Approved as Noted</option>
                    <option value="For Record Only" {% if item.reviewer_response_category == 'For Record Only' %}selected{% endif %}>For Record Only</option>
                    <option value="Rejected" {% if item.reviewer_response_category == 'Rejected' %}selected{% endif %}>Rejected</option>
                    <option value="Revise and Resubmit" {% if item.reviewer_response_category == 'Revise and Resubmit' %}selected{% endif %}>Revise and Resubmit</option>
                </select>
            </div>
            
            <!-- File Selection (shown only for Approve/Modify) -->
            <div class="form-group" id="files-group" style="display: none;">
                <p class="section-title">Confirm Files to Include in Response</p>
                <p style="color: #666; font-size: 13px; margin-bottom: 8px;">Files selected by the Initial Reviewer are pre-checked. You can adjust as needed.</p>
                <div class="file-list">
                    {% if files %}
                        {% for file in files %}
                        <div class="file-item">
                            <input type="checkbox" name="selected_files" value="{{ file }}" id="file_{{ loop.index }}"
                                {% if file in reviewer_files %}checked{% endif %}>
                            <label for="file_{{ loop.index }}">{{ file }}</label>
                        </div>
                        {% endfor %}
                    {% else %}
                        <p style="color: #666; padding: 8px;">No files found in folder.</p>
                    {% endif %}
                </div>
            </div>
            
            <!-- QC Internal Notes (always shown) -->
            <div class="form-group">
                <label for="qcr_internal_notes">Internal Notes <span style="font-weight: normal; color: #888;">(not shared externally)</span></label>
                <textarea name="qcr_internal_notes" id="qcr_internal_notes" placeholder="Add any internal notes for project records..."></textarea>
            </div>
            
            <button type="submit" class="btn" id="submit-btn">Complete QC Review</button>
        </form>
    </div>
    
    <script>
        const reviewerText = document.getElementById('reviewer_original_text')?.innerText || '{{ (item.reviewer_response_text or item.reviewer_notes or "")|e }}';
        const responseModeGroup = document.getElementById('response-mode-group');
        const categoryGroup = document.getElementById('category-group');
        const filesGroup = document.getElementById('files-group');
        const sendBackWarning = document.getElementById('send-back-warning');
        const notesRequiredHint = document.getElementById('notes-required-hint');
        const responseTextReadonly = document.getElementById('response_text_readonly');
        const responseTextArea = document.getElementById('response_text_area');
        const submitBtn = document.getElementById('submit-btn');
        const categorySelect = document.getElementById('response_category');
        
        // Handle QC Action change
        document.querySelectorAll('input[name="qc_action"]').forEach(radio => {
            radio.addEventListener('change', function() {
                const action = this.value;
                
                if (action === 'Send Back') {
                    responseModeGroup.style.display = 'none';
                    categoryGroup.style.display = 'none';
                    filesGroup.style.display = 'none';
                    sendBackWarning.style.display = 'block';
                    notesRequiredHint.style.display = 'inline';
                    categorySelect.required = false;
                    submitBtn.textContent = '↩️ Send Back to Reviewer';
                    submitBtn.style.background = '#f59e0b';
                    // Show textarea for Send Back explanation
                    responseTextReadonly.style.display = 'none';
                    responseTextArea.style.display = 'block';
                    responseTextArea.value = '';
                    responseTextArea.placeholder = 'Explain what revisions are needed...';
                } else {
                    responseModeGroup.style.display = 'block';
                    categoryGroup.style.display = 'block';
                    filesGroup.style.display = 'block';
                    sendBackWarning.style.display = 'none';
                    notesRequiredHint.style.display = 'none';
                    categorySelect.required = true;
                    // Reset to Keep mode display
                    responseTextReadonly.style.display = 'block';
                    responseTextArea.style.display = 'none';
                    responseTextArea.value = reviewerText;
                    
                    if (action === 'Approve') {
                        submitBtn.textContent = '✅ Approve & Complete';
                        submitBtn.style.background = '#10b981';
                    } else {
                        submitBtn.textContent = '✏️ Submit Modifications';
                        submitBtn.style.background = 'linear-gradient(135deg, #667eea 0%, #764ba2 100%)';
                    }
                }
            });
        });
        
        // Handle Response Mode change
        document.querySelectorAll('input[name="response_mode"]').forEach(radio => {
            radio.addEventListener('change', function() {
                const mode = this.value;
                
                if (mode === 'Keep') {
                    responseTextReadonly.style.display = 'block';
                    responseTextArea.style.display = 'none';
                    responseTextArea.value = reviewerText;
                } else if (mode === 'Tweak') {
                    responseTextReadonly.style.display = 'none';
                    responseTextArea.style.display = 'block';
                    responseTextArea.value = reviewerText;
                } else if (mode === 'Revise') {
                    responseTextReadonly.style.display = 'none';
                    responseTextArea.style.display = 'block';
                    responseTextArea.value = '';
                    responseTextArea.placeholder = 'Write your new response text here...';
                }
            });
        });
        
        // Form validation
        document.getElementById('qcr-form').addEventListener('submit', function(e) {
            const action = document.querySelector('input[name="qc_action"]:checked')?.value;
            
            if (!action) {
                e.preventDefault();
                alert('Please select a QC decision.');
                return false;
            }
            
            if (action === 'Send Back' && !responseTextArea.value.trim()) {
                e.preventDefault();
                alert('Please provide a description explaining what revisions are needed.');
                return false;
            }
            
            if ((action === 'Approve' || action === 'Modify') && !categorySelect.value) {
                e.preventDefault();
                alert('Please select a final response category.');
                return false;
            }
        });
    </script>
</body>
</html>
'''

# =============================================================================
# MULTI-REVIEWER FORM TEMPLATES
# =============================================================================

MULTI_REVIEWER_RESPONSE_TEMPLATE = '''
<!DOCTYPE html>
<html>
<head>
    <title>Review Response - {{ item.type }} {{ item.identifier }}</title>
    ''' + BASE_PAGE_STYLE + '''
    <style>
        .version-badge {
            display: inline-block;
            background: #667eea;
            color: white;
            padding: 4px 12px;
            border-radius: 20px;
            font-size: 12px;
            font-weight: 600;
            margin-left: 10px;
        }
        .reviewer-badge {
            display: inline-block;
            background: #10b981;
            color: white;
            padding: 4px 12px;
            border-radius: 20px;
            font-size: 12px;
            font-weight: 600;
        }
        .bluebeam-notice {
            background: #dbeafe;
            border: 1px solid #3b82f6;
            border-radius: 8px;
            padding: 16px;
            margin-bottom: 24px;
        }
        .bluebeam-notice h3 {
            margin: 0 0 8px 0;
            color: #1e40af;
            display: flex;
            align-items: center;
            gap: 8px;
        }
        .bluebeam-notice p {
            color: #1e40af;
            margin: 0;
        }
        .qcr-feedback {
            background: #fef2f2;
            border: 1px solid #fecaca;
            border-radius: 8px;
            padding: 15px;
            margin-bottom: 20px;
        }
        .qcr-feedback h4 {
            margin: 0 0 10px 0;
            color: #991b1b;
        }
        .previous-response {
            background: #f0f9ff;
            border: 1px solid #bae6fd;
            border-radius: 8px;
            padding: 15px;
            margin-bottom: 20px;
        }
        .previous-response h4 {
            margin: 0 0 10px 0;
            color: #0369a1;
        }
        .waiting-notice {
            background: #fef3c7;
            border: 1px solid #f59e0b;
            border-radius: 8px;
            padding: 16px;
            margin-bottom: 24px;
        }
        .waiting-notice h3 {
            margin: 0 0 8px 0;
            color: #92400e;
        }
        .closed-notice {
            background: #f3f4f6;
            border: 1px solid #d1d5db;
            border-radius: 8px;
            padding: 20px;
            text-align: center;
        }
        .closed-notice h3 {
            color: #6b7280;
            margin: 0 0 10px 0;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>Initial Review Response <span class="version-badge">v{{ version }}</span></h1>
        <p class="subtitle">{{ item.type }} {{ item.identifier }}</p>
        <p><span class="reviewer-badge">👤 {{ reviewer_name }}</span></p>
        
        {% if is_closed %}
        <div class="closed-notice">
            <h3>🔒 This Item Has Been Closed</h3>
            <p>No further changes can be submitted. Contact the project administrator if this is unexpected.</p>
        </div>
        {% elif not can_submit %}
        <div class="closed-notice">
            <h3>⚠️ Submission Not Allowed</h3>
            <p>This item has been finalized in QC. Contact the project administrator if additional changes are required.</p>
        </div>
        {% else %}
        
        {% if is_resubmit and qcr_feedback %}
        <div class="qcr-feedback">
            <h4>↩️ QC Reviewer Requested Revisions</h4>
            <p><strong>Feedback:</strong></p>
            <div style="background: white; padding: 10px; border-radius: 4px; margin-top: 8px;">
                {{ qcr_feedback|replace('\\n', '<br>')|safe }}
            </div>
        </div>
        {% endif %}
        
        {% if is_resubmit and previous_response %}
        <div class="previous-response">
            <h4>📄 Your Previous Response (v{{ version - 1 }})</h4>
            <p><strong>Category:</strong> {{ previous_response.category or 'N/A' }}</p>
            {% if previous_response.notes %}
            <p><strong>Internal Notes:</strong></p>
            <div style="background: white; padding: 10px; border-radius: 4px; margin-top: 8px;">
                {{ previous_response.notes|replace('\\n', '<br>')|safe }}
            </div>
            {% endif %}
        </div>
        {% endif %}
        
        <div class="info-box">
            <div class="info-row">
                <span class="info-label">Title:</span>
                <span class="info-value">{{ item.title or 'N/A' }}</span>
            </div>
            <div class="info-row">
                <span class="info-label">Date Received:</span>
                <span class="info-value">{{ item.date_received or 'N/A' }}</span>
            </div>
            <div class="info-row">
                <span class="info-label">Review Due Date:</span>
                <span class="info-value">{{ item.initial_reviewer_due_date or 'N/A' }}</span>
            </div>
            <div class="info-row">
                <span class="info-label">Folder:</span>
                <span class="info-value">{{ item.folder_link or 'N/A' }}</span>
            </div>
        </div>
        
        <!-- Bluebeam Notice instead of file selection -->
        <div class="bluebeam-notice">
            <h3>📐 Markups Instructions</h3>
            <p><strong>Provide markups in the corresponding Bluebeam session.</strong></p>
            <p style="margin-top: 8px; font-size: 13px;">Do not attach files here. All markups should be completed in the shared Bluebeam Studio session for this item.</p>
        </div>
        
        <form method="POST">
            <input type="hidden" name="token" value="{{ token }}">
            
            <div class="form-group">
                <label for="response_category">Response Category *</label>
                <select name="response_category" id="response_category" required>
                    <option value="">-- Select --</option>
                    <option value="Approved" {% if previous_response and previous_response.category == 'Approved' %}selected{% endif %}>Approved</option>
                    <option value="Approved as Noted" {% if previous_response and previous_response.category == 'Approved as Noted' %}selected{% endif %}>Approved as Noted</option>
                    <option value="For Record Only" {% if previous_response and previous_response.category == 'For Record Only' %}selected{% endif %}>For Record Only</option>
                    <option value="Rejected" {% if previous_response and previous_response.category == 'Rejected' %}selected{% endif %}>Rejected</option>
                    <option value="Revise and Resubmit" {% if previous_response and previous_response.category == 'Revise and Resubmit' %}selected{% endif %}>Revise and Resubmit</option>
                </select>
            </div>
            
            <div class="form-group">
                <label for="internal_notes">Internal Notes for QC Reviewer <span style="font-weight: normal; color: #888;">(team only - not shared externally)</span></label>
                <textarea name="internal_notes" id="internal_notes" placeholder="Add any notes, concerns, or recommendations for the QC Reviewer...">{{ previous_response.notes if previous_response else '' }}</textarea>
            </div>
            
            <button type="submit" class="btn">{% if is_resubmit %}Submit Revision (v{{ version }}){% else %}Submit Review{% endif %}</button>
        </form>
        
        {% if pending_reviewers %}
        <div class="waiting-notice" style="margin-top: 24px;">
            <h3>⏳ Other Reviewers</h3>
            <p>This item is assigned to multiple reviewers. The QC Reviewer will be notified once all reviewers submit.</p>
            <ul style="margin-top: 8px; color: #92400e;">
                {% for r in all_reviewers %}
                <li>{{ r.reviewer_name }} - {% if r.response_at %}✅ Submitted{% else %}⏳ Pending{% endif %}</li>
                {% endfor %}
            </ul>
        </div>
        {% endif %}
        {% endif %}
    </div>
</body>
</html>
'''

MULTI_REVIEWER_QCR_TEMPLATE = '''
<!DOCTYPE html>
<html>
<head>
    <title>QC Review - {{ item.type }} {{ item.identifier }}</title>
    ''' + BASE_PAGE_STYLE + '''
    <style>
        .reviewer-response-box {
            background: #f0fdf4;
            border: 1px solid #86efac;
            border-radius: 8px;
            padding: 15px;
            margin: 15px 0;
        }
        .reviewer-response-box h4 {
            margin: 0 0 12px 0;
            color: #166534;
            display: flex;
            align-items: center;
            gap: 8px;
        }
        .reviewer-badge {
            display: inline-block;
            background: #10b981;
            color: white;
            padding: 2px 8px;
            border-radius: 12px;
            font-size: 11px;
            font-weight: 600;
        }
        .category-chip {
            display: inline-block;
            background: #e0e7ff;
            color: #3730a3;
            padding: 4px 10px;
            border-radius: 12px;
            font-size: 12px;
            font-weight: 600;
        }
        .bluebeam-notice {
            background: #dbeafe;
            border: 1px solid #3b82f6;
            border-radius: 8px;
            padding: 16px;
            margin-bottom: 24px;
        }
        .bluebeam-notice h3 {
            margin: 0 0 8px 0;
            color: #1e40af;
        }
        .action-group {
            background: #fef3c7;
            border: 1px solid #fbbf24;
            border-radius: 8px;
            padding: 15px;
            margin: 20px 0;
        }
        .action-group h3 {
            margin: 0 0 12px 0;
            color: #92400e;
        }
        .radio-group {
            display: flex;
            flex-direction: column;
            gap: 10px;
        }
        .radio-option {
            display: flex;
            align-items: flex-start;
            gap: 10px;
            padding: 10px;
            background: white;
            border-radius: 6px;
            border: 1px solid #e5e7eb;
            cursor: pointer;
        }
        .radio-option:hover {
            border-color: #667eea;
        }
        .radio-option input[type="radio"] {
            margin-top: 3px;
        }
        .radio-option-content {
            flex: 1;
        }
        .radio-option-label {
            font-weight: 600;
            color: #333;
        }
        .radio-option-desc {
            font-size: 12px;
            color: #666;
            margin-top: 4px;
        }
        .send-back-warning {
            display: none;
            background: #fef2f2;
            border: 1px solid #fecaca;
            border-radius: 8px;
            padding: 12px;
            margin-top: 15px;
            color: #991b1b;
        }
        .internal-notes-box {
            background: #fff8e6;
            border: 1px solid #ffd966;
            border-radius: 6px;
            padding: 10px;
            margin-top: 10px;
        }
        .internal-notes-box h5 {
            margin: 0 0 6px 0;
            color: #744210;
            font-size: 12px;
        }
        .internal-notes-content {
            color: #744210;
            font-size: 13px;
            white-space: pre-wrap;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>QC Review</h1>
        <p class="subtitle">{{ item.type }} {{ item.identifier }}</p>
        
        <div class="info-box">
            <div class="info-row">
                <span class="info-label">Title:</span>
                <span class="info-value">{{ item.title or 'N/A' }}</span>
            </div>
            <div class="info-row">
                <span class="info-label">Date Received:</span>
                <span class="info-value">{{ item.date_received or 'N/A' }}</span>
            </div>
            <div class="info-row">
                <span class="info-label">Priority:</span>
                <span class="info-value">{{ item.priority or 'Normal' }}</span>
            </div>
            <div class="info-row">
                <span class="info-label">QC Due Date:</span>
                <span class="info-value" style="color: #d97706;">{{ item.qcr_due_date or 'N/A' }}</span>
            </div>
            <div class="info-row">
                <span class="info-label">Contractor Due Date:</span>
                <span class="info-value">{{ item.due_date or 'N/A' }}</span>
            </div>
            <div class="info-row">
                <span class="info-label">Folder:</span>
                <span class="info-value">{{ item.folder_link or 'N/A' }}</span>
            </div>
        </div>
        
        <!-- Bluebeam Notice -->
        <div class="bluebeam-notice">
            <h3>📐 Markups Location</h3>
            <p>All reviewer markups are in the corresponding Bluebeam Studio session for this item.</p>
        </div>
        
        <!-- All Reviewer Responses -->
        <h3 style="margin: 20px 0 10px 0;">📝 Reviewer Responses ({{ reviewer_responses|length }})</h3>
        {% for reviewer in reviewer_responses %}
        <div class="reviewer-response-box">
            <h4>
                <span class="reviewer-badge">{{ loop.index }}</span>
                {{ reviewer.reviewer_name }}
                <span class="category-chip">{{ reviewer.response_category or 'Pending' }}</span>
            </h4>
            {% if reviewer.internal_notes %}
            <div class="internal-notes-box">
                <h5>🔒 Internal Notes (Team Only):</h5>
                <div class="internal-notes-content">{{ reviewer.internal_notes }}</div>
            </div>
            {% else %}
            <p style="color: #666; font-size: 13px;">No internal notes provided.</p>
            {% endif %}
        </div>
        {% endfor %}
        
        <form method="POST" id="qcr-form">
            <input type="hidden" name="token" value="{{ token }}">
            
            <!-- QC Action Selection -->
            <div class="action-group">
                <h3>🎯 Your QC Decision</h3>
                <div class="radio-group">
                    <label class="radio-option">
                        <input type="radio" name="qc_action" value="Complete" required>
                        <div class="radio-option-content">
                            <div class="radio-option-label">✅ Complete Response</div>
                            <div class="radio-option-desc">Write the final response to be sent to the contractor. You'll select a category and provide the official response text.</div>
                        </div>
                    </label>
                    <label class="radio-option">
                        <input type="radio" name="qc_action" value="Send Back" required>
                        <div class="radio-option-content">
                            <div class="radio-option-label">↩️ Send Back to All Reviewers</div>
                            <div class="radio-option-desc">Return to all reviewers for revisions. They will all receive an email with your feedback.</div>
                        </div>
                    </label>
                </div>
            </div>
            
            <div class="send-back-warning" id="send-back-warning">
                ⚠️ <strong>Sending Back:</strong> The item will be returned to ALL Initial Reviewers for revision. Each reviewer will receive an email with your notes explaining what changes are needed.
            </div>
            
            <!-- Response fields (shown only for Complete) -->
            <div id="complete-fields" style="display: none;">
                <div class="form-group">
                    <label for="response_category">Final Response Category *</label>
                    <select name="response_category" id="response_category">
                        <option value="">-- Select --</option>
                        <option value="Approved">Approved</option>
                        <option value="Approved as Noted">Approved as Noted</option>
                        <option value="For Record Only">For Record Only</option>
                        <option value="Rejected">Rejected</option>
                        <option value="Revise and Resubmit">Revise and Resubmit</option>
                    </select>
                </div>
                
                <div class="form-group">
                    <label for="response_text">Final Response Description *</label>
                    <textarea name="response_text" id="response_text" placeholder="Write the official response text to be sent to the contractor..." style="min-height: 150px;"></textarea>
                </div>
            </div>
            
            <!-- Send back notes (shown only for Send Back) -->
            <div id="sendback-fields" style="display: none;">
                <div class="form-group">
                    <label for="sendback_notes">Feedback for Reviewers * <span style="font-weight: normal; color: #888;">(will be sent to all reviewers)</span></label>
                    <textarea name="sendback_notes" id="sendback_notes" placeholder="Explain what revisions are needed from the reviewers..." style="min-height: 120px;"></textarea>
                </div>
            </div>
            
            <!-- QC Internal Notes (always shown) -->
            <div class="form-group">
                <label for="qcr_internal_notes">QC Internal Notes <span style="font-weight: normal; color: #888;">(not shared externally)</span></label>
                <textarea name="qcr_internal_notes" id="qcr_internal_notes" placeholder="Add any internal notes for project records..."></textarea>
            </div>
            
            <button type="submit" class="btn" id="submit-btn">Complete QC Review</button>
        </form>
    </div>
    
    <script>
        const completeFields = document.getElementById('complete-fields');
        const sendbackFields = document.getElementById('sendback-fields');
        const sendBackWarning = document.getElementById('send-back-warning');
        const submitBtn = document.getElementById('submit-btn');
        const categorySelect = document.getElementById('response_category');
        const responseText = document.getElementById('response_text');
        const sendbackNotes = document.getElementById('sendback_notes');
        
        // Handle QC Action change
        document.querySelectorAll('input[name="qc_action"]').forEach(radio => {
            radio.addEventListener('change', function() {
                const action = this.value;
                
                if (action === 'Send Back') {
                    completeFields.style.display = 'none';
                    sendbackFields.style.display = 'block';
                    sendBackWarning.style.display = 'block';
                    categorySelect.required = false;
                    responseText.required = false;
                    sendbackNotes.required = true;
                    submitBtn.textContent = '↩️ Send Back to All Reviewers';
                    submitBtn.style.background = '#f59e0b';
                } else {
                    completeFields.style.display = 'block';
                    sendbackFields.style.display = 'none';
                    sendBackWarning.style.display = 'none';
                    categorySelect.required = true;
                    responseText.required = true;
                    sendbackNotes.required = false;
                    submitBtn.textContent = '✅ Submit Final Response';
                    submitBtn.style.background = '#10b981';
                }
            });
        });
        
        // Form validation
        document.getElementById('qcr-form').addEventListener('submit', function(e) {
            const action = document.querySelector('input[name="qc_action"]:checked')?.value;
            
            if (!action) {
                e.preventDefault();
                alert('Please select a QC decision.');
                return false;
            }
            
            if (action === 'Send Back' && !sendbackNotes.value.trim()) {
                e.preventDefault();
                alert('Please provide feedback explaining what revisions are needed.');
                return false;
            }
            
            if (action === 'Complete') {
                if (!categorySelect.value) {
                    e.preventDefault();
                    alert('Please select a final response category.');
                    return false;
                }
                if (!responseText.value.trim()) {
                    e.preventDefault();
                    alert('Please provide the final response description.');
                    return false;
                }
            }
        });
    </script>
</body>
</html>
'''

# =============================================================================
# DATABASE INITIALIZATION
# =============================================================================

def get_db():
    """Get database connection with row factory."""
    conn = sqlite3.connect(str(DATABASE_PATH), check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    """Initialize database tables."""
    conn = get_db()
    cursor = conn.cursor()
    
    # User table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS user (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT UNIQUE NOT NULL,
            password_hash TEXT,
            display_name TEXT,
            role TEXT DEFAULT 'user' CHECK(role IN ('admin', 'user')),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Item table (RFI/Submittal)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS item (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            type TEXT NOT NULL CHECK(type IN ('RFI', 'Submittal')),
            bucket TEXT NOT NULL DEFAULT 'ALL',
            identifier TEXT NOT NULL,
            title TEXT,
            source_subject TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            last_email_at TIMESTAMP,
            due_date DATE,
            priority TEXT CHECK(priority IN ('High', 'Medium', 'Low', NULL)),
            status TEXT DEFAULT 'Unassigned' CHECK(status IN ('Unassigned', 'Assigned', 'In Review', 'In QC', 'Ready for Response', 'Closed')),
            assigned_to_user_id INTEGER,
            notes TEXT,
            folder_link TEXT,
            response_category TEXT CHECK(response_category IN ('Approved', 'Approved as Noted', 'For Record Only', 'Rejected', 'Revise and Resubmit', NULL)),
            response_text TEXT,
            response_files TEXT,
            closed_at TIMESTAMP,
            read_by TEXT,
            FOREIGN KEY (assigned_to_user_id) REFERENCES user(id),
            UNIQUE(identifier, bucket)
        )
    ''')
    
    # Comment table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS comment (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item_id INTEGER NOT NULL,
            user_id INTEGER NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            body TEXT NOT NULL,
            FOREIGN KEY (item_id) REFERENCES item(id) ON DELETE CASCADE,
            FOREIGN KEY (user_id) REFERENCES user(id)
        )
    ''')
    
    # Email log table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS email_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item_id INTEGER,
            message_id TEXT UNIQUE,
            entry_id TEXT,
            subject TEXT,
            body_preview TEXT,
            received_at TIMESTAMP,
            raw_type TEXT,
            processed INTEGER DEFAULT 0,
            FOREIGN KEY (item_id) REFERENCES item(id)
        )
    ''')
    
    # Migration: Add new columns to existing databases
    try:
        cursor.execute('ALTER TABLE item ADD COLUMN response_category TEXT')
    except:
        pass
    try:
        cursor.execute('ALTER TABLE item ADD COLUMN response_text TEXT')
    except:
        pass
    try:
        cursor.execute('ALTER TABLE item ADD COLUMN response_files TEXT')
    except:
        pass
    try:
        cursor.execute('ALTER TABLE item ADD COLUMN closed_at TIMESTAMP')
    except:
        pass
    try:
        cursor.execute('ALTER TABLE item ADD COLUMN read_by TEXT')
    except:
        pass
    # New columns for two-level review system
    try:
        cursor.execute('ALTER TABLE item ADD COLUMN date_received DATE')
    except:
        pass
    try:
        cursor.execute('ALTER TABLE item ADD COLUMN initial_reviewer_id INTEGER REFERENCES user(id)')
    except:
        pass
    try:
        cursor.execute('ALTER TABLE item ADD COLUMN qcr_id INTEGER REFERENCES user(id)')
    except:
        pass
    try:
        cursor.execute('ALTER TABLE item ADD COLUMN initial_reviewer_due_date DATE')
    except:
        pass
    try:
        cursor.execute('ALTER TABLE item ADD COLUMN qcr_due_date DATE')
    except:
        pass
    try:
        cursor.execute('ALTER TABLE item ADD COLUMN is_contractor_window_insufficient INTEGER DEFAULT 0')
    except:
        pass
    
    # Backfill date_received for existing items that don't have it
    cursor.execute('''
        UPDATE item 
        SET date_received = DATE(created_at)
        WHERE date_received IS NULL AND created_at IS NOT NULL
    ''')
    
    # Recalculate review due dates for ALL items that have date_received and due_date
    # This ensures the calculation is always correct with the latest logic
    cursor.execute('''
        SELECT id, date_received, due_date, priority 
        FROM item 
        WHERE date_received IS NOT NULL 
        AND due_date IS NOT NULL
    ''')
    items_to_update = cursor.fetchall()
    for item_id, date_received, due_date, priority in items_to_update:
        try:
            due_dates = calculate_review_due_dates(date_received, due_date, priority)
            cursor.execute('''
                UPDATE item SET 
                    initial_reviewer_due_date = ?,
                    qcr_due_date = ?,
                    is_contractor_window_insufficient = ?
                WHERE id = ?
            ''', (
                due_dates['initial_reviewer_due_date'],
                due_dates['qcr_due_date'],
                1 if due_dates['is_contractor_window_insufficient'] else 0,
                item_id
            ))
        except Exception as e:
            print(f"Could not calculate due dates for item {item_id}: {e}")
    
    # Email workflow columns
    email_workflow_columns = [
        ('email_token_reviewer', 'TEXT'),
        ('email_token_qcr', 'TEXT'),
        ('reviewer_email_sent_at', 'TIMESTAMP'),
        ('reviewer_response_at', 'TIMESTAMP'),
        ('qcr_email_sent_at', 'TIMESTAMP'),
        ('qcr_response_at', 'TIMESTAMP'),
        ('reviewer_response_status', "TEXT DEFAULT 'Not Sent'"),
        ('qcr_response_status', "TEXT DEFAULT 'Not Sent'"),
        # Reviewer response fields
        ('reviewer_notes', 'TEXT'),  # Description (external)
        ('reviewer_internal_notes', 'TEXT'),  # Internal notes (not shared externally)
        ('reviewer_response_category', 'TEXT'),
        ('reviewer_selected_files', 'TEXT'),
        ('reviewer_response_text', 'TEXT'),
        # QCR response fields
        ('qcr_notes', 'TEXT'),  # Description (external)
        ('qcr_internal_notes', 'TEXT'),  # Internal notes (not shared externally)
        ('qcr_response_category', 'TEXT'),
        ('qcr_selected_files', 'TEXT'),
        ('qcr_action', 'TEXT'),  # Approve, Modify, Send Back
        ('qcr_response_mode', 'TEXT'),  # Keep, Tweak, Revise
        ('qcr_response_text', 'TEXT'),
        # Final official response fields
        ('final_response_category', 'TEXT'),
        ('final_response_text', 'TEXT'),
        ('final_response_files', 'TEXT'),
    ]
    for col_name, col_type in email_workflow_columns:
        try:
            cursor.execute(f'ALTER TABLE item ADD COLUMN {col_name} {col_type}')
        except:
            pass
    
    # Add reviewer_response_version column for version tracking
    # First submission is v0, revisions are v1, v2, etc.
    try:
        cursor.execute('ALTER TABLE item ADD COLUMN reviewer_response_version INTEGER DEFAULT 0')
    except:
        pass
    
    # Add email_entry_id column for opening original email
    try:
        cursor.execute('ALTER TABLE item ADD COLUMN email_entry_id TEXT')
    except:
        pass
    
    # Reviewer response history table for version tracking
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS reviewer_response_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item_id INTEGER NOT NULL,
            version INTEGER NOT NULL,
            submitted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            response_category TEXT,
            response_text TEXT,
            response_files TEXT,
            submitted_by_user_id INTEGER,
            FOREIGN KEY (item_id) REFERENCES item(id) ON DELETE CASCADE,
            FOREIGN KEY (submitted_by_user_id) REFERENCES user(id)
        )
    ''')
    
    # Add missing columns to reviewer_response_history if they don't exist
    try:
        cursor.execute('ALTER TABLE reviewer_response_history ADD COLUMN notes TEXT')
    except:
        pass
    try:
        cursor.execute('ALTER TABLE reviewer_response_history ADD COLUMN selected_files TEXT')
    except:
        pass
    
    # ==========================================================================
    # MULTI-REVIEWER SUPPORT - item_reviewers table
    # ==========================================================================
    # This table tracks multiple reviewers assigned to a single item
    # Each reviewer gets their own form and must submit before QCR is notified
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS item_reviewers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item_id INTEGER NOT NULL,
            user_id INTEGER,
            reviewer_name TEXT NOT NULL,
            reviewer_email TEXT NOT NULL,
            email_token TEXT,
            email_sent_at TIMESTAMP,
            response_at TIMESTAMP,
            response_category TEXT,
            internal_notes TEXT,
            response_version INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (item_id) REFERENCES item(id) ON DELETE CASCADE,
            FOREIGN KEY (user_id) REFERENCES user(id)
        )
    ''')
    
    # Add multi_reviewer_mode column to item table
    try:
        cursor.execute('ALTER TABLE item ADD COLUMN multi_reviewer_mode INTEGER DEFAULT 0')
    except:
        pass
    
    # Add needs_response column to item_reviewers table (for selective send-back)
    try:
        cursor.execute('ALTER TABLE item_reviewers ADD COLUMN needs_response INTEGER DEFAULT 1')
    except:
        pass
    
    # Add rfi_question column to item table
    try:
        cursor.execute('ALTER TABLE item ADD COLUMN rfi_question TEXT')
    except:
        pass
    
    # Notification table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS notification (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            type TEXT NOT NULL,
            title TEXT NOT NULL,
            message TEXT,
            item_id INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            read_at TIMESTAMP,
            action_url TEXT,
            action_label TEXT,
            FOREIGN KEY (item_id) REFERENCES item(id) ON DELETE CASCADE
        )
    ''')
    
    # Reminder tracking table - tracks which reminders have been sent
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS reminder_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item_id INTEGER NOT NULL,
            reminder_type TEXT NOT NULL,
            recipient_email TEXT NOT NULL,
            recipient_role TEXT NOT NULL,
            due_date DATE NOT NULL,
            reminder_stage TEXT NOT NULL CHECK(reminder_stage IN ('due_today', 'overdue', 'manual')),
            sent_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (item_id) REFERENCES item(id) ON DELETE CASCADE,
            UNIQUE(item_id, recipient_email, recipient_role, reminder_stage)
        )
    ''')
    
    # Add item_reviewer_id column to reminder_log for multi-reviewer tracking
    try:
        cursor.execute('ALTER TABLE reminder_log ADD COLUMN item_reviewer_id INTEGER')
    except:
        pass
    
    # Migration: Update reminder_log CHECK constraint to allow 'manual' stage
    # SQLite doesn't support ALTER TABLE to modify constraints, so we recreate the table
    try:
        # Check if we need to migrate (old constraint doesn't allow 'manual')
        cursor.execute("SELECT sql FROM sqlite_master WHERE name = 'reminder_log'")
        table_sql = cursor.fetchone()
        if table_sql and "'manual'" not in table_sql[0]:
            # Need to migrate - recreate table with new constraint
            cursor.execute('ALTER TABLE reminder_log RENAME TO reminder_log_old')
            cursor.execute('''
                CREATE TABLE reminder_log (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    item_id INTEGER NOT NULL,
                    reminder_type TEXT NOT NULL,
                    recipient_email TEXT NOT NULL,
                    recipient_role TEXT NOT NULL,
                    due_date DATE NOT NULL,
                    reminder_stage TEXT NOT NULL CHECK(reminder_stage IN ('due_today', 'overdue', 'manual')),
                    sent_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    item_reviewer_id INTEGER,
                    FOREIGN KEY (item_id) REFERENCES item(id) ON DELETE CASCADE,
                    UNIQUE(item_id, recipient_email, recipient_role, reminder_stage)
                )
            ''')
            cursor.execute('''
                INSERT INTO reminder_log (id, item_id, reminder_type, recipient_email, recipient_role, due_date, reminder_stage, sent_at, item_reviewer_id)
                SELECT id, item_id, reminder_type, recipient_email, recipient_role, due_date, reminder_stage, sent_at, item_reviewer_id
                FROM reminder_log_old
            ''')
            cursor.execute('DROP TABLE reminder_log_old')
            print("Migrated reminder_log table to support 'manual' reminder stage")
    except Exception as e:
        print(f"Note: reminder_log migration skipped or failed: {e}")
    
    # ==========================================================================
    # CONTRACTOR UPDATE TRACKING - for handling ACC updates during/after review
    # ==========================================================================
    # Add columns to track contractor updates from ACC
    acc_update_columns = [
        ('has_pending_update', 'INTEGER DEFAULT 0'),  # Flag for admin review
        ('update_type', 'TEXT'),  # 'due_date_only' or 'content_change'
        ('update_detected_at', 'TIMESTAMP'),  # When the update was detected
        ('update_reviewed_at', 'TIMESTAMP'),  # When admin reviewed the update
        ('update_admin_note', 'TEXT'),  # Admin's note about the change
        ('previous_due_date', 'DATE'),  # Store old due date before update
        ('previous_title', 'TEXT'),  # Store old title before update
        ('previous_priority', 'TEXT'),  # Store old priority before update
        ('update_email_body', 'TEXT'),  # Store relevant portion of update email
        ('reopened_from_closed', 'INTEGER DEFAULT 0'),  # If item was reopened from Closed
        ('status_before_update', 'TEXT'),  # Status before the update came in
    ]
    for col_name, col_def in acc_update_columns:
        try:
            cursor.execute(f'ALTER TABLE item ADD COLUMN {col_name} {col_def}')
        except:
            pass
    
    # Item update history table - tracks all updates from ACC
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS item_update_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item_id INTEGER NOT NULL,
            detected_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            update_type TEXT NOT NULL,
            old_due_date DATE,
            new_due_date DATE,
            old_title TEXT,
            new_title TEXT,
            old_priority TEXT,
            new_priority TEXT,
            email_entry_id TEXT,
            email_subject TEXT,
            email_body_preview TEXT,
            admin_reviewed_at TIMESTAMP,
            admin_reviewed_by INTEGER,
            admin_note TEXT,
            action_taken TEXT,
            FOREIGN KEY (item_id) REFERENCES item(id) ON DELETE CASCADE,
            FOREIGN KEY (admin_reviewed_by) REFERENCES user(id)
        )
    ''')
    
    # Create default admin user if no users exist
    cursor.execute('SELECT COUNT(*) FROM user')
    if cursor.fetchone()[0] == 0:
        default_password = 'admin123'  # Change this!
        password_hash = bcrypt.hashpw(default_password.encode('utf-8'), bcrypt.gensalt())
        cursor.execute('''
            INSERT INTO user (email, password_hash, display_name, role)
            VALUES (?, ?, ?, ?)
        ''', ('admin@local', password_hash.decode('utf-8'), 'Administrator', 'admin'))
        print(f"Created default admin user: admin@local / {default_password}")
    
    conn.commit()
    conn.close()

# =============================================================================
# NOTIFICATION HELPERS
# =============================================================================

def show_windows_toast(title, message):
    """Show a Windows toast notification in the system tray."""
    if HAS_WINOTIFY:
        try:
            toast = Notification(
                app_id="LEB Tracker",
                title=title,
                msg=message,
                duration="short"
            )
            toast.set_audio(audio.Default, loop=False)
            # Add action to open the app
            toast.add_actions(label="Open App", launch="http://localhost:5000")
            toast.show()
        except Exception as e:
            print(f"Toast notification error: {e}")

def create_notification(notification_type, title, message, item_id=None, action_url=None, action_label=None):
    """Create a new notification and show Windows toast."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO notification (type, title, message, item_id, action_url, action_label)
        VALUES (?, ?, ?, ?, ?, ?)
    ''', (notification_type, title, message, item_id, action_url, action_label))
    conn.commit()
    notification_id = cursor.lastrowid
    conn.close()
    
    # Also show Windows toast notification
    show_windows_toast(title, message)
    
    return notification_id

# =============================================================================
# EMAIL PARSING UTILITIES
# =============================================================================

# Bucket mapping based on subject patterns
BUCKET_PATTERNS = [
    (r'LEB\s*-\s*Turner', 'ACC_TURNER'),
    (r'LEB\s*-\s*Mortenson', 'ACC_MORTENSON'),
    (r'LEB\s*-\s*Faith', 'ACC_FTI'),
    (r'LEB\s*-\s*FTI', 'ACC_FTI'),
    (r'LEB', 'ALL'),  # Default fallback
]

def determine_bucket(subject):
    """Determine the bucket from email subject."""
    for pattern, bucket in BUCKET_PATTERNS:
        if re.search(pattern, subject, re.IGNORECASE):
            return bucket
    return 'ALL'

def parse_item_type(subject):
    """Determine if this is an RFI or Submittal."""
    if re.search(r'\bSubmittal\b', subject, re.IGNORECASE):
        return 'Submittal'
    elif re.search(r'\bRFI\b', subject, re.IGNORECASE):
        return 'RFI'
    return None

def parse_identifier(subject, item_type):
    """Extract the identifier from the subject."""
    if item_type == 'Submittal':
        # Match patterns like "Submittal #13 34 19-2" or "Submittal #123"
        match = re.search(r'Submittal\s*#?([\d\s\-\.]+)', subject, re.IGNORECASE)
        if match:
            return f"Submittal #{match.group(1).strip()}"
    elif item_type == 'RFI':
        # Match patterns like "RFI #123" or "RFI-123"
        match = re.search(r'RFI\s*[#\-]?\s*(\d+)', subject, re.IGNORECASE)
        if match:
            return f"RFI #{match.group(1).strip()}"
    return None

def parse_title(subject, identifier, body=None):
    """Extract a title from the email body (full item name, NOT Spec Section)."""
    title = None
    
    if body:
        # First, try to get the FULL item title from ACC email body
        # ACC format in item link: "item #23 00 00-1 LEB1,2,10_230000_MOFE_Modular Central Utility Plant_Product Data & Drawings_Mech Yard"
        # We want the full title, NOT just the Spec Section
        
        # Extract just the number portion of the identifier (e.g., "23 00 00-1" from "Submittal #23 00 00-1")
        id_number = identifier
        if identifier:
            id_match = re.search(r'#(.+)$', identifier)
            if id_match:
                id_number = id_match.group(1).strip()
        
        if id_number:
            # Pattern 1: Look for "item #23 00 00-1 TITLE" in email body
            # Title starts after identifier and includes everything up to "What's changed" or newline
            item_pattern = rf'item\s*#?\s*{re.escape(id_number)}\s+([^\n\r]+?)(?:\s*What|\s*$)'
            item_match = re.search(item_pattern, body, re.IGNORECASE)
            if item_match:
                title = item_match.group(1).strip()
                # Remove any leading code prefix like "LEB1,2,10_230000_"
                title = re.sub(r'^[A-Z0-9,]+_\d+_', '', title)
                if title:
                    return title
        
        # Pattern 2: Look for "Title" field in ACC emails
        title_match = re.search(r'(?:^|\n)\s*Title[:\s\t]+([^\n\r]+)', body, re.IGNORECASE)
        if title_match:
            title = title_match.group(1).strip()
            if title and len(title) > 5:  # Make sure it's substantial
                return title
        
        # Pattern 3: Submittal/RFI with full title  
        if id_number:
            submittal_pattern = rf'(?:Submittal|RFI)\s*#?\s*{re.escape(id_number)}\s+([^\n\r]+?)(?:\s*What|\s*$)'
            submittal_match = re.search(submittal_pattern, body, re.IGNORECASE)
            if submittal_match:
                title = submittal_match.group(1).strip()
                title = re.sub(r'^[A-Z0-9,]+_\d+_', '', title)
                if title:
                    return title
        
        # Fallback: Use Spec Section if nothing else found
        spec_match = re.search(r'Spec\s*Section[:\s\t]+([^\n\r]+)', body, re.IGNORECASE)
        if spec_match:
            title = spec_match.group(1).strip()
            if title:
                return title
    
    # Fallback: extract from subject
    if identifier and subject:
        title = subject
        # Remove common prefixes
        title = re.sub(r'^(Re:\s*)?(Fwd?:\s*)?(Action Required:\s*)?', '', title, flags=re.IGNORECASE)
        # Remove project prefix like "LEB - Turner (NB.TypeF2.0) -"
        title = re.sub(r'LEB\s*-?\s*[\w\s]*\([^)]+\)\s*-?\s*', '', title)
        # Remove the identifier (handle both "Submittal #25 00 00-3" formats)
        title = re.sub(re.escape(identifier), '', title, flags=re.IGNORECASE)
        title = re.sub(r'Submittal\s*#?[\d\s\-\.]+', '', title, flags=re.IGNORECASE)
        title = re.sub(r'RFI\s*#?[\d\s\-\.]+', '', title, flags=re.IGNORECASE)
        # Remove trailing actions
        title = re.sub(r'\s*(was assigned to you|was assigned to your role|needs your review|requires action).*$', '', title, flags=re.IGNORECASE)
        title = title.strip(' -–—:,')
        
    return title if title else None

def parse_due_date(body):
    """Extract due date from email body."""
    if not body:
        return None
    
    # Patterns to look for (ACC uses tab/whitespace separation, not just colons)
    patterns = [
        r'Due\s*Date[:\s\t]+([A-Za-z]{3,9}\s+\d{1,2},?\s+\d{4})',  # "Due Date    Jan 22, 2026"
        r'Due\s*Date[:\s\t]+(\d{1,2}/\d{1,2}/\d{4})',  # "Due Date    01/22/2026"
        r'Due\s*Date[:\s\t]+(\d{4}-\d{2}-\d{2})',  # "Due Date    2026-01-22"
        r'Due[:\s\t]+([A-Za-z]{3,9}\s+\d{1,2},?\s+\d{4})',  # "Due: Jan 22, 2026"
        r'Due[:\s\t]+(\d{1,2}/\d{1,2}/\d{4})',
        r'Response\s*Due[:\s\t]+(.+?)(?:\n|$)',
        r'Required\s*[Bb]y[:\s\t]+(.+?)(?:\n|$)',
    ]
    
    for pattern in patterns:
        match = re.search(pattern, body, re.IGNORECASE)
        if match:
            date_str = match.group(1).strip()
            # Try to parse the date
            if HAS_DATEUTIL:
                try:
                    # Parse without timezone awareness to avoid UTC conversion issues
                    parsed_date = date_parser.parse(date_str, fuzzy=True, ignoretz=True)
                    # Return just the date portion (no time conversion)
                    return parsed_date.strftime('%Y-%m-%d')
                except:
                    pass
            else:
                # Basic date parsing without dateutil
                # Try common formats
                for fmt in ['%m/%d/%Y', '%Y-%m-%d', '%B %d, %Y', '%b %d, %Y', '%b %d %Y']:
                    try:
                        parsed_date = datetime.strptime(date_str, fmt)
                        return parsed_date.strftime('%Y-%m-%d')
                    except:
                        pass
    return None

def parse_priority(body):
    """Extract priority from email body."""
    if not body:
        return None
    
    # ACC uses tab/whitespace separation and may use "Normal" instead of "Medium"
    match = re.search(r'Priority[:\s\t]+(High|Medium|Normal|Low|Urgent|Critical)', body, re.IGNORECASE)
    if match:
        priority = match.group(1).capitalize()
        # Map ACC priority values to our standard values
        priority_map = {
            'Normal': 'Medium',
            'Urgent': 'High',
            'Critical': 'High'
        }
        return priority_map.get(priority, priority)
    return None

def sanitize_folder_name(name):
    """Create a filesystem-safe folder name."""
    # Remove or replace invalid characters
    invalid_chars = '<>:"/\\|?*'
    for char in invalid_chars:
        name = name.replace(char, '-')
    # Remove multiple dashes
    name = re.sub(r'-+', '-', name)
    # Remove leading/trailing dashes and spaces
    name = name.strip('- ')
    return name

def create_item_folder(item_type, identifier, bucket, title=None):
    """Create a folder for the item and return the path."""
    base_path = Path(CONFIG['base_folder_path'])
    
    # Create bucket subfolder
    bucket_folder = bucket.replace('ACC_', '').title()
    if bucket == 'ALL':
        bucket_folder = 'General'
    
    # Create type subfolder
    type_folder = 'Submittals' if item_type == 'Submittal' else 'RFIs'
    
    # Create item folder name - include title for easier distinction
    clean_id = sanitize_folder_name(identifier)
    folder_id = clean_id.replace(f'{item_type} #', '')
    if title:
        clean_title = sanitize_folder_name(title)[:100]  # Limit title length to 100 chars
        item_folder = f"{item_type} - {folder_id} - {clean_title}"
    else:
        item_folder = f"{item_type} - {folder_id}"
    
    # Full path
    full_path = base_path / bucket_folder / type_folder / item_folder
    
    try:
        full_path.mkdir(parents=True, exist_ok=True)
        return str(full_path)
    except Exception as e:
        print(f"Error creating folder {full_path}: {e}")
        return None

# =============================================================================
# FILE-BASED RESPONSE FORMS
# =============================================================================

TEMPLATES_DIR = BASE_DIR / "templates"

def generate_reviewer_form_html(item_id):
    """Generate a self-contained HTML form for reviewer response and save it to the item folder."""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT i.*, 
               ir.display_name as reviewer_name, ir.email as reviewer_email,
               qcr.display_name as qcr_name, qcr.email as qcr_email
        FROM item i
        LEFT JOIN user ir ON i.initial_reviewer_id = ir.id
        LEFT JOIN user qcr ON i.qcr_id = qcr.id
        WHERE i.id = ?
    ''', (item_id,))
    item = cursor.fetchone()
    
    if not item:
        conn.close()
        return {'success': False, 'error': 'Item not found'}
    
    if not item['folder_link']:
        conn.close()
        return {'success': False, 'error': 'Item has no folder assigned'}
    
    # Generate token if not exists
    token = item['email_token_reviewer']
    if not token:
        token = generate_token()
        cursor.execute('UPDATE item SET email_token_reviewer = ? WHERE id = ?', (token, item_id))
        conn.commit()
    
    # Calculate due dates
    reviewer_due = item['initial_reviewer_due_date'] or 'N/A'
    qcr_due = item['qcr_due_date'] or 'N/A'
    
    conn.close()
    
    # Note: We no longer pre-populate files - HTA will scan the folder
    folder_files = []  # Empty - HTA loads files from folder directly
    
    # Load template (use HTA template for automatic file saving)
    template_path = TEMPLATES_DIR / "_RESPONSE_FORM_TEMPLATE_v3.hta"
    if not template_path.exists():
        # Fallback to HTML templates
        template_path = TEMPLATES_DIR / "_RESPONSE_FORM_TEMPLATE_v3.html"
        if not template_path.exists():
            template_path = TEMPLATES_DIR / "_RESPONSE_FORM_TEMPLATE_v2.html"
            if not template_path.exists():
                template_path = TEMPLATES_DIR / "_RESPONSE_FORM_TEMPLATE.html"
                if not template_path.exists():
                    return {'success': False, 'error': 'Reviewer form template not found'}
    
    with open(template_path, 'r', encoding='utf-8') as f:
        template = f.read()
    
    # Escape special characters for JavaScript embedding
    def js_escape(s):
        if not s:
            return ''
        return s.replace('\\', '\\\\').replace('"', '\\"').replace("'", "\\'").replace('\n', '\\n').replace('\r', '')
    
    # Create folder path URL for file:// link
    folder_path_url = str(item['folder_link']).replace('\\', '/')
    
    # Replace placeholders
    html = template.replace('{{ITEM_ID}}', str(item['id']))
    html = html.replace('{{ITEM_TYPE}}', item['type'] or '')
    html = html.replace('{{ITEM_IDENTIFIER}}', item['identifier'] or '')
    html = html.replace('{{ITEM_TITLE}}', js_escape(item['title']) or 'N/A')
    html = html.replace('{{DATE_RECEIVED}}', item['date_received'] or 'N/A')
    html = html.replace('{{REVIEWER_DUE_DATE}}', reviewer_due)
    html = html.replace('{{QCR_DUE_DATE}}', qcr_due)
    html = html.replace('{{CONTRACTOR_DUE_DATE}}', item['due_date'] or 'N/A')
    html = html.replace('{{REVIEWER_NAME}}', js_escape(item['reviewer_name']) or 'N/A')
    html = html.replace('{{REVIEWER_EMAIL}}', item['reviewer_email'] or '')
    html = html.replace('{{TOKEN}}', token)
    html = html.replace('{{FOLDER_PATH}}', js_escape(item['folder_link']) or '')
    html = html.replace('{{FOLDER_PATH_RAW}}', item['folder_link'] or '')
    html = html.replace('{{FOLDER_PATH_URL}}', folder_path_url)
    html = html.replace('{{FOLDER_FILES_JSON}}', json.dumps(folder_files))
    
    # Save to Responses subfolder (use .hta extension if HTA template, else .html)
    folder_path = Path(item['folder_link'])
    responses_folder = folder_path / "Responses"
    
    # Create Responses subfolder if it doesn't exist
    try:
        responses_folder.mkdir(exist_ok=True)
    except Exception as e:
        return {'success': False, 'error': f'Failed to create Responses folder: {e}'}
    
    if template_path.suffix == '.hta':
        form_path = responses_folder / "_RESPONSE_FORM.hta"
    else:
        form_path = responses_folder / "_RESPONSE_FORM.html"
    
    try:
        with open(form_path, 'w', encoding='utf-8') as f:
            f.write(html)
        return {'success': True, 'path': str(form_path)}
    except Exception as e:
        return {'success': False, 'error': f'Failed to save form: {e}'}


def generate_qcr_form_html(item_id):
    """Generate a self-contained HTML form for QCR response and save it to the item folder."""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT i.*, 
               ir.display_name as reviewer_name, ir.email as reviewer_email,
               qcr.display_name as qcr_name, qcr.email as qcr_email
        FROM item i
        LEFT JOIN user ir ON i.initial_reviewer_id = ir.id
        LEFT JOIN user qcr ON i.qcr_id = qcr.id
        WHERE i.id = ?
    ''', (item_id,))
    item = cursor.fetchone()
    
    if not item:
        conn.close()
        return {'success': False, 'error': 'Item not found'}
    
    if not item['folder_link']:
        conn.close()
        return {'success': False, 'error': 'Item has no folder assigned'}
    
    # Generate token if not exists
    token = item['email_token_qcr']
    if not token:
        token = generate_token()
        cursor.execute('UPDATE item SET email_token_qcr = ? WHERE id = ?', (token, item_id))
        conn.commit()
    
    # Parse reviewer selected files
    reviewer_files = []
    if item['reviewer_selected_files']:
        try:
            reviewer_files = json.loads(item['reviewer_selected_files'])
        except:
            pass
    
    conn.close()
    
    # Load template (use HTA template for automatic file saving)
    template_path = TEMPLATES_DIR / "_QCR_FORM_TEMPLATE_v3.hta"
    if not template_path.exists():
        # Fall back to HTML templates
        template_path = TEMPLATES_DIR / "_QCR_FORM_TEMPLATE_v2.html"
        if not template_path.exists():
            template_path = TEMPLATES_DIR / "_QCR_FORM_TEMPLATE.html"
            if not template_path.exists():
                return {'success': False, 'error': 'QCR form template not found'}
    
    with open(template_path, 'r', encoding='utf-8') as f:
        template = f.read()
    
    # Escape special characters for JavaScript embedding
    def js_escape(s):
        if not s:
            return ''
        return s.replace('\\', '\\\\').replace('"', '\\"').replace("'", "\\'")
    
    # Format reviewer selected files as HTML list items
    reviewer_files_html = ''
    reviewer_files_text = 'None selected'
    reviewer_files_js = ''
    if reviewer_files:
        for f in reviewer_files:
            reviewer_files_html += f'<li>{f}</li>'
        reviewer_files_text = '; '.join(reviewer_files)
        reviewer_files_js = ', '.join([f'"{js_escape(f)}"' for f in reviewer_files])
    else:
        reviewer_files_html = '<li><em>None selected</em></li>'
    
    # Get response version
    response_version = item['reviewer_response_version'] if item['reviewer_response_version'] else 1
    
    # Replace placeholders
    html = template.replace('{{ITEM_ID}}', str(item['id']))
    html = html.replace('{{ITEM_TYPE}}', item['type'] or '')
    html = html.replace('{{ITEM_IDENTIFIER}}', item['identifier'] or '')
    html = html.replace('{{ITEM_TITLE}}', js_escape(item['title']) or 'N/A')
    html = html.replace('{{DATE_RECEIVED}}', item['date_received'] or 'N/A')
    html = html.replace('{{QCR_DUE_DATE}}', item['qcr_due_date'] or 'N/A')
    html = html.replace('{{CONTRACTOR_DUE_DATE}}', item['due_date'] or 'N/A')
    html = html.replace('{{PRIORITY}}', item['priority'] or 'Normal')
    html = html.replace('{{REVIEWER_NAME}}', js_escape(item['reviewer_name']) or 'N/A')
    html = html.replace('{{QCR_NAME}}', js_escape(item['qcr_name']) or 'N/A')
    html = html.replace('{{QCR_EMAIL}}', item['qcr_email'] or '')
    html = html.replace('{{TOKEN}}', token)
    html = html.replace('{{FOLDER_PATH}}', js_escape(item['folder_link']) or '')
    html = html.replace('{{FOLDER_PATH_RAW}}', item['folder_link'] or '')
    html = html.replace('{{REVIEWER_RESPONSE_CATEGORY}}', item['reviewer_response_category'] or 'Not specified')
    html = html.replace('{{REVIEWER_NOTES}}', js_escape(item['reviewer_notes'] or item['reviewer_response_text'] or 'No notes provided'))
    html = html.replace('{{REVIEWER_INTERNAL_NOTES}}', js_escape(item['reviewer_internal_notes'] or ''))
    html = html.replace('{{REVIEWER_INTERNAL_NOTES_DISPLAY}}', 'block' if item['reviewer_internal_notes'] else 'none')
    html = html.replace('{{REVIEWER_SELECTED_FILES}}', reviewer_files_html)
    html = html.replace('{{REVIEWER_SELECTED_FILES_TEXT}}', reviewer_files_text)
    html = html.replace('{{REVIEWER_SELECTED_FILES_JS}}', reviewer_files_js)
    html = html.replace('{{RESPONSE_VERSION}}', str(response_version))
    
    # Save to Responses subfolder (use .hta extension if HTA template, else .html)
    folder_path = Path(item['folder_link'])
    responses_folder = folder_path / "Responses"
    
    # Create Responses subfolder if it doesn't exist
    try:
        responses_folder.mkdir(exist_ok=True)
    except Exception as e:
        return {'success': False, 'error': f'Failed to create Responses folder: {e}'}
    
    if template_path.suffix == '.hta':
        form_path = responses_folder / "_QCR_RESPONSE_FORM.hta"
    else:
        form_path = responses_folder / "_QCR_FORM.html"
    
    try:
        with open(form_path, 'w', encoding='utf-8') as f:
            f.write(html)
        return {'success': True, 'path': str(form_path)}
    except Exception as e:
        return {'success': False, 'error': f'Failed to save form: {e}'}


def process_reviewer_response_json(json_path):
    """Process a _reviewer_response.json file and import it into the database."""
    try:
        with open(json_path, 'r', encoding='utf-8-sig') as f:
            data = json.load(f)
        
        # Validate it's the right type
        if data.get('_form_type') != 'reviewer_response':
            return {'success': False, 'error': 'Invalid form type'}
        
        token = data.get('token')
        if not token:
            return {'success': False, 'error': 'Missing token'}
        
        conn = get_db()
        cursor = conn.cursor()
        
        # Find item by token
        cursor.execute('SELECT id, reviewer_response_version FROM item WHERE email_token_reviewer = ?', (token,))
        item = cursor.fetchone()
        
        if not item:
            conn.close()
            return {'success': False, 'error': 'Invalid token - item not found'}
        
        item_id = item['id']
        current_version = item['reviewer_response_version'] or 0
        
        # Save to history if there's an existing response
        cursor.execute('SELECT reviewer_response_at FROM item WHERE id = ?', (item_id,))
        existing = cursor.fetchone()
        if existing and existing['reviewer_response_at']:
            cursor.execute('''
                INSERT INTO reviewer_response_history 
                (item_id, version, response_category, response_text, notes, selected_files, submitted_at)
                SELECT id, reviewer_response_version, reviewer_response_category, 
                       reviewer_response_text, reviewer_notes, reviewer_selected_files, reviewer_response_at
                FROM item WHERE id = ?
            ''', (item_id,))
            current_version += 1
        
        # Update item with new response
        selected_files_json = json.dumps(data.get('selected_files', []))
        cursor.execute('''
            UPDATE item SET
                reviewer_response_category = ?,
                reviewer_notes = ?,
                reviewer_internal_notes = ?,
                reviewer_selected_files = ?,
                reviewer_response_at = ?,
                reviewer_response_status = 'Responded',
                reviewer_response_version = ?
            WHERE id = ?
        ''', (
            data.get('response_category'),
            data.get('notes'),
            data.get('internal_notes'),
            selected_files_json,
            data.get('_submitted_at', datetime.now().isoformat()),
            current_version,
            item_id
        ))
        
        conn.commit()
        conn.close()
        
        # Rename processed file
        processed_path = json_path.parent / f"_reviewer_response_processed_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        json_path.rename(processed_path)
        
        # Send QCR assignment email now that reviewer has responded
        try:
            is_revision = current_version > 1
            qcr_result = send_qcr_assignment_email(item_id, is_revision=is_revision, version=current_version)
            if qcr_result['success']:
                print(f"  [Watcher] QCR email sent for item {item_id}")
            else:
                print(f"  [Watcher] Failed to send QCR email: {qcr_result.get('error')}")
        except Exception as e:
            print(f"  [Watcher] Error sending QCR email: {e}")
        
        return {'success': True, 'item_id': item_id, 'version': current_version}
        
    except json.JSONDecodeError:
        return {'success': False, 'error': 'Invalid JSON file'}
    except Exception as e:
        return {'success': False, 'error': str(e)}


def process_qcr_response_json(json_path):
    """Process a _qcr_response.json file and import it into the database."""
    try:
        with open(json_path, 'r', encoding='utf-8-sig') as f:
            data = json.load(f)
        
        # Validate it's the right type
        if data.get('_form_type') != 'qcr_response':
            return {'success': False, 'error': 'Invalid form type'}
        
        token = data.get('token')
        if not token:
            return {'success': False, 'error': 'Missing token'}
        
        conn = get_db()
        cursor = conn.cursor()
        
        # Find item by token
        cursor.execute('SELECT id FROM item WHERE email_token_qcr = ?', (token,))
        item = cursor.fetchone()
        
        if not item:
            conn.close()
            return {'success': False, 'error': 'Invalid token - item not found'}
        
        item_id = item['id']
        qc_action = data.get('qc_action')
        
        if qc_action == 'Send Back':
            # Send back to reviewer
            cursor.execute('''
                UPDATE item SET
                    qcr_action = 'Send Back',
                    qcr_notes = ?,
                    qcr_internal_notes = ?,
                    qcr_response_at = ?,
                    qcr_response_status = 'Waiting for Revision',
                    reviewer_response_status = 'Revision Requested'
                WHERE id = ?
            ''', (
                data.get('qcr_notes'),
                data.get('qcr_internal_notes'),
                data.get('_submitted_at', datetime.now().isoformat()),
                item_id
            ))
        else:
            # Approve or Modify
            selected_files_json = json.dumps(data.get('selected_files', []))
            final_response_text = data.get('response_text', '')  # HTA sends 'response_text'
            cursor.execute('''
                UPDATE item SET
                    qcr_action = ?,
                    qcr_notes = ?,
                    qcr_internal_notes = ?,
                    qcr_response_at = ?,
                    qcr_response_status = 'Responded',
                    qcr_response_mode = ?,
                    qcr_response_text = ?,
                    qcr_response_category = ?,
                    final_response_category = ?,
                    final_response_text = ?,
                    final_response_files = ?,
                    status = 'Ready for Response'
                WHERE id = ?
            ''', (
                qc_action,
                data.get('qcr_notes'),
                data.get('qcr_internal_notes'),
                data.get('_submitted_at', datetime.now().isoformat()),
                data.get('response_mode'),
                final_response_text,
                data.get('response_category'),
                data.get('response_category'),
                final_response_text,
                selected_files_json,
                item_id
            ))
        
        conn.commit()
        conn.close()
        
        # Rename processed file
        processed_path = json_path.parent / f"_qcr_response_processed_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        json_path.rename(processed_path)
        
        # Get item details for notifications
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT i.*, 
                   ir.display_name as reviewer_name, ir.email as reviewer_email,
                   qcr.display_name as qcr_name, qcr.email as qcr_email
            FROM item i
            LEFT JOIN user ir ON i.initial_reviewer_id = ir.id
            LEFT JOIN user qcr ON i.qcr_id = qcr.id
            WHERE i.id = ?
        ''', (item_id,))
        item_info = cursor.fetchone()
        conn.close()
        
        if item_info:
            qcr_notes = data.get('qcr_notes', '')
            final_category = data.get('response_category')
            final_text = data.get('response_text', '')  # HTA sends 'response_text'
            
            # Create system notifications based on QC action
            if qc_action == 'Approve' or qc_action == 'Modify':
                create_notification(
                    'response_ready',
                    f'✅ Response Ready: {item_info["type"]} {item_info["identifier"]}',
                    f'QC review complete. The response for "{item_info["title"] or item_info["identifier"]}" is ready to be sent to the contractor. Final category: {final_category}',
                    item_id=item_id,
                    action_url=f'/api/items/{item_id}/complete',
                    action_label='Mark Complete'
                )
                
                # Send confirmation emails to both QCR and reviewer
                try:
                    email_result = send_qcr_completion_confirmation_email(
                        item_id, qc_action, qcr_notes, 
                        final_category=final_category, 
                        final_text=final_text
                    )
                    if email_result.get('success'):
                        print(f"  [Watcher] QC completion confirmation emails sent for item {item_id}")
                    else:
                        print(f"  [Watcher] Failed to send QC confirmation emails: {email_result.get('error')}")
                except Exception as e:
                    print(f"  [Watcher] Error sending QC confirmation emails: {e}")
                
            elif qc_action == 'Send Back':
                create_notification(
                    'sent_back',
                    f'↩️ Sent Back: {item_info["type"]} {item_info["identifier"]}',
                    f'The item "{item_info["title"] or item_info["identifier"]}" has been sent back to the reviewer for revisions.',
                    item_id=item_id
                )
                
                # Send revision request to reviewer
                try:
                    reviewer_result = send_reviewer_assignment_email(item_id, is_revision=True, qcr_notes=qcr_notes)
                    if reviewer_result['success']:
                        print(f"  [Watcher] Revision request email sent to reviewer for item {item_id}")
                    else:
                        print(f"  [Watcher] Failed to send revision email: {reviewer_result.get('error')}")
                except Exception as e:
                    print(f"  [Watcher] Error sending revision email: {e}")
        
        return {'success': True, 'item_id': item_id, 'action': qc_action}
        
    except json.JSONDecodeError:
        return {'success': False, 'error': 'Invalid JSON file'}
    except Exception as e:
        return {'success': False, 'error': str(e)}


def process_multi_reviewer_response_json(json_path):
    """Process a multi-reviewer response JSON file from local HTA form."""
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # Verify this is a multi-reviewer response
        if data.get('_form_type') != 'multi_reviewer_response':
            return {'success': False, 'error': 'Not a multi-reviewer response file'}
        
        item_id = data.get('item_id')
        token = data.get('token')
        response_category = data.get('response_category')
        internal_notes = data.get('internal_notes', '')
        reviewer_name = data.get('reviewer_name', '')
        
        if not item_id or not token:
            return {'success': False, 'error': 'Missing item_id or token'}
        
        conn = get_db()
        cursor = conn.cursor()
        
        # Find the reviewer by token
        cursor.execute('''
            SELECT ir.*, i.qcr_id, i.qcr_email_sent_at
            FROM item_reviewers ir
            JOIN item i ON ir.item_id = i.id
            WHERE ir.email_token = ?
        ''', (token,))
        reviewer = cursor.fetchone()
        
        if not reviewer:
            conn.close()
            return {'success': False, 'error': 'Invalid token - reviewer not found'}
        
        # Check if this is a resubmission
        is_resubmission = reviewer['response_at'] is not None
        qcr_email_already_sent = reviewer['qcr_email_sent_at'] is not None
        
        # Calculate new version
        new_version = (reviewer['response_version'] or 0) + 1
        
        # Update reviewer response (allow resubmissions)
        cursor.execute('''
            UPDATE item_reviewers SET
                response_at = ?,
                response_category = ?,
                internal_notes = ?,
                response_version = ?,
                needs_response = 0
            WHERE id = ?
        ''', (
            data.get('_submitted_at', datetime.now().isoformat()),
            response_category,
            internal_notes,
            new_version,
            reviewer['id']
        ))
        
        item_id = reviewer['item_id']
        
        # Check if all reviewers have now responded
        # Note: needs_response is used for selective send-back, but for initial completion
        # we check all reviewers regardless of needs_response
        cursor.execute('''
            SELECT COUNT(*) as total, SUM(CASE WHEN response_at IS NOT NULL THEN 1 ELSE 0 END) as responded
            FROM item_reviewers
            WHERE item_id = ?
        ''', (item_id,))
        count_result = cursor.fetchone()
        all_responded = (count_result['total'] > 0 and count_result['total'] == count_result['responded'])
        
        # Handle resubmission after QCR email was already sent
        if is_resubmission and qcr_email_already_sent:
            # This is an updated response after QCR was notified - send updated QCR email
            conn.commit()
            conn.close()
            
            if reviewer['qcr_id']:
                email_result = send_email_with_retry(
                    send_multi_reviewer_qcr_email, item_id, 'multi_reviewer_qcr'
                )
                if email_result.get('skipped'):
                    print(f"  [Watcher] QCR already has latest info for item {item_id}")
                elif email_result.get('success'):
                    print(f"  [Watcher] Updated QCR email sent for multi-reviewer item {item_id} (reviewer {reviewer_name} resubmitted)")
                elif email_result.get('queued'):
                    print(f"  [Watcher] Updated QCR email queued for retry for item {item_id}")
                else:
                    print(f"  [Watcher] Failed to send updated QCR email: {email_result.get('error')}")
            
            # Rename processed file
            try:
                processed_path = json_path.parent / f"_multi_reviewer_response_resubmit_{datetime.now().strftime('%Y%m%d_%H%M%S%f')}.json"
                json_path.rename(processed_path)
            except Exception as e:
                print(f"  [Watcher] Warning: Could not rename file {json_path.name}: {e}")
            
            return {
                'success': True, 
                'item_id': item_id, 
                'reviewer': reviewer_name,
                'resubmission': True,
                'qcr_notified': True
            }
        
        if all_responded:
            # Update item status to In QC
            cursor.execute('''
                UPDATE item SET 
                    status = 'In QC',
                    reviewer_response_status = 'All Responded'
                WHERE id = ?
            ''', (item_id,))
            
            conn.commit()
            
            # Re-check if QCR email was already sent (in case another response was processed first)
            cursor.execute('SELECT qcr_email_sent_at FROM item WHERE id = ?', (item_id,))
            current_item = cursor.fetchone()
            qcr_email_already_sent_now = current_item['qcr_email_sent_at'] is not None
            
            conn.close()
            
            # Send QCR assignment email now that all reviewers have responded
            # Only send if QCR email hasn't already been sent (avoid duplicates)
            if reviewer['qcr_id'] and not qcr_email_already_sent_now:
                email_result = send_email_with_retry(
                    send_multi_reviewer_qcr_email, item_id, 'multi_reviewer_qcr'
                )
                if email_result.get('success'):
                    print(f"  [Watcher] QCR email sent for multi-reviewer item {item_id} (all reviewers responded)")
                elif email_result.get('queued'):
                    print(f"  [Watcher] QCR email queued for retry for item {item_id}")
                else:
                    print(f"  [Watcher] Failed to send QCR email: {email_result.get('error')}")
            elif qcr_email_already_sent_now:
                print(f"  [Watcher] QCR already notified for item {item_id}, skipping duplicate email")
            
            # Rename processed file (with retry for timestamp collision)
            try:
                processed_path = json_path.parent / f"_multi_reviewer_response_processed_{datetime.now().strftime('%Y%m%d_%H%M%S%f')}.json"
                json_path.rename(processed_path)
            except Exception as e:
                print(f"  [Watcher] Warning: Could not rename file {json_path.name}: {e}")
                # Try with a unique suffix
                try:
                    import random
                    processed_path = json_path.parent / f"_multi_reviewer_response_processed_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{random.randint(1000,9999)}.json"
                    json_path.rename(processed_path)
                except:
                    pass  # File will be reprocessed but that's OK, reviewer already marked as responded
            
            return {
                'success': True, 
                'item_id': item_id, 
                'reviewer': reviewer_name,
                'all_responded': True
            }
        else:
            cursor.execute('''
                UPDATE item SET status = 'In Review' WHERE id = ? AND status = 'Assigned'
            ''', (item_id,))
            conn.commit()
            conn.close()
            
            # Rename processed file (with retry for timestamp collision)
            try:
                processed_path = json_path.parent / f"_multi_reviewer_response_processed_{datetime.now().strftime('%Y%m%d_%H%M%S%f')}.json"
                json_path.rename(processed_path)
            except Exception as e:
                print(f"  [Watcher] Warning: Could not rename file {json_path.name}: {e}")
                try:
                    import random
                    processed_path = json_path.parent / f"_multi_reviewer_response_processed_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{random.randint(1000,9999)}.json"
                    json_path.rename(processed_path)
                except:
                    pass
            
            return {
                'success': True, 
                'item_id': item_id, 
                'reviewer': reviewer_name,
                'all_responded': False,
                'responded': count_result['responded'],
                'total': count_result['total']
            }
        
    except json.JSONDecodeError:
        return {'success': False, 'error': 'Invalid JSON file'}
    except Exception as e:
        return {'success': False, 'error': str(e)}


def process_multi_reviewer_qcr_response_json(json_path):
    """Process a multi-reviewer QCR response JSON file from local HTA form."""
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # Verify this is a multi-reviewer QCR response
        if data.get('_form_type') != 'multi_reviewer_qcr_response':
            return {'success': False, 'error': 'Not a multi-reviewer QCR response file'}
        
        item_id = data.get('item_id')
        token = data.get('token')
        qcr_action = data.get('qcr_action')
        
        if not item_id or not token:
            return {'success': False, 'error': 'Missing item_id or token'}
        
        conn = get_db()
        cursor = conn.cursor()
        
        # Verify token matches
        cursor.execute('SELECT * FROM item WHERE id = ? AND email_token_qcr = ?', (item_id, token))
        item = cursor.fetchone()
        
        if not item:
            conn.close()
            return {'success': False, 'error': 'Invalid token - item not found'}
        
        if qcr_action == 'Complete':
            # Complete the response
            response_category = data.get('response_category')
            response_text = data.get('response_text', '')
            qcr_internal_notes = data.get('qcr_internal_notes', '')
            
            cursor.execute('''
                UPDATE item SET
                    qcr_action = 'Approve',
                    qcr_notes = ?,
                    qcr_internal_notes = ?,
                    qcr_response_at = ?,
                    qcr_response_status = 'Responded',
                    qcr_response_mode = 'Revise',
                    qcr_response_text = ?,
                    qcr_response_category = ?,
                    final_response_category = ?,
                    final_response_text = ?,
                    status = 'Ready for Response'
                WHERE id = ?
            ''', (
                response_text,
                qcr_internal_notes,
                data.get('_submitted_at', datetime.now().isoformat()),
                response_text,
                response_category,
                response_category,
                response_text,
                item_id
            ))
            conn.commit()
            conn.close()
            
            # Rename processed file
            try:
                processed_path = json_path.parent / f"_multi_reviewer_qcr_response_processed_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
                json_path.rename(processed_path)
            except Exception as e:
                print(f"  [Watcher] Warning: Could not rename file {json_path.name}: {e}")
            
            # Create notification
            create_notification(
                'response_ready',
                f'Response Ready: {item["type"]} {item["identifier"]}',
                f'QC review complete. The response for "{item["title"] or item["identifier"]}" is ready to be sent.',
                item_id=item_id
            )
            
            # Send completion confirmation email to QCR and ALL reviewers
            try:
                email_result = send_multi_reviewer_completion_email(item_id, response_category, response_text)
                if email_result.get('success'):
                    print(f"  [Watcher] Completion email sent for multi-reviewer item {item_id}")
                else:
                    print(f"  [Watcher] Failed to send completion email: {email_result.get('error')}")
            except Exception as e:
                print(f"  [Watcher] Error sending completion email: {e}")
            
            return {'success': True, 'item_id': item_id, 'action': 'Complete'}
            
        elif qcr_action == 'Send Back':
            # Send back to selected reviewers (or all if none specified)
            sendback_notes = data.get('sendback_notes', '')
            qcr_internal_notes = data.get('qcr_internal_notes', '')
            sendback_reviewer_ids = data.get('sendback_reviewer_ids', [])  # List of selected reviewer IDs
            
            cursor.execute('''
                UPDATE item SET
                    qcr_action = 'Send Back',
                    qcr_notes = ?,
                    qcr_internal_notes = ?,
                    qcr_response_at = ?,
                    qcr_response_status = 'Revision Requested',
                    status = 'In Review'
                WHERE id = ?
            ''', (
                sendback_notes,
                qcr_internal_notes,
                data.get('_submitted_at', datetime.now().isoformat()),
                item_id
            ))
            
            # If specific reviewers selected, only reset and require those
            if sendback_reviewer_ids and len(sendback_reviewer_ids) > 0:
                # First, set all reviewers to NOT need response
                cursor.execute('''
                    UPDATE item_reviewers SET needs_response = 0 WHERE item_id = ?
                ''', (item_id,))
                
                # Reset only selected reviewer responses and mark them as needing response
                for reviewer_id in sendback_reviewer_ids:
                    cursor.execute('''
                        UPDATE item_reviewers SET
                            response_at = NULL,
                            response_category = NULL,
                            internal_notes = NULL,
                            response_version = response_version + 1,
                            needs_response = 1
                        WHERE item_id = ? AND id = ?
                    ''', (item_id, reviewer_id))
            else:
                # Reset all reviewer responses (original behavior)
                cursor.execute('''
                    UPDATE item_reviewers SET
                        response_at = NULL,
                        response_category = NULL,
                        internal_notes = NULL,
                        response_version = response_version + 1,
                        needs_response = 1
                    WHERE item_id = ?
                ''', (item_id,))
            
            # Reset QCR notification tracking since we're sending back
            cursor.execute('''
                UPDATE item SET qcr_notified_at = NULL WHERE id = ?
            ''', (item_id,))
            
            conn.commit()
            conn.close()
            
            # Rename processed file
            try:
                processed_path = json_path.parent / f"_multi_reviewer_qcr_response_processed_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
                json_path.rename(processed_path)
            except Exception as e:
                print(f"  [Watcher] Warning: Could not rename file {json_path.name}: {e}")
            
            # Send sendback emails to selected reviewers (or all if none specified)
            # Use retry logic since Outlook COM can fail intermittently
            reviewer_ids_to_send = sendback_reviewer_ids if sendback_reviewer_ids else None
            email_result = send_email_with_retry(
                lambda item_id, **kw: send_multi_reviewer_sendback_emails(item_id, kw.get('feedback', ''), kw.get('reviewer_ids')),
                item_id,
                'multi_reviewer_sendback',
                feedback=sendback_notes,
                reviewer_ids=reviewer_ids_to_send
            )
            if email_result.get('success'):
                sent_count = email_result.get('sent_count', 'unknown')
                print(f"  [Watcher] Sendback emails sent to {sent_count} reviewers for item {item_id}")
            elif email_result.get('queued'):
                print(f"  [Watcher] Sendback emails queued for retry for item {item_id}")
            else:
                print(f"  [Watcher] Failed to send sendback emails: {email_result.get('error')}")
            
            return {'success': True, 'item_id': item_id, 'action': 'Send Back', 'reviewers_sent_back': len(sendback_reviewer_ids) if sendback_reviewer_ids else 'all'}
        else:
            conn.close()
            return {'success': False, 'error': f'Unknown QCR action: {qcr_action}'}
        
    except json.JSONDecodeError:
        return {'success': False, 'error': 'Invalid JSON file'}
    except Exception as e:
        return {'success': False, 'error': str(e)}


def scan_folders_for_responses():
    """Scan all item folders for JSON response files and process them."""
    base_path = Path(CONFIG['base_folder_path'])
    results = {
        'reviewer_responses': [],
        'qcr_responses': [],
        'multi_reviewer_responses': [],
        'errors': []
    }
    
    if not base_path.exists():
        return results
    
    # Scan for _reviewer_response.json files
    for json_file in base_path.rglob('_reviewer_response.json'):
        result = process_reviewer_response_json(json_file)
        if result['success']:
            results['reviewer_responses'].append({
                'path': str(json_file),
                'item_id': result.get('item_id'),
                'version': result.get('version')
            })
        else:
            results['errors'].append({
                'path': str(json_file),
                'error': result.get('error')
            })
    
    # Scan for _qcr_response.json files
    for json_file in base_path.rglob('_qcr_response.json'):
        result = process_qcr_response_json(json_file)
        if result['success']:
            results['qcr_responses'].append({
                'path': str(json_file),
                'item_id': result.get('item_id'),
                'action': result.get('action')
            })
        else:
            results['errors'].append({
                'path': str(json_file),
                'error': result.get('error')
            })
    
    # Scan for _multi_reviewer_response_*.json files
    for json_file in base_path.rglob('_multi_reviewer_response_*.json'):
        # Skip already processed files (both _processed_ and _resubmit_ variants)
        if '_processed_' in json_file.name or '_already_processed_' in json_file.name or '_resubmit_' in json_file.name:
            continue
        result = process_multi_reviewer_response_json(json_file)
        if result['success']:
            results['multi_reviewer_responses'].append({
                'path': str(json_file),
                'item_id': result.get('item_id'),
                'reviewer': result.get('reviewer'),
                'all_responded': result.get('all_responded')
            })
        else:
            results['errors'].append({
                'path': str(json_file),
                'error': result.get('error')
            })
    
    # Scan for _multi_reviewer_qcr_response.json files
    for json_file in base_path.rglob('_multi_reviewer_qcr_response.json'):
        result = process_multi_reviewer_qcr_response_json(json_file)
        if result['success']:
            results['qcr_responses'].append({
                'path': str(json_file),
                'item_id': result.get('item_id'),
                'action': result.get('action'),
                'multi_reviewer': True
            })
        else:
            results['errors'].append({
                'path': str(json_file),
                'error': result.get('error')
            })
    
    return results


class FolderResponseWatcher:
    """Background watcher for JSON response files in item folders."""
    
    def __init__(self, interval_seconds=60):
        self.running = False
        self.thread = None
        self.interval = interval_seconds
        self.last_scan = None
        self.scan_count = 0
        self.logged_errors = {}  # Track logged errors to avoid spam
    
    def start(self):
        """Start the watcher thread."""
        self.running = True
        self.thread = threading.Thread(target=self._watch_loop, daemon=True)
        self.thread.start()
        print(f"Folder response watcher started (scanning every {self.interval}s)")
    
    def stop(self):
        """Stop the watcher thread."""
        self.running = False
        if self.thread:
            self.thread.join(timeout=5)
        print("Folder response watcher stopped")
    
    def _watch_loop(self):
        """Main watch loop."""
        while self.running:
            try:
                results = scan_folders_for_responses()
                self.last_scan = datetime.now()
                self.scan_count += 1
                
                # Log any processed responses
                for resp in results['reviewer_responses']:
                    print(f"  [Watcher] Imported reviewer response for item {resp['item_id']} (v{resp['version']})")
                for resp in results['qcr_responses']:
                    print(f"  [Watcher] Imported QCR response for item {resp['item_id']} ({resp['action']})")
                for resp in results.get('multi_reviewer_responses', []):
                    if resp.get('all_responded'):
                        print(f"  [Watcher] Imported multi-reviewer response for item {resp['item_id']} from {resp['reviewer']} - ALL RESPONDED, QCR notified")
                    else:
                        print(f"  [Watcher] Imported multi-reviewer response for item {resp['item_id']} from {resp['reviewer']} - waiting for others")
                
                # Only log new/changed errors to avoid spam
                for err in results['errors']:
                    error_key = f"{err['path']}:{err['error']}"
                    if error_key not in self.logged_errors:
                        print(f"  [Watcher] Error processing {err['path']}: {err['error']}")
                        self.logged_errors[error_key] = True
                        # Limit logged errors to prevent memory growth
                        if len(self.logged_errors) > 100:
                            # Clear oldest half
                            keys_to_remove = list(self.logged_errors.keys())[:50]
                            for key in keys_to_remove:
                                del self.logged_errors[key]
                
                # Process any pending emails that failed earlier
                process_pending_emails()
                    
            except Exception as e:
                print(f"  [Watcher] Scan error: {e}")
            
            # Sleep in small increments so we can stop quickly
            for _ in range(self.interval):
                if not self.running:
                    break
                time.sleep(1)


# Global watcher instance
folder_watcher = FolderResponseWatcher(interval_seconds=30)

# =============================================================================
# REMINDER EMAIL SYSTEM
# =============================================================================

# PST timezone offset (UTC-8, or UTC-7 during DST)
# For simplicity, we'll use 8 AM PST = 16:00 UTC (or 15:00 UTC during DST)
REMINDER_HOUR_PST = 8  # 8 AM PST

def get_pst_now():
    """Get current time in PST (approximate, using UTC-8)."""
    utc_now = datetime.utcnow()
    pst_offset = timedelta(hours=-8)
    return utc_now + pst_offset

def is_past_reminder_time_today():
    """Check if we're past the reminder time (8 AM PST) today."""
    pst_now = get_pst_now()
    return pst_now.hour >= REMINDER_HOUR_PST

def has_reminder_been_sent(item_id, recipient_email, recipient_role, reminder_stage, item_reviewer_id=None):
    """Check if a specific reminder has already been sent."""
    conn = get_db()
    cursor = conn.cursor()
    
    if item_reviewer_id:
        cursor.execute('''
            SELECT id FROM reminder_log 
            WHERE item_id = ? AND recipient_email = ? AND recipient_role = ? AND reminder_stage = ? AND item_reviewer_id = ?
        ''', (item_id, recipient_email, recipient_role, reminder_stage, item_reviewer_id))
    else:
        cursor.execute('''
            SELECT id FROM reminder_log 
            WHERE item_id = ? AND recipient_email = ? AND recipient_role = ? AND reminder_stage = ?
        ''', (item_id, recipient_email, recipient_role, reminder_stage))
    
    result = cursor.fetchone()
    conn.close()
    return result is not None

def record_reminder_sent(item_id, reminder_type, recipient_email, recipient_role, due_date, reminder_stage, item_reviewer_id=None):
    """Record that a reminder has been sent."""
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute('''
            INSERT OR IGNORE INTO reminder_log (item_id, reminder_type, recipient_email, recipient_role, due_date, reminder_stage, item_reviewer_id)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (item_id, reminder_type, recipient_email, recipient_role, due_date, reminder_stage, item_reviewer_id))
        conn.commit()
    except Exception as e:
        print(f"  [Reminder] Error recording reminder: {e}")
    finally:
        conn.close()

def check_response_exists_local(item_id, role, reviewer_name=None):
    """Check if a response file exists for an item in local mode.
    
    Args:
        item_id: The item ID
        role: 'reviewer' or 'qcr'
        reviewer_name: For multi-reviewer mode, the reviewer's name
    
    Returns True if response exists, False otherwise.
    """
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('SELECT folder_link, multi_reviewer_mode FROM item WHERE id = ?', (item_id,))
    item = cursor.fetchone()
    conn.close()
    
    if not item or not item['folder_link']:
        return False
    
    folder_path = Path(item['folder_link'])
    responses_folder = folder_path / 'Responses'
    
    if not responses_folder.exists():
        return False
    
    if item['multi_reviewer_mode']:
        if role == 'reviewer' and reviewer_name:
            # Check for multi-reviewer response file
            safe_name = re.sub(r'[^a-zA-Z0-9]', '_', reviewer_name)
            response_file = responses_folder / f'_multi_reviewer_response_{safe_name}.json'
            # Also check for processed versions
            processed_file = responses_folder / f'_processed__multi_reviewer_response_{safe_name}.json'
            return response_file.exists() or processed_file.exists()
        elif role == 'qcr':
            response_file = responses_folder / '_multi_reviewer_qcr_response.json'
            processed_file = responses_folder / '_processed__multi_reviewer_qcr_response.json'
            return response_file.exists() or processed_file.exists()
    else:
        # Single reviewer mode
        if role == 'reviewer':
            response_file = responses_folder / '_reviewer_response.json'
            processed_file = responses_folder / '_processed__reviewer_response.json'
            return response_file.exists() or processed_file.exists()
        elif role == 'qcr':
            response_file = responses_folder / '_qcr_response.json'
            processed_file = responses_folder / '_processed__qcr_response.json'
            return response_file.exists() or processed_file.exists()
    
    return False

def get_items_needing_reminders():
    """Get all items that need reminder emails today.
    
    Returns dict with:
        - single_reviewer_items: Items where reviewer or QCR needs reminder
        - multi_reviewer_items: Multi-reviewer items with reviewers needing reminder
    """
    today = datetime.now().date()
    yesterday = today - timedelta(days=1)
    
    conn = get_db()
    cursor = conn.cursor()
    
    result = {
        'single_reviewer': [],  # (item, role, due_date, reminder_stage)
        'multi_reviewer': [],   # (item, reviewer_record, role, due_date, reminder_stage)
        'multi_reviewer_qcr': [] # (item, due_date, reminder_stage)
    }
    
    # =====================================================================
    # SINGLE REVIEWER MODE
    # =====================================================================
    # Get items where reviewer hasn't responded and reviewer due date is today or yesterday
    # Only for items that are open (not closed) and in the reviewer's court
    cursor.execute('''
        SELECT i.*, 
               ir.email as reviewer_email, ir.display_name as reviewer_name,
               qcr.email as qcr_email, qcr.display_name as qcr_name
        FROM item i
        LEFT JOIN user ir ON i.initial_reviewer_id = ir.id
        LEFT JOIN user qcr ON i.qcr_id = qcr.id
        WHERE i.multi_reviewer_mode = 0 
        AND i.closed_at IS NULL
        AND i.status IN ('Assigned', 'In Review')
        AND i.initial_reviewer_due_date IS NOT NULL
        AND DATE(i.initial_reviewer_due_date) <= ?
        AND i.reviewer_response_at IS NULL
        AND i.reviewer_email_sent_at IS NOT NULL
    ''', (today.strftime('%Y-%m-%d'),))
    
    for item in cursor.fetchall():
        item = dict(item)
        due_date = datetime.strptime(item['initial_reviewer_due_date'], '%Y-%m-%d').date()
        
        if due_date == today:
            reminder_stage = 'due_today'
        elif due_date < today:
            # Only send overdue reminder on the day after (not every day after)
            if due_date == yesterday:
                reminder_stage = 'overdue'
            else:
                continue  # Don't send reminders for items overdue by more than 1 day
        else:
            continue  # Future due date
        
        # Check if response file exists in local mode
        if is_local_mode() and check_response_exists_local(item['id'], 'reviewer'):
            continue  # Response already exists
        
        if item['reviewer_email']:
            result['single_reviewer'].append((item, 'reviewer', due_date, reminder_stage))
    
    # Get items where QCR hasn't responded and QCR due date is today or yesterday
    # (Item must be in 'In QC' status, meaning reviewer has submitted)
    # Only for items that are open (not closed) and in QCR's court
    cursor.execute('''
        SELECT i.*, 
               ir.email as reviewer_email, ir.display_name as reviewer_name,
               qcr.email as qcr_email, qcr.display_name as qcr_name
        FROM item i
        LEFT JOIN user ir ON i.initial_reviewer_id = ir.id
        LEFT JOIN user qcr ON i.qcr_id = qcr.id
        WHERE i.multi_reviewer_mode = 0 
        AND i.closed_at IS NULL
        AND i.status = 'In QC'
        AND i.qcr_due_date IS NOT NULL
        AND DATE(i.qcr_due_date) <= ?
        AND i.qcr_response_at IS NULL
        AND i.qcr_email_sent_at IS NOT NULL
        AND DATE(i.qcr_email_sent_at) < ?
    ''', (today.strftime('%Y-%m-%d'), today.strftime('%Y-%m-%d')))
    
    for item in cursor.fetchall():
        item = dict(item)
        due_date = datetime.strptime(item['qcr_due_date'], '%Y-%m-%d').date()
        qcr_email_sent_date = datetime.strptime(item['qcr_email_sent_at'][:10], '%Y-%m-%d').date() if item['qcr_email_sent_at'] else None
        
        if due_date == today:
            reminder_stage = 'due_today'
        elif due_date < today:
            # Send overdue reminder if:
            # 1. Due date was yesterday (normal case), OR
            # 2. Due date passed but assignment was sent yesterday (late assignment case)
            if due_date == yesterday:
                reminder_stage = 'overdue'
            elif qcr_email_sent_date == yesterday:
                # Assignment was sent yesterday for an already-overdue item
                # Send one overdue reminder today
                reminder_stage = 'overdue'
            else:
                continue
        else:
            continue
        
        # Check if response file exists in local mode
        if is_local_mode() and check_response_exists_local(item['id'], 'qcr'):
            continue
        
        if item['qcr_email']:
            result['single_reviewer'].append((item, 'qcr', due_date, reminder_stage))
    
    # =====================================================================
    # MULTI-REVIEWER MODE - Individual Reviewers
    # =====================================================================
    # Only for items that are open (not closed) and in reviewer's court
    cursor.execute('''
        SELECT i.*, 
               qcr.email as qcr_email, qcr.display_name as qcr_name
        FROM item i
        LEFT JOIN user qcr ON i.qcr_id = qcr.id
        WHERE i.multi_reviewer_mode = 1 
        AND i.closed_at IS NULL
        AND i.status IN ('Assigned', 'In Review')
        AND i.initial_reviewer_due_date IS NOT NULL
        AND DATE(i.initial_reviewer_due_date) <= ?
    ''', (today.strftime('%Y-%m-%d'),))
    
    for item in cursor.fetchall():
        item = dict(item)
        due_date = datetime.strptime(item['initial_reviewer_due_date'], '%Y-%m-%d').date()
        
        if due_date == today:
            reminder_stage = 'due_today'
        elif due_date < today:
            if due_date == yesterday:
                reminder_stage = 'overdue'
            else:
                continue
        else:
            continue
        
        # Get individual reviewers who haven't responded
        cursor.execute('''
            SELECT * FROM item_reviewers 
            WHERE item_id = ? 
            AND response_at IS NULL 
            AND email_sent_at IS NOT NULL
            AND needs_response = 1
        ''', (item['id'],))
        
        for reviewer in cursor.fetchall():
            reviewer = dict(reviewer)
            
            # Check if response file exists in local mode
            if is_local_mode() and check_response_exists_local(item['id'], 'reviewer', reviewer['reviewer_name']):
                continue
            
            result['multi_reviewer'].append((item, reviewer, 'reviewer', due_date, reminder_stage))
    
    # =====================================================================
    # MULTI-REVIEWER MODE - QCR
    # =====================================================================
    # Only for items that are open (not closed) and in QCR's court
    cursor.execute('''
        SELECT i.*, 
               qcr.email as qcr_email, qcr.display_name as qcr_name
        FROM item i
        LEFT JOIN user qcr ON i.qcr_id = qcr.id
        WHERE i.multi_reviewer_mode = 1 
        AND i.closed_at IS NULL
        AND i.status = 'In QC'
        AND i.qcr_due_date IS NOT NULL
        AND DATE(i.qcr_due_date) <= ?
        AND i.qcr_response_at IS NULL
        AND i.qcr_email_sent_at IS NOT NULL
        AND DATE(i.qcr_email_sent_at) < ?
    ''', (today.strftime('%Y-%m-%d'), today.strftime('%Y-%m-%d')))
    
    for item in cursor.fetchall():
        item = dict(item)
        due_date = datetime.strptime(item['qcr_due_date'], '%Y-%m-%d').date()
        qcr_email_sent_date = datetime.strptime(item['qcr_email_sent_at'][:10], '%Y-%m-%d').date() if item['qcr_email_sent_at'] else None
        
        if due_date == today:
            reminder_stage = 'due_today'
        elif due_date < today:
            # Send overdue reminder if:
            # 1. Due date was yesterday (normal case), OR
            # 2. Due date passed but assignment was sent yesterday (late assignment case)
            if due_date == yesterday:
                reminder_stage = 'overdue'
            elif qcr_email_sent_date == yesterday:
                # Assignment was sent yesterday for an already-overdue item
                # Send one overdue reminder today
                reminder_stage = 'overdue'
            else:
                continue
        else:
            continue
        
        # Check if response file exists in local mode
        if is_local_mode() and check_response_exists_local(item['id'], 'qcr'):
            continue
        
        if item['qcr_email']:
            result['multi_reviewer_qcr'].append((item, due_date, reminder_stage))
    
    conn.close()
    return result

def send_single_reviewer_reminder_email(item, role, due_date, reminder_stage):
    """Send a reminder email for single-reviewer mode.
    
    This re-sends the original assignment email but with a modified subject line.
    Only emails the person whose turn it is (no CCs).
    """
    if not HAS_WIN32COM:
        return {'success': False, 'error': 'Outlook not available'}
    
    item_id = item['id']
    
    # Check if this reminder has already been sent
    recipient_email = item['reviewer_email'] if role == 'reviewer' else item['qcr_email']
    if has_reminder_been_sent(item_id, recipient_email, role, reminder_stage):
        return {'success': True, 'skipped': True, 'reason': 'Already sent'}
    
    # Determine subject prefix
    if reminder_stage == 'due_today':
        subject_prefix = "REMINDER: DUE TODAY"
    else:
        subject_prefix = "REMINDER: OVERDUE"
    
    # Build subject line
    subject = f"{subject_prefix} - [LEB] {item['identifier']}"
    
    # Priority color
    priority_color = '#e67e22' if item['priority'] == 'Medium' else '#c0392b' if item['priority'] == 'High' else '#27ae60'
    
    # Folder link
    folder_path = item['folder_link'] or 'Not set'
    if folder_path != 'Not set':
        folder_link_html = f'<a href="file:///{folder_path.replace(chr(92), "/")}" style="color:#0078D4; text-decoration:underline;">{folder_path}</a>'
    else:
        folder_link_html = 'Not set'
    
    # Determine if using file-based forms
    use_file_form = is_local_mode() and folder_path != 'Not set'
    
    if role == 'reviewer':
        # Send reminder to reviewer
        reviewer_due_date = format_date_for_email(item['initial_reviewer_due_date'])
        qcr_due_date = format_date_for_email(item['qcr_due_date'])
        contractor_due_date = format_date_for_email(item['due_date'])
        
        if use_file_form:
            form_result = generate_reviewer_form_html(item_id)
            if form_result['success']:
                form_file_path = form_result['path']
                form_file_link = f'file:///{form_file_path.replace(chr(92), "/")}'
                
                html_body = f"""<div style="font-family:Segoe UI, Helvetica, Arial, sans-serif; color:#333; font-size:14px; line-height:1.5;">

    <!-- REMINDER BANNER -->
    <div style="background: {'#dc3545' if reminder_stage == 'overdue' else '#ffc107'}; color: {'white' if reminder_stage == 'overdue' else '#212529'}; padding: 15px; border-radius: 8px; margin-bottom: 20px; text-align: center;">
        <strong style="font-size: 16px;">{"⚠️ OVERDUE - RESPONSE REQUIRED IMMEDIATELY" if reminder_stage == 'overdue' else "⏰ REMINDER - DUE TODAY"}</strong>
        <div style="margin-top: 8px; font-size: 14px;">Your response was due: {due_date.strftime('%B %d, %Y')}</div>
    </div>

    <!-- HEADER -->
    <h2 style="color:#444; margin-bottom:6px;">
        [LEB] {item['identifier']} – Assigned to You
    </h2>

    <p style="margin-top:0; font-size:13px; color:#666;">
        This is a reminder that you have been assigned a review task that requires your response.
    </p>

    <!-- DIRECT LINK TO FORM -->
    <div style="margin:20px 0; text-align:center;">
        <table cellpadding="0" cellspacing="0" border="0" style="margin:0 auto;">
            <tr>
                <td align="center" bgcolor="#667eea" style="background:#667eea; border-radius:8px; padding:0;">
                    <a href="{form_file_link}" target="_blank"
                       style="background:#667eea; color:#ffffff; display:inline-block; font-family:Segoe UI,Arial,sans-serif;
                              font-size:16px; font-weight:bold; line-height:50px; text-align:center; text-decoration:none;
                              width:320px; -webkit-text-size-adjust:none; border-radius:8px;">
                        OPEN RESPONSE FORM
                    </a>
                </td>
            </tr>
        </table>
    </div>

    <!-- INFO TABLE -->
    <table cellpadding="8" cellspacing="0" width="100%" style="border-collapse:collapse; margin-top:10px;">
        <tr><td colspan="2" style="background:#f2f2f2; font-weight:bold; border:1px solid #ddd;">Item Information</td></tr>
        <tr><td style="width:160px; border:1px solid #ddd; font-weight:bold;">Type</td><td style="border:1px solid #ddd;">{item['type']}</td></tr>
        <tr><td style="border:1px solid #ddd; font-weight:bold;">Identifier</td><td style="border:1px solid #ddd;">{item['identifier']}</td></tr>
        <tr><td style="border:1px solid #ddd; font-weight:bold;">Title</td><td style="border:1px solid #ddd;">{item['title'] or 'N/A'}</td></tr>
        <tr><td style="border:1px solid #ddd; font-weight:bold;">Priority</td><td style="border:1px solid #ddd; color:{priority_color}; font-weight:bold;">{item['priority'] or 'Normal'}</td></tr>
        <tr><td style="border:1px solid #ddd; font-weight:bold;">Your Due Date</td><td style="border:1px solid #ddd; color:#c0392b; font-weight:bold;">{reviewer_due_date or 'N/A'}</td></tr>
        <tr><td style="border:1px solid #ddd; font-weight:bold;">QCR Due Date</td><td style="border:1px solid #ddd; color:#27ae60; font-weight:bold;">{qcr_due_date or 'N/A'}</td></tr>
        <tr><td style="border:1px solid #ddd; font-weight:bold;">Contractor Due Date</td><td style="border:1px solid #ddd; color:#c0392b; font-weight:bold;">{item['due_date'] or 'N/A'}</td></tr>
        <tr><td style="border:1px solid #ddd; font-weight:bold;">Project Folder</td><td style="border:1px solid #ddd;">{folder_link_html}</td></tr>
    </table>

    <p style="margin-top:20px; font-size:12px; color:#777;">
        <em>This is an automated reminder. Please submit your response as soon as possible.</em>
    </p>
</div>"""
            else:
                return {'success': False, 'error': f'Could not generate form: {form_result.get("error")}'}
        else:
            # Server-based form fallback
            app_host = get_app_host()
            token = item['email_token_reviewer']
            review_url = f"{app_host}/respond/reviewer?token={token}"
            html_body = f"""<div style="font-family:Segoe UI, Helvetica, Arial, sans-serif; color:#333;">
                <h2>{subject_prefix} - {item['identifier']}</h2>
                <p>Please submit your response: <a href="{review_url}">{review_url}</a></p>
            </div>"""
        
        # Send via Outlook - NO CC
        # Record BEFORE sending to prevent duplicates if Outlook blocks then releases the email
        record_reminder_sent(item_id, 'single_reviewer', item['reviewer_email'], 'reviewer', due_date.strftime('%Y-%m-%d'), reminder_stage)
        try:
            pythoncom.CoInitialize()
            outlook = win32com.client.Dispatch("Outlook.Application")
            mail = outlook.CreateItem(0)
            mail.Subject = subject
            mail.HTMLBody = html_body
            mail.To = item['reviewer_email']
            # NO CC for reminder emails
            mail.Send()
            
            print(f"  [Reminder] Sent {reminder_stage} reminder to reviewer for item {item_id}")
            return {'success': True}
        except Exception as e:
            # Note: reminder is already recorded to prevent duplicates even if send fails
            print(f"  [Reminder] Failed to send {reminder_stage} reminder to reviewer for item {item_id}: {e}")
            return {'success': False, 'error': str(e)}
        finally:
            pythoncom.CoUninitialize()
    
    else:  # role == 'qcr'
        # Send reminder to QCR
        qcr_due_date = format_date_for_email(item['qcr_due_date'])
        contractor_due_date = format_date_for_email(item['due_date'])
        
        if use_file_form:
            form_result = generate_qcr_form_html(item_id)
            if form_result['success']:
                form_file_path = form_result['path']
                form_file_link = f'file:///{form_file_path.replace(chr(92), "/")}'
                
                html_body = f"""<div style="font-family:Segoe UI, Helvetica, Arial, sans-serif; color:#333; font-size:14px; line-height:1.5;">

    <!-- REMINDER BANNER -->
    <div style="background: {'#dc3545' if reminder_stage == 'overdue' else '#ffc107'}; color: {'white' if reminder_stage == 'overdue' else '#212529'}; padding: 15px; border-radius: 8px; margin-bottom: 20px; text-align: center;">
        <strong style="font-size: 16px;">{"⚠️ OVERDUE - QC REVIEW REQUIRED IMMEDIATELY" if reminder_stage == 'overdue' else "⏰ REMINDER - QC REVIEW DUE TODAY"}</strong>
        <div style="margin-top: 8px; font-size: 14px;">Your QC review was due: {due_date.strftime('%B %d, %Y')}</div>
    </div>

    <!-- HEADER -->
    <h2 style="color:#444; margin-bottom:6px;">
        [LEB] {item['identifier']} – Ready for Your Review
    </h2>

    <p style="margin-top:0; font-size:13px; color:#666;">
        This is a reminder that a QC review is awaiting your action.
    </p>

    <!-- DIRECT LINK TO FORM -->
    <div style="margin:20px 0; text-align:center;">
        <table cellpadding="0" cellspacing="0" border="0" style="margin:0 auto;">
            <tr>
                <td align="center" bgcolor="#27ae60" style="background:#27ae60; border-radius:8px; padding:0;">
                    <a href="{form_file_link}" target="_blank"
                       style="background:#27ae60; color:#ffffff; display:inline-block; font-family:Segoe UI,Arial,sans-serif;
                              font-size:16px; font-weight:bold; line-height:50px; text-align:center; text-decoration:none;
                              width:320px; -webkit-text-size-adjust:none; border-radius:8px;">
                        OPEN QC REVIEW FORM
                    </a>
                </td>
            </tr>
        </table>
    </div>

    <!-- INFO TABLE -->
    <table cellpadding="8" cellspacing="0" width="100%" style="border-collapse:collapse; margin-top:10px;">
        <tr><td colspan="2" style="background:#f2f2f2; font-weight:bold; border:1px solid #ddd;">Item Information</td></tr>
        <tr><td style="width:160px; border:1px solid #ddd; font-weight:bold;">Type</td><td style="border:1px solid #ddd;">{item['type']}</td></tr>
        <tr><td style="border:1px solid #ddd; font-weight:bold;">Identifier</td><td style="border:1px solid #ddd;">{item['identifier']}</td></tr>
        <tr><td style="border:1px solid #ddd; font-weight:bold;">Title</td><td style="border:1px solid #ddd;">{item['title'] or 'N/A'}</td></tr>
        <tr><td style="border:1px solid #ddd; font-weight:bold;">Priority</td><td style="border:1px solid #ddd; color:{priority_color}; font-weight:bold;">{item['priority'] or 'Normal'}</td></tr>
        <tr><td style="border:1px solid #ddd; font-weight:bold;">Your Due Date</td><td style="border:1px solid #ddd; color:#c0392b; font-weight:bold;">{qcr_due_date}</td></tr>
        <tr><td style="border:1px solid #ddd; font-weight:bold;">Contractor Due Date</td><td style="border:1px solid #ddd; color:#c0392b; font-weight:bold;">{contractor_due_date}</td></tr>
        <tr><td style="border:1px solid #ddd; font-weight:bold;">Project Folder</td><td style="border:1px solid #ddd;">{folder_link_html}</td></tr>
    </table>

    <p style="margin-top:20px; font-size:12px; color:#777;">
        <em>This is an automated reminder. Please complete your QC review as soon as possible.</em>
    </p>
</div>"""
            else:
                return {'success': False, 'error': f'Could not generate form: {form_result.get("error")}'}
        else:
            app_host = get_app_host()
            token = item['email_token_qcr']
            review_url = f"{app_host}/respond/qcr?token={token}"
            html_body = f"""<div style="font-family:Segoe UI, Helvetica, Arial, sans-serif; color:#333;">
                <h2>{subject_prefix} - {item['identifier']}</h2>
                <p>Please complete your QC review: <a href="{review_url}">{review_url}</a></p>
            </div>"""
        
        # Send via Outlook - NO CC
        # Record BEFORE sending to prevent duplicates if Outlook blocks then releases the email
        record_reminder_sent(item_id, 'single_reviewer', item['qcr_email'], 'qcr', due_date.strftime('%Y-%m-%d'), reminder_stage)
        try:
            pythoncom.CoInitialize()
            outlook = win32com.client.Dispatch("Outlook.Application")
            mail = outlook.CreateItem(0)
            mail.Subject = subject
            mail.HTMLBody = html_body
            mail.To = item['qcr_email']
            # NO CC for reminder emails
            mail.Send()
            
            print(f"  [Reminder] Sent {reminder_stage} reminder to QCR for item {item_id}")
            return {'success': True}
        except Exception as e:
            # Note: reminder is already recorded to prevent duplicates even if send fails
            print(f"  [Reminder] Failed to send {reminder_stage} reminder to QCR for item {item_id}: {e}")
            return {'success': False, 'error': str(e)}
        finally:
            pythoncom.CoUninitialize()

def send_multi_reviewer_reminder_email(item, reviewer, role, due_date, reminder_stage):
    """Send a reminder email for a specific reviewer in multi-reviewer mode."""
    if not HAS_WIN32COM:
        return {'success': False, 'error': 'Outlook not available'}
    
    item_id = item['id']
    
    # Check if this reminder has already been sent
    if has_reminder_been_sent(item_id, reviewer['reviewer_email'], role, reminder_stage, reviewer['id']):
        return {'success': True, 'skipped': True, 'reason': 'Already sent'}
    
    # Determine subject prefix
    if reminder_stage == 'due_today':
        subject_prefix = "REMINDER: DUE TODAY"
    else:
        subject_prefix = "REMINDER: OVERDUE"
    
    subject = f"{subject_prefix} - [LEB] {item['identifier']}"
    
    # Priority color
    priority_color = '#e67e22' if item['priority'] == 'Medium' else '#c0392b' if item['priority'] == 'High' else '#27ae60'
    
    # Folder link
    folder_path = item['folder_link'] or 'Not set'
    use_file_form = is_local_mode() and folder_path != 'Not set'
    
    reviewer_due_date = format_date_for_email(item['initial_reviewer_due_date'])
    qcr_due_date = format_date_for_email(item['qcr_due_date'])
    contractor_due_date = format_date_for_email(item['due_date'])
    
    if use_file_form:
        form_result = generate_multi_reviewer_form(item_id, reviewer)
        if form_result['success']:
            form_file_path = form_result['path']
            form_file_link = f'file:///{form_file_path.replace(chr(92), "/")}'
            
            html_body = f"""<div style="font-family:Segoe UI, Helvetica, Arial, sans-serif; color:#333; font-size:14px; line-height:1.5;">

    <!-- REMINDER BANNER -->
    <div style="background: {'#dc3545' if reminder_stage == 'overdue' else '#ffc107'}; color: {'white' if reminder_stage == 'overdue' else '#212529'}; padding: 15px; border-radius: 8px; margin-bottom: 20px; text-align: center;">
        <strong style="font-size: 16px;">{"⚠️ OVERDUE - RESPONSE REQUIRED IMMEDIATELY" if reminder_stage == 'overdue' else "⏰ REMINDER - DUE TODAY"}</strong>
        <div style="margin-top: 8px; font-size: 14px;">Your response was due: {due_date.strftime('%B %d, %Y')}</div>
    </div>

    <!-- BLUEBEAM INSTRUCTIONS -->
    <div style="margin:15px 0; padding:15px; background:#dbeafe; border:1px solid #3b82f6; border-radius:8px;">
        <div style="font-size:14px; color:#1e40af; font-weight:bold;">📐 Markups Instructions</div>
        <div style="font-size:13px; color:#1e40af; margin-top:8px;">
            <strong>Provide markups in the corresponding Bluebeam session.</strong>
        </div>
    </div>

    <!-- HEADER -->
    <h2 style="color:#444; margin-bottom:6px;">
        [LEB] {item['identifier']} – Assigned to You
    </h2>

    <p style="margin-top:0; font-size:13px; color:#666;">
        This is a reminder that you have been assigned a review task that requires your response.
    </p>

    <!-- DIRECT LINK TO FORM -->
    <div style="margin:20px 0; text-align:center;">
        <table cellpadding="0" cellspacing="0" border="0" style="margin:0 auto;">
            <tr>
                <td align="center" bgcolor="#27ae60" style="background:#27ae60; border-radius:8px; padding:0;">
                    <a href="{form_file_link}" target="_blank"
                       style="background:#27ae60; color:#ffffff; display:inline-block; font-family:Segoe UI,Arial,sans-serif;
                              font-size:16px; font-weight:bold; line-height:50px; text-align:center; text-decoration:none;
                              width:320px; -webkit-text-size-adjust:none; border-radius:8px;">
                        OPEN RESPONSE FORM
                    </a>
                </td>
            </tr>
        </table>
    </div>

    <!-- INFO TABLE -->
    <table cellpadding="8" cellspacing="0" width="100%" style="border-collapse:collapse; margin-top:10px;">
        <tr><td colspan="2" style="background:#f2f2f2; font-weight:bold; border:1px solid #ddd;">Item Information</td></tr>
        <tr><td style="width:160px; border:1px solid #ddd; font-weight:bold;">Type</td><td style="border:1px solid #ddd;">{item['type']}</td></tr>
        <tr><td style="border:1px solid #ddd; font-weight:bold;">Identifier</td><td style="border:1px solid #ddd;">{item['identifier']}</td></tr>
        <tr><td style="border:1px solid #ddd; font-weight:bold;">Title</td><td style="border:1px solid #ddd;">{item['title'] or 'N/A'}</td></tr>
        <tr><td style="border:1px solid #ddd; font-weight:bold;">Priority</td><td style="border:1px solid #ddd; color:{priority_color}; font-weight:bold;">{item['priority'] or 'Normal'}</td></tr>
        <tr><td style="border:1px solid #ddd; font-weight:bold;">Your Due Date</td><td style="border:1px solid #ddd; color:#c0392b; font-weight:bold;">{reviewer_due_date}</td></tr>
        <tr><td style="border:1px solid #ddd; font-weight:bold;">QCR Due Date</td><td style="border:1px solid #ddd; color:#27ae60; font-weight:bold;">{qcr_due_date}</td></tr>
        <tr><td style="border:1px solid #ddd; font-weight:bold;">Contractor Due Date</td><td style="border:1px solid #ddd; color:#c0392b; font-weight:bold;">{contractor_due_date}</td></tr>
    </table>

    <p style="margin-top:20px; font-size:12px; color:#777;">
        <em>This is an automated reminder. Please submit your response as soon as possible.</em>
    </p>
</div>"""
        else:
            return {'success': False, 'error': f'Could not generate form: {form_result.get("error")}'}
    else:
        app_host = get_app_host()
        token = reviewer['email_token']
        review_url = f"{app_host}/respond/multi-reviewer?token={token}"
        html_body = f"""<div style="font-family:Segoe UI, Helvetica, Arial, sans-serif; color:#333;">
            <h2>{subject_prefix} - {item['identifier']}</h2>
            <p>Please submit your response: <a href="{review_url}">{review_url}</a></p>
        </div>"""
    
    # Send via Outlook - Only to this specific reviewer, NO CC
    # Record BEFORE sending to prevent duplicates if Outlook blocks then releases the email
    record_reminder_sent(item_id, 'multi_reviewer', reviewer['reviewer_email'], 'reviewer', due_date.strftime('%Y-%m-%d'), reminder_stage, reviewer['id'])
    try:
        pythoncom.CoInitialize()
        outlook = win32com.client.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)
        mail.Subject = subject
        mail.HTMLBody = html_body
        mail.To = reviewer['reviewer_email']
        # NO CC for reminder emails
        mail.Send()
        
        print(f"  [Reminder] Sent {reminder_stage} reminder to {reviewer['reviewer_name']} for item {item_id}")
        return {'success': True}
    except Exception as e:
        # Note: reminder is already recorded to prevent duplicates even if send fails
        print(f"  [Reminder] Failed to send {reminder_stage} reminder to {reviewer['reviewer_name']} for item {item_id}: {e}")
        return {'success': False, 'error': str(e)}
    finally:
        pythoncom.CoUninitialize()

def send_multi_reviewer_qcr_reminder_email(item, due_date, reminder_stage):
    """Send a reminder email to QCR in multi-reviewer mode."""
    if not HAS_WIN32COM:
        return {'success': False, 'error': 'Outlook not available'}
    
    item_id = item['id']
    
    # Check if this reminder has already been sent
    if has_reminder_been_sent(item_id, item['qcr_email'], 'qcr', reminder_stage):
        return {'success': True, 'skipped': True, 'reason': 'Already sent'}
    
    # Determine subject prefix
    if reminder_stage == 'due_today':
        subject_prefix = "REMINDER: DUE TODAY"
    else:
        subject_prefix = "REMINDER: OVERDUE"
    
    subject = f"{subject_prefix} - [LEB] {item['identifier']}"
    
    # Priority color
    priority_color = '#e67e22' if item['priority'] == 'Medium' else '#c0392b' if item['priority'] == 'High' else '#27ae60'
    
    # Folder link
    folder_path = item['folder_link'] or 'Not set'
    use_file_form = is_local_mode() and folder_path != 'Not set'
    
    qcr_due_date = item['qcr_due_date']
    
    if use_file_form:
        form_result = generate_multi_reviewer_qcr_form(item_id)
        if form_result['success']:
            form_file_path = form_result['path']
            form_file_link = f'file:///{form_file_path.replace(chr(92), "/")}'
            
            html_body = f"""<div style="font-family:Segoe UI, Helvetica, Arial, sans-serif; color:#333; font-size:14px; line-height:1.5;">

    <!-- REMINDER BANNER -->
    <div style="background: {'#dc3545' if reminder_stage == 'overdue' else '#ffc107'}; color: {'white' if reminder_stage == 'overdue' else '#212529'}; padding: 15px; border-radius: 8px; margin-bottom: 20px; text-align: center;">
        <strong style="font-size: 16px;">{"⚠️ OVERDUE - QC REVIEW REQUIRED IMMEDIATELY" if reminder_stage == 'overdue' else "⏰ REMINDER - QC REVIEW DUE TODAY"}</strong>
        <div style="margin-top: 8px; font-size: 14px;">Your QC review was due: {due_date.strftime('%B %d, %Y')}</div>
    </div>

    <!-- HEADER -->
    <h2 style="color:#444; margin-bottom:6px;">
        [LEB] {item['identifier']} – QC Review Ready
    </h2>

    <p style="margin-top:0; font-size:13px; color:#666;">
        This is a reminder that all reviewers have submitted and QC review is required.
    </p>

    <!-- DIRECT LINK TO FORM -->
    <div style="margin:20px 0; text-align:center;">
        <table cellpadding="0" cellspacing="0" border="0" style="margin:0 auto;">
            <tr>
                <td align="center" bgcolor="#27ae60" style="background:#27ae60; border-radius:8px; padding:0;">
                    <a href="{form_file_link}" target="_blank"
                       style="background:#27ae60; color:#ffffff; display:inline-block; font-family:Segoe UI,Arial,sans-serif;
                              font-size:16px; font-weight:bold; line-height:50px; text-align:center; text-decoration:none;
                              width:320px; -webkit-text-size-adjust:none; border-radius:8px;">
                        OPEN QC REVIEW FORM
                    </a>
                </td>
            </tr>
        </table>
    </div>

    <!-- INFO TABLE -->
    <table cellpadding="8" cellspacing="0" width="100%" style="border-collapse:collapse; margin-top:10px;">
        <tr><td colspan="2" style="background:#f2f2f2; font-weight:bold; border:1px solid #ddd;">Item Information</td></tr>
        <tr><td style="width:160px; border:1px solid #ddd; font-weight:bold;">Type</td><td style="border:1px solid #ddd;">{item['type']}</td></tr>
        <tr><td style="border:1px solid #ddd; font-weight:bold;">Identifier</td><td style="border:1px solid #ddd;">{item['identifier']}</td></tr>
        <tr><td style="border:1px solid #ddd; font-weight:bold;">Title</td><td style="border:1px solid #ddd;">{item['title'] or 'N/A'}</td></tr>
        <tr><td style="border:1px solid #ddd; font-weight:bold;">Priority</td><td style="border:1px solid #ddd; color:{priority_color}; font-weight:bold;">{item['priority'] or 'Normal'}</td></tr>
        <tr><td style="border:1px solid #ddd; font-weight:bold;">Your Due Date</td><td style="border:1px solid #ddd; color:#c0392b; font-weight:bold;">{qcr_due_date or 'N/A'}</td></tr>
        <tr><td style="border:1px solid #ddd; font-weight:bold;">Contractor Due Date</td><td style="border:1px solid #ddd; color:#c0392b; font-weight:bold;">{item['due_date'] or 'N/A'}</td></tr>
    </table>

    <p style="margin-top:20px; font-size:12px; color:#777;">
        <em>This is an automated reminder. Please complete your QC review as soon as possible.</em>
    </p>
</div>"""
        else:
            return {'success': False, 'error': f'Could not generate form: {form_result.get("error")}'}
    else:
        app_host = get_app_host()
        token = item['email_token_qcr']
        review_url = f"{app_host}/respond/multi-qcr?token={token}"
        html_body = f"""<div style="font-family:Segoe UI, Helvetica, Arial, sans-serif; color:#333;">
            <h2>{subject_prefix} - {item['identifier']}</h2>
            <p>Please complete your QC review: <a href="{review_url}">{review_url}</a></p>
        </div>"""
    
    # Send via Outlook - NO CC
    # Record BEFORE sending to prevent duplicates if Outlook blocks then releases the email
    record_reminder_sent(item_id, 'multi_reviewer', item['qcr_email'], 'qcr', due_date.strftime('%Y-%m-%d'), reminder_stage)
    try:
        pythoncom.CoInitialize()
        outlook = win32com.client.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)
        mail.Subject = subject
        mail.HTMLBody = html_body
        mail.To = item['qcr_email']
        # NO CC for reminder emails
        mail.Send()
        
        print(f"  [Reminder] Sent {reminder_stage} reminder to QCR for item {item_id}")
        return {'success': True}
    except Exception as e:
        # Note: reminder is already recorded to prevent duplicates even if send fails
        print(f"  [Reminder] Failed to send {reminder_stage} reminder to QCR for item {item_id}: {e}")
        return {'success': False, 'error': str(e)}
    finally:
        pythoncom.CoUninitialize()

def process_all_reminders():
    """Process all due/overdue reminders. Called by the reminder scheduler."""
    if not is_past_reminder_time_today():
        return {'processed': False, 'reason': 'Not yet reminder time (8 AM PST)'}
    
    print(f"  [Reminder] Processing reminders at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    items_needing_reminders = get_items_needing_reminders()
    
    results = {
        'single_reviewer_sent': 0,
        'single_reviewer_skipped': 0,
        'multi_reviewer_sent': 0,
        'multi_reviewer_skipped': 0,
        'multi_reviewer_qcr_sent': 0,
        'multi_reviewer_qcr_skipped': 0,
        'errors': []
    }
    
    # Process single reviewer reminders
    for item, role, due_date, reminder_stage in items_needing_reminders['single_reviewer']:
        try:
            result = send_single_reviewer_reminder_email(item, role, due_date, reminder_stage)
            if result.get('success'):
                if result.get('skipped'):
                    results['single_reviewer_skipped'] += 1
                else:
                    results['single_reviewer_sent'] += 1
            else:
                results['errors'].append(f"Item {item['id']} ({role}): {result.get('error')}")
        except Exception as e:
            results['errors'].append(f"Item {item['id']} ({role}): {str(e)}")
    
    # Process multi-reviewer individual reminders
    for item, reviewer, role, due_date, reminder_stage in items_needing_reminders['multi_reviewer']:
        try:
            result = send_multi_reviewer_reminder_email(item, reviewer, role, due_date, reminder_stage)
            if result.get('success'):
                if result.get('skipped'):
                    results['multi_reviewer_skipped'] += 1
                else:
                    results['multi_reviewer_sent'] += 1
            else:
                results['errors'].append(f"Item {item['id']} ({reviewer['reviewer_name']}): {result.get('error')}")
        except Exception as e:
            results['errors'].append(f"Item {item['id']} ({reviewer['reviewer_name']}): {str(e)}")
    
    # Process multi-reviewer QCR reminders
    for item, due_date, reminder_stage in items_needing_reminders['multi_reviewer_qcr']:
        try:
            result = send_multi_reviewer_qcr_reminder_email(item, due_date, reminder_stage)
            if result.get('success'):
                if result.get('skipped'):
                    results['multi_reviewer_qcr_skipped'] += 1
                else:
                    results['multi_reviewer_qcr_sent'] += 1
            else:
                results['errors'].append(f"Item {item['id']} (QCR): {result.get('error')}")
        except Exception as e:
            results['errors'].append(f"Item {item['id']} (QCR): {str(e)}")
    
    total_sent = results['single_reviewer_sent'] + results['multi_reviewer_sent'] + results['multi_reviewer_qcr_sent']
    if total_sent > 0:
        print(f"  [Reminder] Sent {total_sent} reminder(s)")
    
    return results


class ReminderScheduler:
    """Background scheduler for sending reminder emails at 8 AM PST."""
    
    def __init__(self, check_interval_seconds=300):  # Check every 5 minutes
        self.running = False
        self.thread = None
        self.interval = check_interval_seconds
        self.last_check = None
        self.last_reminder_date = None  # Track when we last processed reminders for the day
    
    def start(self):
        """Start the reminder scheduler thread."""
        self.running = True
        self.thread = threading.Thread(target=self._scheduler_loop, daemon=True)
        self.thread.start()
        print(f"Reminder scheduler started (checking every {self.interval}s)")
    
    def stop(self):
        """Stop the reminder scheduler thread."""
        self.running = False
        if self.thread:
            self.thread.join(timeout=5)
        print("Reminder scheduler stopped")
    
    def _scheduler_loop(self):
        """Main scheduler loop."""
        while self.running:
            try:
                self.last_check = datetime.now()
                pst_now = get_pst_now()
                today_pst = pst_now.date()
                
                # Only process reminders once per day, after 8 AM PST
                if is_past_reminder_time_today():
                    if self.last_reminder_date != today_pst:
                        # Process reminders for today
                        results = process_all_reminders()
                        self.last_reminder_date = today_pst
                        
                        if results.get('errors'):
                            for err in results['errors']:
                                print(f"  [Reminder] Error: {err}")
                
            except Exception as e:
                print(f"  [Reminder] Scheduler error: {e}")
            
            # Sleep in small increments so we can stop quickly
            for _ in range(self.interval):
                if not self.running:
                    break
                time.sleep(1)


# Global reminder scheduler instance
reminder_scheduler = ReminderScheduler(check_interval_seconds=300)

# =============================================================================
# OUTLOOK EMAIL POLLING
# =============================================================================

class EmailPoller:
    """Background email polling service."""
    
    def __init__(self):
        self.running = False
        self.thread = None
        self.last_poll = None
        self.poll_count = 0
        self.error_count = 0
        self.last_error = None
    
    def start(self):
        """Start the polling thread."""
        if not HAS_WIN32COM:
            print("Email polling disabled: pywin32 not available")
            return
        
        self.running = True
        self.thread = threading.Thread(target=self._poll_loop, daemon=True)
        self.thread.start()
        print("Email polling started")
    
    def stop(self):
        """Stop the polling thread."""
        self.running = False
        if self.thread:
            self.thread.join(timeout=5)
        print("Email polling stopped")
    
    def _poll_loop(self):
        """Main polling loop."""
        while self.running:
            try:
                self._poll_emails()
                self.last_poll = datetime.now()
                self.poll_count += 1
            except Exception as e:
                self.error_count += 1
                self.last_error = str(e)
                print(f"Email polling error: {e}")
            
            # Wait for next poll
            interval = CONFIG['poll_interval_minutes'] * 60
            for _ in range(interval):
                if not self.running:
                    break
                time.sleep(1)
    
    def _poll_emails(self):
        """Poll Outlook for new emails."""
        # Initialize COM for this thread
        pythoncom.CoInitialize()
        
        try:
            outlook = win32com.client.Dispatch("Outlook.Application")
            namespace = outlook.GetNamespace("MAPI")
            
            # Get the folder (Inbox or custom folder)
            folder_name = CONFIG.get('outlook_folder', 'Inbox')
            if folder_name.lower() == 'inbox':
                folder = namespace.GetDefaultFolder(6)  # 6 = Inbox
            else:
                # Try to find a subfolder
                inbox = namespace.GetDefaultFolder(6)
                try:
                    folder = inbox.Folders[folder_name]
                except:
                    folder = inbox  # Fall back to Inbox
            
            # Get messages
            messages = folder.Items
            messages.Sort("[ReceivedTime]", True)  # Sort by newest first
            
            conn = get_db()
            cursor = conn.cursor()
            
            processed_count = 0
            scanned_count = 0
            for message in messages:
                scanned_count += 1
                if scanned_count > 200:  # Limit to last 200 messages
                    break
                    
                try:
                    # Check if already processed
                    message_id = getattr(message, 'EntryID', None)
                    if not message_id:
                        continue
                    
                    cursor.execute('SELECT id FROM email_log WHERE entry_id = ?', (message_id,))
                    if cursor.fetchone():
                        continue  # Already processed
                    
                    # Check sender - only process emails from ACC
                    sender_email = ''
                    try:
                        sender = getattr(message, 'SenderEmailAddress', '') or ''
                        sender_email = sender.lower()
                        # Sometimes Outlook returns Exchange format, try to get SMTP address
                        if '@' not in sender_email:
                            try:
                                sender_email = message.Sender.GetExchangeUser().PrimarySmtpAddress.lower()
                            except:
                                pass
                    except:
                        pass
                    
                    # Only process emails from Autodesk Construction Cloud
                    if 'no-reply@acc.autodesk.com' not in sender_email and 'acc.autodesk.com' not in sender_email:
                        continue
                    
                    subject = getattr(message, 'Subject', '') or ''
                    body = getattr(message, 'Body', '') or ''
                    received_time = getattr(message, 'ReceivedTime', None)
                    
                    # Check if this is a relevant email
                    if 'LEB' not in subject.upper():
                        continue
                    
                    item_type = parse_item_type(subject)
                    if not item_type:
                        continue
                    
                    # Parse the email
                    identifier = parse_identifier(subject, item_type)
                    if not identifier:
                        continue
                    
                    bucket = determine_bucket(subject)
                    title = parse_title(subject, identifier, body)
                    due_date = parse_due_date(body)
                    priority = parse_priority(body)
                    
                    received_at = None
                    if received_time:
                        try:
                            received_at = datetime(
                                received_time.year, received_time.month, received_time.day,
                                received_time.hour, received_time.minute, received_time.second
                            ).isoformat()
                        except:
                            received_at = datetime.now().isoformat()
                    
                    # Check if item exists
                    cursor.execute('''
                        SELECT id, due_date, priority, title, status, closed_at,
                               reviewer_response_status, qcr_response_status,
                               initial_reviewer_id, qcr_id, folder_link
                        FROM item 
                        WHERE identifier = ? AND bucket = ?
                    ''', (identifier, bucket))
                    existing = cursor.fetchone()
                    
                    item_id = None
                    if existing:
                        # EXISTING ITEM - Check for updates from contractor
                        item_id = existing['id']
                        existing_due = existing['due_date']
                        existing_priority = existing['priority']
                        existing_title = existing['title']
                        current_status = existing['status']
                        was_closed = existing['closed_at'] is not None
                        
                        # Detect what changed
                        due_date_changed = due_date and existing_due and due_date != existing_due
                        priority_changed = priority and existing_priority and priority != existing_priority
                        title_changed = title and existing_title and title.strip() != existing_title.strip()
                        
                        # Check if this is a meaningful update (item is in workflow or was closed)
                        in_active_workflow = current_status in ('Assigned', 'In Review', 'In QC', 'Ready for Response')
                        
                        if (due_date_changed or priority_changed or title_changed) and (in_active_workflow or was_closed):
                            # This is a contractor update that needs admin attention
                            update_type = 'due_date_only' if (due_date_changed and not title_changed and not priority_changed) else 'content_change'
                            
                            # Log the update in history
                            cursor.execute('''
                                INSERT INTO item_update_history (
                                    item_id, update_type, 
                                    old_due_date, new_due_date,
                                    old_title, new_title,
                                    old_priority, new_priority,
                                    email_entry_id, email_subject, email_body_preview
                                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            ''', (
                                item_id, update_type,
                                existing_due, due_date,
                                existing_title, title if title_changed else None,
                                existing_priority, priority if priority_changed else None,
                                message_id, subject, body[:500]
                            ))
                            
                            # Flag the item for admin review
                            cursor.execute('''
                                UPDATE item SET
                                    has_pending_update = 1,
                                    update_type = ?,
                                    update_detected_at = ?,
                                    previous_due_date = ?,
                                    previous_title = ?,
                                    previous_priority = ?,
                                    update_email_body = ?,
                                    reopened_from_closed = ?,
                                    status_before_update = ?,
                                    last_email_at = ?
                                WHERE id = ?
                            ''', (
                                update_type, received_at,
                                existing_due if due_date_changed else None,
                                existing_title if title_changed else None,
                                existing_priority if priority_changed else None,
                                body[:1000],
                                1 if was_closed else 0,
                                current_status,
                                received_at,
                                item_id
                            ))
                            
                            # Create notification for admin
                            update_desc = []
                            if due_date_changed:
                                update_desc.append(f"Due Date: {existing_due} → {due_date}")
                            if title_changed:
                                update_desc.append("Title changed")
                            if priority_changed:
                                update_desc.append(f"Priority: {existing_priority} → {priority}")
                            
                            notification_msg = f"Contractor updated {item_type} {identifier}. Changes: {'; '.join(update_desc)}"
                            if was_closed:
                                notification_msg = f"⚠️ CLOSED ITEM UPDATED: {notification_msg}"
                            
                            create_notification(
                                'contractor_update',
                                f'🔄 Contractor Update: {identifier}',
                                notification_msg,
                                item_id=item_id,
                                action_url=f'/api/item/{item_id}/review-update',
                                action_label='Review Update'
                            )
                            
                            print(f"  [ACC Update] Detected update for {identifier}: {update_type}")
                        else:
                            # Normal update - just update last_email_at
                            updates = ['last_email_at = ?']
                            params = [received_at]
                            
                            # Only update due_date if currently empty
                            if not existing_due and due_date:
                                updates.append('due_date = ?')
                                params.append(due_date)
                            
                            # Only update priority if currently empty
                            if not existing_priority and priority:
                                updates.append('priority = ?')
                                params.append(priority)
                            
                            params.append(item_id)
                            cursor.execute(f'''
                                UPDATE item SET {', '.join(updates)} WHERE id = ?
                            ''', params)
                    else:
                        # Create new item
                        folder_link = create_item_folder(item_type, identifier, bucket, title)
                        
                        # Extract date_received from email received time
                        if received_at:
                            date_received = received_at[:10]  # Get YYYY-MM-DD part
                        else:
                            date_received = datetime.now().strftime('%Y-%m-%d')
                        
                        # Calculate review due dates if we have enough info
                        initial_reviewer_due = None
                        qcr_due = None
                        is_insufficient = 0
                        
                        if due_date and date_received:
                            due_dates = calculate_review_due_dates(date_received, due_date, priority)
                            initial_reviewer_due = due_dates['initial_reviewer_due_date']
                            qcr_due = due_dates['qcr_due_date']
                            is_insufficient = 1 if due_dates['is_contractor_window_insufficient'] else 0
                        
                        cursor.execute('''
                            INSERT INTO item (type, bucket, identifier, title, source_subject, 
                                            created_at, last_email_at, due_date, priority, folder_link,
                                            date_received, initial_reviewer_due_date, qcr_due_date, 
                                            is_contractor_window_insufficient, email_entry_id)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        ''', (item_type, bucket, identifier, title, subject,
                              received_at, received_at, due_date, priority, folder_link,
                              date_received, initial_reviewer_due, qcr_due, is_insufficient, message_id))
                        item_id = cursor.lastrowid
                    
                    # Log the email
                    cursor.execute('''
                        INSERT INTO email_log (item_id, message_id, entry_id, subject, body_preview, 
                                              received_at, raw_type, processed)
                        VALUES (?, ?, ?, ?, ?, ?, ?, 1)
                    ''', (item_id, message_id, message_id, subject, body[:500], 
                          received_at, item_type))
                    
                    processed_count += 1
                    
                except Exception as e:
                    print(f"Error processing message: {e}")
                    continue
            
            conn.commit()
            conn.close()
            
            print(f"Scanned {scanned_count} messages, processed {processed_count} new emails")
                
        finally:
            pythoncom.CoUninitialize()
    
    def get_status(self):
        """Get polling status."""
        return {
            'running': self.running,
            'last_poll': self.last_poll.isoformat() if self.last_poll else None,
            'poll_count': self.poll_count,
            'error_count': self.error_count,
            'last_error': self.last_error,
            'outlook_available': HAS_WIN32COM
        }

# Global email poller instance
email_poller = EmailPoller()

# =============================================================================
# EMAIL SENDING FOR WORKFLOW (Reviewer/QCR Assignment)
# =============================================================================

def generate_token():
    """Generate a secure random token for magic links."""
    return secrets.token_urlsafe(32)

def get_app_host():
    """Get the host URL for the application."""
    # When deployed to web, use the configured web host URL
    if CONFIG.get('deployment_mode') == 'web' and CONFIG.get('web_host_url'):
        return CONFIG['web_host_url'].rstrip('/')
    # For local development, use localhost
    return f"http://localhost:{CONFIG.get('server_port', 5000)}"

def is_local_mode():
    """Check if running in local mode (use file-based forms)."""
    return CONFIG.get('deployment_mode', 'local') == 'local'

def send_reviewer_assignment_email(item_id, is_revision=False, qcr_notes=None):
    """Send assignment email to the Initial Reviewer with magic link or file-based form.
    
    Args:
        item_id: The item ID
        is_revision: If True, this is a revision request from QCR
        qcr_notes: QCR's notes when sending back for revision
    """
    if not HAS_WIN32COM:
        return {'success': False, 'error': 'Outlook not available'}
    
    conn = get_db()
    cursor = conn.cursor()
    
    # Get item with reviewer info
    cursor.execute('''
        SELECT i.*, 
               ir.email as reviewer_email, ir.display_name as reviewer_name,
               qcr.email as qcr_email, qcr.display_name as qcr_name
        FROM item i
        LEFT JOIN user ir ON i.initial_reviewer_id = ir.id
        LEFT JOIN user qcr ON i.qcr_id = qcr.id
        WHERE i.id = ?
    ''', (item_id,))
    item = cursor.fetchone()
    
    if not item:
        conn.close()
        return {'success': False, 'error': 'Item not found'}
    
    if not item['reviewer_email']:
        conn.close()
        return {'success': False, 'error': 'No Initial Reviewer assigned'}
    
    # Calculate due dates if not already set (ensures consistency with app display)
    reviewer_due_date = item['initial_reviewer_due_date']
    qcr_due_date = item['qcr_due_date']
    
    if item['date_received'] and item['due_date']:
        calculated = calculate_review_due_dates(
            item['date_received'],
            item['due_date'],
            item['priority']
        )
        # Use calculated values if database values are missing
        if not reviewer_due_date:
            reviewer_due_date = calculated['initial_reviewer_due_date']
        if not qcr_due_date:
            qcr_due_date = calculated['qcr_due_date']
        
        # Also update database if values were missing
        if not item['initial_reviewer_due_date'] or not item['qcr_due_date']:
            cursor.execute('''
                UPDATE item SET 
                    initial_reviewer_due_date = COALESCE(initial_reviewer_due_date, ?),
                    qcr_due_date = COALESCE(qcr_due_date, ?)
                WHERE id = ?
            ''', (calculated['initial_reviewer_due_date'], calculated['qcr_due_date'], item_id))
    
    # Generate token if not exists
    token = item['email_token_reviewer']
    if not token:
        token = generate_token()
        cursor.execute('UPDATE item SET email_token_reviewer = ? WHERE id = ?', (token, item_id))
    
    conn.commit()
    
    # Priority color
    priority_color = '#e67e22' if item['priority'] == 'Medium' else '#c0392b' if item['priority'] == 'High' else '#27ae60'
    
    # Build email content - different subject for revision
    if is_revision:
        subject = f"[LEB] {item['identifier']} – REVISION REQUESTED"
    else:
        subject = f"[LEB] {item['identifier']} – Assigned to You"
    
    # Create clickable folder link
    folder_path = item['folder_link'] or 'Not set'
    if folder_path != 'Not set':
        folder_link_html = f'<a href="file:///{folder_path.replace(chr(92), "/")}" style="color:#0078D4; text-decoration:underline;">{folder_path}</a>'
    else:
        folder_link_html = 'Not set'
    
    # Determine if using file-based forms (local mode) or server-based
    use_file_form = is_local_mode() and folder_path != 'Not set'
    
    if use_file_form:
        # Generate the HTA form file in the item folder
        form_result = generate_reviewer_form_html(item_id)
        if form_result['success']:
            form_file_path = form_result['path']
            form_file_link = f'file:///{form_file_path.replace(chr(92), "/")}'
            
            # Build revision notice HTML if applicable
            if is_revision:
                revision_notice = f"""
    <!-- REVISION NOTICE -->
    <div style="margin:15px 0; padding:15px; background:#fff3cd; border:2px solid #ffc107; border-radius:8px;">
        <div style="font-size:15px; color:#856404; font-weight:bold;">
            ⚠️ REVISION REQUESTED BY QCR
        </div>
        <div style="font-size:13px; color:#856404; margin-top:8px;">
            <strong>{item['qcr_name'] or 'The QCR'}</strong> has reviewed your response and is requesting revisions.
        </div>
        {f'<div style="margin-top:10px; padding:10px; background:#fff8e1; border-radius:4px; font-size:13px;"><strong>QCR Notes:</strong><br/>{qcr_notes}</div>' if qcr_notes else ''}
    </div>
"""
                header_title = f"[LEB] {item['identifier']} - REVISION REQUESTED"
                header_text = "Your response has been sent back for revision. Please review the QCR's notes and submit an updated response."
            else:
                revision_notice = ""
                header_title = f"[LEB] {item['identifier']} - Assigned to You"
                header_text = "You have been assigned a new review task. Please review the details below."
            
            # Email content for file-based form - DIRECT LINK to HTA file
            html_body = f"""<div style="font-family:Segoe UI, Helvetica, Arial, sans-serif; color:#333; font-size:14px; line-height:1.5;">

    <!-- HEADER -->
    <h2 style="color:{'#c0392b' if is_revision else '#444'}; margin-bottom:6px;">
        {header_title}
    </h2>

    <p style="margin-top:0; font-size:13px; color:#666;">
        {header_text}
    </p>
{revision_notice}
    <!-- DIRECT LINK TO FORM - PROMINENT BUTTON -->
    <div style="margin:20px 0; text-align:center;">
        <table cellpadding="0" cellspacing="0" border="0" style="margin:0 auto;">
            <tr>
                <td align="center" bgcolor="#27ae60" style="background:#27ae60; border-radius:8px; padding:0;">
                    <a href="{form_file_link}" target="_blank"
                       style="background:#27ae60; color:#ffffff; display:inline-block; font-family:Segoe UI,Arial,sans-serif;
                              font-size:16px; font-weight:bold; line-height:50px; text-align:center; text-decoration:none;
                              width:320px; -webkit-text-size-adjust:none; border-radius:8px;">
                        {'SUBMIT REVISED RESPONSE' if is_revision else 'OPEN RESPONSE FORM'}
                    </a>
                </td>
            </tr>
        </table>
    </div>

    <!-- INSTRUCTIONS FOR HTA -->
    <div style="margin:20px 0; padding:15px; background:#e8f5e9; border:1px solid #4caf50; border-radius:8px;">
        <div style="font-size:14px; color:#2e7d32;">
            <strong>Instructions:</strong>
            <ol style="margin:8px 0 0 0; padding-left:20px;">
                <li>Click the green button above (or navigate to the item folder's <strong>Responses</strong> subfolder)</li>
                <li>Double-click <strong>_RESPONSE_FORM.hta</strong> to open it</li>
                <li>If prompted, select <strong>"Microsoft (R) HTML Application host"</strong> - choose <strong>Open</strong>, do NOT save the file</li>
                <li>Select your response category and any files to include</li>
                <li>Click <strong>Submit Response</strong> - your response will be saved automatically</li>
            </ol>
        </div>
    </div>

    <!-- INFO TABLE -->
    <table cellpadding="8" cellspacing="0" width="100%" style="border-collapse:collapse; margin-top:10px;">
        <tr>
            <td colspan="2" style="background:#f2f2f2; font-weight:bold; border:1px solid #ddd;">
                Item Information
            </td>
        </tr>

        <tr>
            <td style="width:160px; border:1px solid #ddd; font-weight:bold;">Type</td>
            <td style="border:1px solid #ddd;">{item['type']}</td>
        </tr>

        <tr>
            <td style="width:160px; border:1px solid #ddd; font-weight:bold;">Identifier</td>
            <td style="border:1px solid #ddd;">{item['identifier']}</td>
        </tr>

        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Title</td>
            <td style="border:1px solid #ddd;">{item['title'] or 'N/A'}</td>
        </tr>

        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Date Received</td>
            <td style="border:1px solid #ddd;">{item['date_received'] or 'N/A'}</td>
        </tr>

        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Priority</td>
            <td style="border:1px solid #ddd; color:{priority_color}; font-weight:bold;">{item['priority'] or 'Normal'}</td>
        </tr>

        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Initial Review Due Date</td>
            <td style="border:1px solid #ddd; color:#2980b9; font-weight:bold;">{reviewer_due_date or 'N/A'}</td>
        </tr>

        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">QCR Due Date</td>
            <td style="border:1px solid #ddd; color:#27ae60; font-weight:bold;">{qcr_due_date or 'N/A'}</td>
        </tr>

        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Contractor Due Date</td>
            <td style="border:1px solid #ddd; color:#c0392b; font-weight:bold;">{item['due_date'] or 'N/A'}</td>
        </tr>

        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Initial Reviewer</td>
            <td style="border:1px solid #ddd;">{item['reviewer_name'] or 'Not assigned'}</td>
        </tr>

        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">QC Reviewer</td>
            <td style="border:1px solid #ddd;">{item['qcr_name'] or 'Not assigned'}</td>
        </tr>
    </table>

    <!-- FILE PATH SECTION -->
    <div style="margin-top:18px;">
        <div style="font-weight:bold; margin-bottom:4px;">📁 Designated Folder:</div>
        <div style="padding:10px; border:1px solid #ddd; background:#fafafa; font-family:Consolas, monospace; font-size:12px; border-radius:4px;">
            {folder_link_html}
        </div>
    </div>

    <!-- CC NOTE -->
    <p style="margin-top:16px; font-size:13px; color:#555;">
        <strong>{item['qcr_name'] or 'The QCR'}</strong> has been CC'd so they are aware this review is in progress.
    </p>

    <!-- FOOTER -->
    <p style="margin-top:20px; font-size:12px; color:#777;">
        <em>This message was automatically generated. If you believe you received this by mistake, please contact the project administrator.</em>
    </p>

</div>"""
        else:
            # Fall back to server URL if form generation fails
            use_file_form = False
            print(f"Warning: Could not generate form file: {form_result.get('error')}")
    
    if not use_file_form:
        # Use server-based form (original behavior)
        server_url = f"{get_app_host()}/respond/reviewer?token={token}"
        
        if is_local_mode() and HAS_AIRTABLE:
            airtable_url = get_airtable_form_url('reviewer', dict(item), token)
            if airtable_url:
                respond_url = airtable_url
                button_text = "SUBMIT REVISED RESPONSE" if is_revision else "OPEN RESPONSE FORM"
            else:
                respond_url = server_url
                button_text = "SUBMIT REVISED RESPONSE" if is_revision else "OPEN RESPONSE FORM"
        else:
            respond_url = server_url
            button_text = "SUBMIT REVISED RESPONSE" if is_revision else "OPEN RESPONSE FORM"
        
        # Build revision notice for server-based email
        if is_revision:
            revision_notice_server = f"""
    <!-- REVISION NOTICE -->
    <div style="margin:15px 0; padding:15px; background:#fff3cd; border:2px solid #ffc107; border-radius:8px;">
        <div style="font-size:15px; color:#856404; font-weight:bold;">
            ⚠️ REVISION REQUESTED BY QCR
        </div>
        <div style="font-size:13px; color:#856404; margin-top:8px;">
            <strong>{item['qcr_name'] or 'The QCR'}</strong> has reviewed your response and is requesting revisions.
        </div>
        {f'<div style="margin-top:10px; padding:10px; background:#fff8e1; border-radius:4px; font-size:13px;"><strong>QCR Notes:</strong><br/>{qcr_notes}</div>' if qcr_notes else ''}
    </div>
"""
            header_title_server = f"[LEB] {item['identifier']} – REVISION REQUESTED"
            header_text_server = "Your response has been sent back for revision. Please review the QCR's notes and submit an updated response."
        else:
            revision_notice_server = ""
            header_title_server = f"[LEB] {item['identifier']} – Assigned to You"
            header_text_server = "You have been assigned a new review task. Please review the details below."
        
        html_body = f"""<div style="font-family:Segoe UI, Helvetica, Arial, sans-serif; color:#333; font-size:14px; line-height:1.5;">

    <!-- HEADER -->
    <h2 style="color:{'#c0392b' if is_revision else '#444'}; margin-bottom:6px;">
        {header_title_server}
    </h2>

    <p style="margin-top:0; font-size:13px; color:#666;">
        {header_text_server}
    </p>
{revision_notice_server}
    <!-- ACTION BUTTON - AT TOP -->
    <div style="margin:20px 0; text-align:center;">
        <table cellpadding="0" cellspacing="0" border="0" style="margin:0;">
            <tr>
                <td align="center" bgcolor="#27ae60" style="background:#27ae60; border-radius:8px; padding:0;">
                    <a href="{respond_url}" target="_blank"
                       style="background:#27ae60; color:#ffffff; display:inline-block; font-family:Segoe UI,Arial,sans-serif;
                              font-size:16px; font-weight:bold; line-height:50px; text-align:center; text-decoration:none;
                              width:280px; -webkit-text-size-adjust:none; border-radius:8px;">
                        {button_text}
                    </a>
                </td>
            </tr>
        </table>
    </div>

    <!-- INFO TABLE -->
    <table cellpadding="8" cellspacing="0" width="100%" style="border-collapse:collapse; margin-top:10px;">
        <tr>
            <td colspan="2" style="background:#f2f2f2; font-weight:bold; border:1px solid #ddd;">
                Item Information
            </td>
        </tr>

        <tr>
            <td style="width:160px; border:1px solid #ddd; font-weight:bold;">Type</td>
            <td style="border:1px solid #ddd;">{item['type']}</td>
        </tr>

        <tr>
            <td style="width:160px; border:1px solid #ddd; font-weight:bold;">Identifier</td>
            <td style="border:1px solid #ddd;">{item['identifier']}</td>
        </tr>

        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Title</td>
            <td style="border:1px solid #ddd;">{item['title'] or 'N/A'}</td>
        </tr>

        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Date Received</td>
            <td style="border:1px solid #ddd;">{item['date_received'] or 'N/A'}</td>
        </tr>

        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Priority</td>
            <td style="border:1px solid #ddd; color:{priority_color}; font-weight:bold;">{item['priority'] or 'Normal'}</td>
        </tr>

        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Initial Review Due Date</td>
            <td style="border:1px solid #ddd; color:#2980b9; font-weight:bold;">{reviewer_due_date or 'N/A'}</td>
        </tr>

        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">QCR Due Date</td>
            <td style="border:1px solid #ddd; color:#27ae60; font-weight:bold;">{qcr_due_date or 'N/A'}</td>
        </tr>

        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Contractor Due Date</td>
            <td style="border:1px solid #ddd; color:#c0392b; font-weight:bold;">{item['due_date'] or 'N/A'}</td>
        </tr>

        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Initial Reviewer</td>
            <td style="border:1px solid #ddd;">{item['reviewer_name'] or 'Not assigned'}</td>
        </tr>

        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">QC Reviewer</td>
            <td style="border:1px solid #ddd;">{item['qcr_name'] or 'Not assigned'}</td>
        </tr>
    </table>

    <!-- FILE PATH SECTION -->
    <div style="margin-top:18px;">
        <div style="font-weight:bold; margin-bottom:4px;">📁 Designated Folder:</div>
        <div style="padding:10px; border:1px solid #ddd; background:#fafafa; font-family:Consolas, monospace; font-size:12px; border-radius:4px;">
            {folder_link_html}
        </div>
        <p style="font-size:11px; color:#888; margin-top:4px;">Click to open folder location</p>
    </div>

    <!-- INSTRUCTIONS -->
    <div style="margin-top:18px; padding:12px; background:#f7f9fc; border-left:4px solid #0078D4;">
        <strong>In the form, you will:</strong>
        <ul style="margin-top:8px; padding-left:20px;">
            <li>Select a response category (Approved, Approved as Noted, For Record Only, Rejected, Revise/Resubmit)</li>
            <li>Select files from the designated folder to include in the official response</li>
            <li>Add notes/comments for QC and project records</li>
        </ul>
    </div>

    <!-- CC NOTE -->
    <p style="margin-top:16px; font-size:13px; color:#555;">
        <strong>{item['qcr_name'] or 'The QCR'}</strong> has been CC'd so they are aware this review is in progress.
    </p>

    <!--AIRTABLE_FALLBACK_PLACEHOLDER-->

    <!-- FOOTER -->
    <p style="margin-top:20px; font-size:12px; color:#777;">
        <em>This message was automatically generated. If you believe you received this by mistake, please contact the project administrator.</em>
    </p>

</div>"""
        
        # Generate Airtable fallback link only when deployed to web (not in local mode)
        airtable_fallback_section = ""
        if not is_local_mode() and HAS_AIRTABLE:
            try:
                airtable_url = get_airtable_form_url('reviewer', dict(item), token)
                if airtable_url:
                    airtable_fallback_section = f'''
    <!-- FALLBACK OPTION -->
    <div style="margin-top:24px; padding:16px; background:#fff8e6; border:1px solid #ffd666; border-radius:8px;">
        <div style="font-weight:bold; color:#d48806; margin-bottom:8px;">
            🌐 Server Unavailable?
        </div>
        <p style="font-size:13px; color:#666; margin-bottom:12px;">
            If the button above doesn't work, use this alternative form:
        </p>
        <a href="{airtable_url}"
           style="background:#ffd666; color:#333; padding:10px 20px; 
                  font-size:14px; font-weight:600; text-decoration:none; border-radius:6px; display:inline-block;">
            📝 Use Backup Form
        </a>
    </div>
'''
            except:
                pass
        
        html_body = html_body.replace('<!--AIRTABLE_FALLBACK_PLACEHOLDER-->', airtable_fallback_section)
    
    try:
        pythoncom.CoInitialize()
        try:
            outlook = win32com.client.Dispatch("Outlook.Application")
            mail = outlook.CreateItem(0)  # 0 = olMailItem
            mail.To = item['reviewer_email']
            if item['qcr_email']:
                mail.CC = item['qcr_email']
            mail.Subject = subject
            mail.HTMLBody = html_body
            mail.Send()
            
            # Update database
            cursor.execute('''
                UPDATE item SET 
                    reviewer_email_sent_at = ?,
                    reviewer_response_status = 'Email Sent',
                    status = 'Assigned'
                WHERE id = ?
            ''', (datetime.now().isoformat(), item_id))
            conn.commit()
            
        finally:
            pythoncom.CoUninitialize()
        
        conn.close()
        return {'success': True, 'message': 'Reviewer assignment email sent'}
        
    except Exception as e:
        conn.close()
        return {'success': False, 'error': str(e)}

# =============================================================================
# CONTRACTOR UPDATE NOTIFICATION EMAILS
# =============================================================================

def send_due_date_update_email(item_id, recipient_type, new_due_date, admin_note='', was_reopened=False):
    """Send email notification about a due date update from contractor.
    
    Args:
        item_id: The item ID
        recipient_type: 'reviewer' or 'qcr'
        new_due_date: The new due date from contractor
        admin_note: Optional note from admin about the update
        was_reopened: Whether the item was reopened from closed status
    """
    if not HAS_WIN32COM:
        return {'success': False, 'error': 'Outlook not available'}
    
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT i.*, 
               ir.email as reviewer_email, ir.display_name as reviewer_name,
               qcr.email as qcr_email, qcr.display_name as qcr_name
        FROM item i
        LEFT JOIN user ir ON i.initial_reviewer_id = ir.id
        LEFT JOIN user qcr ON i.qcr_id = qcr.id
        WHERE i.id = ?
    ''', (item_id,))
    item = cursor.fetchone()
    
    if not item:
        conn.close()
        return {'success': False, 'error': 'Item not found'}
    
    # Determine recipient
    if recipient_type == 'qcr':
        to_email = item['qcr_email']
        to_name = item['qcr_name']
        role_label = 'QCR'
        due_date_field = item['qcr_due_date']
    else:
        to_email = item['reviewer_email']
        to_name = item['reviewer_name']
        role_label = 'Initial Reviewer'
        due_date_field = item['initial_reviewer_due_date']
    
    if not to_email:
        conn.close()
        return {'success': False, 'error': f'No {recipient_type} email found'}
    
    conn.close()
    
    # Recalculate the appropriate due date based on new contractor due date
    if item['date_received'] and new_due_date:
        due_dates = calculate_review_due_dates(
            item['date_received'], new_due_date, item['priority']
        )
        if recipient_type == 'qcr':
            new_internal_due = due_dates['qcr_due_date']
        else:
            new_internal_due = due_dates['initial_reviewer_due_date']
    else:
        new_internal_due = due_date_field
    
    # Create folder link
    folder_path = item['folder_link'] or 'Not set'
    if folder_path != 'Not set':
        folder_link_html = f'<a href="file:///{folder_path.replace(chr(92), "/")}" style="color:#0078D4;">{folder_path}</a>'
    else:
        folder_link_html = 'Not set'
    
    # Build subject with REOPENED or UPDATED prefix
    if was_reopened:
        subject = f"[LEB] {item['identifier']} – REOPENED: Due Date Updated by Contractor"
        header_text = "📅🔄 ITEM REOPENED - Due Date Update"
        header_color = "#dc2626"
        subtext = "This item was <strong>previously closed</strong> but has been reopened by the contractor with a new due date."
    else:
        subject = f"[LEB] {item['identifier']} – UPDATED: Due Date Changed by Contractor"
        header_text = "📅 Due Date Update"
        header_color = "#2563eb"
        subtext = "The contractor has updated the due date for this item."
    
    # Build the email
    html_body = f"""<div style="font-family:Segoe UI, Helvetica, Arial, sans-serif; color:#333; font-size:14px; line-height:1.5;">

    <!-- HEADER -->
    <h2 style="color:{header_color}; margin-bottom:6px;">
        {header_text}: {item['identifier']}
    </h2>

    <p style="margin-top:0; font-size:13px; color:#666;">
        {subtext}
    </p>

    <!-- UPDATE NOTICE -->
    <div style="margin:15px 0; padding:15px; background:#dbeafe; border:2px solid #3b82f6; border-radius:8px;">
        <div style="font-size:14px; color:#1e40af;">
            <strong>📢 Due Date Changed</strong>
        </div>
        <div style="margin-top:10px; font-size:13px; color:#1e40af;">
            <table cellpadding="4" cellspacing="0" style="border-collapse:collapse;">
                <tr>
                    <td><strong>Previous Contractor Due Date:</strong></td>
                    <td style="text-decoration:line-through; color:#dc2626;">{item['previous_due_date'] or item['due_date'] or 'N/A'}</td>
                </tr>
                <tr>
                    <td><strong>New Contractor Due Date:</strong></td>
                    <td style="color:#059669; font-weight:bold;">{new_due_date}</td>
                </tr>
                <tr>
                    <td><strong>Your Updated Due Date:</strong></td>
                    <td style="color:#2563eb; font-weight:bold;">{format_date_for_email(new_internal_due)}</td>
                </tr>
            </table>
        </div>
    </div>
    
    {f'''<!-- ADMIN NOTE -->
    <div style="margin:15px 0; padding:15px; background:#fef3c7; border:2px solid #f59e0b; border-radius:8px;">
        <div style="font-size:14px; color:#92400e; font-weight:bold;">📋 Note from Administrator:</div>
        <div style="margin-top:8px; font-size:13px; color:#78350f;">{admin_note}</div>
    </div>''' if admin_note else ''}

    <p style="font-size:13px; color:#666;">
        Please adjust your timeline accordingly. Your review workflow continues as normal.
    </p>

    <!-- INFO TABLE -->
    <table cellpadding="8" cellspacing="0" width="100%" style="border-collapse:collapse; margin-top:15px;">
        <tr>
            <td colspan="2" style="background:#f2f2f2; font-weight:bold; border:1px solid #ddd;">
                Item Information
            </td>
        </tr>
        <tr>
            <td style="width:160px; border:1px solid #ddd; font-weight:bold;">Type</td>
            <td style="border:1px solid #ddd;">{item['type']}</td>
        </tr>
        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Identifier</td>
            <td style="border:1px solid #ddd;">{item['identifier']}</td>
        </tr>
        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Title</td>
            <td style="border:1px solid #ddd;">{item['title'] or 'N/A'}</td>
        </tr>
        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Your Role</td>
            <td style="border:1px solid #ddd;">{role_label}</td>
        </tr>
    </table>

    <!-- FOLDER LINK -->
    <div style="margin-top:18px;">
        <div style="font-weight:bold; margin-bottom:4px;">📁 Item Folder:</div>
        <div style="padding:10px; border:1px solid #ddd; background:#fafafa; font-family:Consolas, monospace; font-size:12px; border-radius:4px;">
            {folder_link_html}
        </div>
    </div>

    <!-- FOOTER -->
    <p style="margin-top:20px; font-size:12px; color:#777;">
        <em>This is an automated notification. If you have questions, please contact the project administrator.</em>
    </p>

</div>"""

    # Send email
    pythoncom.CoInitialize()
    try:
        outlook = win32com.client.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)
        mail.To = to_email
        mail.Subject = subject
        mail.HTMLBody = html_body
        mail.Send()
        
        return {'success': True, 'message': f'Due date update email sent to {recipient_type}'}
    except Exception as e:
        return {'success': False, 'error': str(e)}
    finally:
        pythoncom.CoUninitialize()


def send_workflow_restart_email(item_id, admin_note='', was_closed=False):
    """Send email to reviewer(s) notifying them the workflow is restarting due to content changes.
    
    Args:
        item_id: The item ID
        admin_note: Admin's note about what changed
        was_closed: Whether the item was previously closed
    """
    if not HAS_WIN32COM:
        return {'success': False, 'error': 'Outlook not available'}
    
    conn = get_db()
    cursor = conn.cursor()
    
    # Get item details
    cursor.execute('''
        SELECT i.*, 
               ir.email as reviewer_email, ir.display_name as reviewer_name,
               qcr.email as qcr_email, qcr.display_name as qcr_name
        FROM item i
        LEFT JOIN user ir ON i.initial_reviewer_id = ir.id
        LEFT JOIN user qcr ON i.qcr_id = qcr.id
        WHERE i.id = ?
    ''', (item_id,))
    item = cursor.fetchone()
    
    if not item:
        conn.close()
        return {'success': False, 'error': 'Item not found'}
    
    # Check if multi-reviewer mode
    is_multi = item['multi_reviewer_mode']
    
    if is_multi:
        # Get all reviewers
        cursor.execute('''
            SELECT reviewer_name, reviewer_email, email_token
            FROM item_reviewers
            WHERE item_id = ?
        ''', (item_id,))
        reviewers = cursor.fetchall()
    else:
        # Single reviewer
        reviewers = [{
            'reviewer_name': item['reviewer_name'],
            'reviewer_email': item['reviewer_email'],
            'email_token': item['email_token_reviewer']
        }] if item['reviewer_email'] else []
    
    if not reviewers:
        conn.close()
        return {'success': False, 'error': 'No reviewers found'}
    
    conn.close()
    
    # Create folder link
    folder_path = item['folder_link'] or 'Not set'
    if folder_path != 'Not set':
        folder_link_html = f'<a href="file:///{folder_path.replace(chr(92), "/")}" style="color:#0078D4;">{folder_path}</a>'
    else:
        folder_link_html = 'Not set'
    
    # Build status message
    if was_closed:
        status_msg = "This item was previously <strong>CLOSED</strong> but has been reopened due to contractor changes."
        header_color = "#dc2626"
        icon = "🔄⚠️"
    else:
        status_msg = "The contractor has made changes to this item that require a fresh review."
        header_color = "#f59e0b"
        icon = "🔄"
    
    subject = f"[LEB] {item['identifier']} – {'REOPENED: ' if was_closed else ''}Review Restart Required"
    
    # Priority color
    priority_color = '#e67e22' if item['priority'] == 'Medium' else '#c0392b' if item['priority'] == 'High' else '#27ae60'
    
    emails_sent = 0
    errors = []
    
    pythoncom.CoInitialize()
    try:
        outlook = win32com.client.Dispatch("Outlook.Application")
        
        for reviewer in reviewers:
            if not reviewer['reviewer_email']:
                continue
            
            # Generate form file path for link
            form_file_link = ''
            if is_local_mode() and folder_path != 'Not set':
                form_result = generate_reviewer_form_html(item_id)
                if form_result['success']:
                    form_file_link = f'file:///{form_result["path"].replace(chr(92), "/")}'
            
            html_body = f"""<div style="font-family:Segoe UI, Helvetica, Arial, sans-serif; color:#333; font-size:14px; line-height:1.5;">

    <!-- HEADER -->
    <h2 style="color:{header_color}; margin-bottom:6px;">
        {icon} {'ITEM REOPENED' if was_closed else 'REVIEW RESTART'}: {item['identifier']}
    </h2>

    <p style="margin-top:0; font-size:13px; color:#666;">
        {status_msg}
    </p>

    <!-- CHANGE NOTICE -->
    <div style="margin:15px 0; padding:15px; background:#fef3c7; border:2px solid #f59e0b; border-radius:8px;">
        <div style="font-size:15px; color:#92400e; font-weight:bold;">
            {'⚠️ ITEM REOPENED - NEW REVIEW REQUIRED' if was_closed else '⚠️ CONTRACTOR CONTENT CHANGE'}
        </div>
        <div style="font-size:13px; color:#92400e; margin-top:8px;">
            Your previous response has been cleared. Please review the updated materials and submit a new response.
        </div>
    </div>
    
    {f'''<!-- ADMIN NOTE -->
    <div style="margin:15px 0; padding:15px; background:#e0e7ff; border:2px solid #4f46e5; border-radius:8px;">
        <div style="font-size:14px; color:#3730a3; font-weight:bold;">📋 What Changed (Note from Administrator):</div>
        <div style="margin-top:8px; font-size:13px; color:#312e81;">{admin_note}</div>
    </div>''' if admin_note else '''<!-- NO ADMIN NOTE PROVIDED -->
    <div style="margin:10px 0; padding:10px; background:#fef2f2; border:1px solid #fecaca; border-radius:8px;">
        <div style="font-size:13px; color:#991b1b;">No specific changes were noted by the administrator. Please review the updated materials carefully.</div>
    </div>'''}
    </div>

    <!-- ACTION BUTTON -->
    {f'''<div style="margin:20px 0; text-align:center;">
        <table cellpadding="0" cellspacing="0" border="0" style="margin:0 auto;">
            <tr>
                <td align="center" bgcolor="#dc2626" style="background:#dc2626; border-radius:8px; padding:0;">
                    <a href="{form_file_link}" target="_blank"
                       style="background:#dc2626; color:#ffffff; display:inline-block; font-family:Segoe UI,Arial,sans-serif;
                              font-size:16px; font-weight:bold; line-height:50px; text-align:center; text-decoration:none;
                              width:280px; -webkit-text-size-adjust:none; border-radius:8px;">
                        OPEN NEW RESPONSE FORM
                    </a>
                </td>
            </tr>
        </table>
    </div>''' if form_file_link else ''}

    <!-- INFO TABLE -->
    <table cellpadding="8" cellspacing="0" width="100%" style="border-collapse:collapse; margin-top:15px;">
        <tr>
            <td colspan="2" style="background:#f2f2f2; font-weight:bold; border:1px solid #ddd;">
                Item Information
            </td>
        </tr>
        <tr>
            <td style="width:160px; border:1px solid #ddd; font-weight:bold;">Type</td>
            <td style="border:1px solid #ddd;">{item['type']}</td>
        </tr>
        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Identifier</td>
            <td style="border:1px solid #ddd;">{item['identifier']}</td>
        </tr>
        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Title</td>
            <td style="border:1px solid #ddd;">{item['title'] or 'N/A'}</td>
        </tr>
        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Priority</td>
            <td style="border:1px solid #ddd; color:{priority_color}; font-weight:bold;">{item['priority'] or 'Normal'}</td>
        </tr>
        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Your Due Date</td>
            <td style="border:1px solid #ddd; color:#2980b9; font-weight:bold;">{format_date_for_email(item['initial_reviewer_due_date'])}</td>
        </tr>
        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Contractor Due Date</td>
            <td style="border:1px solid #ddd; color:#c0392b; font-weight:bold;">{format_date_for_email(item['due_date'])}</td>
        </tr>
    </table>

    <!-- FOLDER LINK -->
    <div style="margin-top:18px;">
        <div style="font-weight:bold; margin-bottom:4px;">📁 Item Folder (review updated materials here):</div>
        <div style="padding:10px; border:1px solid #ddd; background:#fafafa; font-family:Consolas, monospace; font-size:12px; border-radius:4px;">
            {folder_link_html}
        </div>
    </div>

    <!-- FOOTER -->
    <p style="margin-top:20px; font-size:12px; color:#777;">
        <em>This is an automated notification. Please contact the project administrator if you have questions about the changes.</em>
    </p>

</div>"""

            try:
                mail = outlook.CreateItem(0)
                mail.To = reviewer['reviewer_email']
                if item['qcr_email']:
                    mail.CC = item['qcr_email']
                mail.Subject = subject
                mail.HTMLBody = html_body
                mail.Send()
                emails_sent += 1
            except Exception as e:
                errors.append(f"{reviewer['reviewer_name']}: {str(e)}")
        
        # Update item to show email was sent for restart
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('''
            UPDATE item SET 
                reviewer_email_sent_at = ?,
                reviewer_response_status = 'Email Sent'
            WHERE id = ?
        ''', (datetime.now().isoformat(), item_id))
        conn.commit()
        conn.close()
        
        if errors:
            return {'success': False, 'error': f'Sent {emails_sent}, failed: {"; ".join(errors)}'}
        
        return {'success': True, 'message': f'Workflow restart email sent to {emails_sent} reviewer(s)'}
    except Exception as e:
        return {'success': False, 'error': str(e)}
    finally:
        pythoncom.CoUninitialize()


def send_qcr_assignment_email(item_id, is_revision=False, version=None):
    """Send assignment email to the QCR with magic link or file-based form."""
    if not HAS_WIN32COM:
        return {'success': False, 'error': 'Outlook not available'}
    
    conn = get_db()
    cursor = conn.cursor()
    
    # Get item with reviewer info
    cursor.execute('''
        SELECT i.*, 
               ir.email as reviewer_email, ir.display_name as reviewer_name,
               qcr.email as qcr_email, qcr.display_name as qcr_name
        FROM item i
        LEFT JOIN user ir ON i.initial_reviewer_id = ir.id
        LEFT JOIN user qcr ON i.qcr_id = qcr.id
        WHERE i.id = ?
    ''', (item_id,))
    item = cursor.fetchone()
    
    if not item:
        conn.close()
        return {'success': False, 'error': 'Item not found'}
    
    if not item['qcr_email']:
        conn.close()
        return {'success': False, 'error': 'No QCR assigned'}
    
    # Calculate due dates if not already set (ensures consistency with app display)
    qcr_due_date_email = item['qcr_due_date']
    
    if item['date_received'] and item['due_date']:
        calculated = calculate_review_due_dates(
            item['date_received'],
            item['due_date'],
            item['priority']
        )
        # Use calculated value if database value is missing
        if not qcr_due_date_email:
            qcr_due_date_email = calculated['qcr_due_date']
        
        # Also update database if value was missing
        if not item['qcr_due_date']:
            cursor.execute('''
                UPDATE item SET qcr_due_date = ? WHERE id = ?
            ''', (calculated['qcr_due_date'], item_id))
    
    # Generate token if not exists
    token = item['email_token_qcr']
    if not token:
        token = generate_token()
        cursor.execute('UPDATE item SET email_token_qcr = ? WHERE id = ?', (token, item_id))
    
    conn.commit()
    
    # Get version info
    current_version = version if version is not None else (item['reviewer_response_version'] if item['reviewer_response_version'] is not None else 0)
    
    # Get version history for display
    cursor.execute('''
        SELECT version, submitted_at 
        FROM reviewer_response_history 
        WHERE item_id = ? 
        ORDER BY version DESC
        LIMIT 5
    ''', (item_id,))
    history = cursor.fetchall()
    version_history_html = ''
    if history:
        history_parts = [f"v{h['version']} ({h['submitted_at'][:16].replace('T', ' ')})" for h in history]
        version_history_html = f"<p style='font-size: 12px; color: #666;'><strong>Previous versions:</strong> {', '.join(history_parts)}</p>"
    
    # Format reviewer response time
    reviewer_response_time = item['reviewer_response_at'] or 'N/A'
    if reviewer_response_time != 'N/A':
        try:
            dt = datetime.fromisoformat(reviewer_response_time.replace('Z', '+00:00'))
            reviewer_response_time = dt.strftime('%Y-%m-%d %H:%M')
        except:
            pass
    
    # Format reviewer selected files
    reviewer_files_display = 'None selected'
    reviewer_files_list = []
    if item['reviewer_selected_files']:
        try:
            reviewer_files_list = json.loads(item['reviewer_selected_files'])
            if reviewer_files_list:
                reviewer_files_display = '<br>'.join([f"• {f}" for f in reviewer_files_list])
        except:
            reviewer_files_display = item['reviewer_selected_files']
    
    # Format reviewer description (external notes)
    reviewer_notes_html = (item['reviewer_notes'] or 'No description provided').replace('\n', '<br>')
    
    # Format reviewer internal notes (for team visibility only)
    reviewer_internal_notes_html = ''
    if item['reviewer_internal_notes']:
        reviewer_internal_notes_html = item['reviewer_internal_notes'].replace('\n', '<br>')
    
    # Priority color
    priority_color = '#e67e22' if item['priority'] == 'Medium' else '#c0392b' if item['priority'] == 'High' else '#27ae60'
    
    # Build subject and intro based on whether this is a revision
    if is_revision:
        subject = f"[LEB] {item['identifier']} – Ready for Your Review (v{current_version})"
        intro_text = f"<strong style='color: #f59e0b;'>📝 Revision v{current_version}</strong> - The Initial Reviewer has submitted an updated response after your feedback. Please complete a new QC review."
    else:
        subject = f"[LEB] {item['identifier']} – Ready for Your Review"
        intro_text = "The Initial Reviewer has submitted their response. Please complete the QC review."
    
    # Create clickable folder link for QCR email
    folder_path = item['folder_link'] or 'Not set'
    if folder_path != 'Not set':
        folder_link_html = f'<a href="file:///{folder_path.replace(chr(92), "/")}" style="color:#27ae60; text-decoration:underline;">{folder_path}</a>'
    else:
        folder_link_html = 'Not set'
    
    # Determine if using file-based forms (local mode) or server-based
    use_file_form = is_local_mode() and folder_path != 'Not set'
    
    if use_file_form:
        # Generate the HTA form file in the item folder
        form_result = generate_qcr_form_html(item_id)
        if form_result['success']:
            form_file_path = form_result['path']
            form_file_link = f'file:///{form_file_path.replace(chr(92), "/")}'
            
            # Email content for file-based form - DIRECT LINK to HTA file
            html_body = f"""<div style="font-family:Segoe UI, Helvetica, Arial, sans-serif; color:#333; font-size:14px; line-height:1.5;">

    <!-- HEADER -->
    <h2 style="color:#444; margin-bottom:6px;">
        [LEB] {item['identifier']} - Ready for Your Review {f'(v{current_version})' if is_revision else ''}
    </h2>

    <p style="margin-top:0; font-size:13px; color:#666;">
        {intro_text}
    </p>

    <!-- DIRECT LINK TO QCR FORM - PROMINENT BUTTON -->
    <div style="margin:20px 0; text-align:center;">
        <table cellpadding="0" cellspacing="0" border="0" style="margin:0 auto;">
            <tr>
                <td align="center" bgcolor="#11998e" style="background:#11998e; border-radius:8px; padding:0;">
                    <a href="{form_file_link}" target="_blank"
                       style="background:#11998e; color:#ffffff; display:inline-block; font-family:Segoe UI,Arial,sans-serif;
                              font-size:16px; font-weight:bold; line-height:50px; text-align:center; text-decoration:none;
                              width:320px; -webkit-text-size-adjust:none; border-radius:8px;">
                        OPEN QC REVIEW FORM
                    </a>
                </td>
            </tr>
        </table>
    </div>

    <!-- INSTRUCTIONS FOR HTA -->
    <div style="margin:20px 0; padding:15px; background:#e8f5e9; border:1px solid #4caf50; border-radius:8px;">
        <div style="font-size:14px; color:#2e7d32;">
            <strong>Instructions:</strong>
            <ol style="margin:8px 0 0 0; padding-left:20px;">
                <li>Click the green button above (or navigate to the item folder's <strong>Responses</strong> subfolder)</li>
                <li>Double-click <strong>_QCR_RESPONSE_FORM.hta</strong> to open it</li>
                <li>If prompted, select <strong>"Microsoft (R) HTML Application host"</strong> - choose <strong>Open</strong>, do NOT save the file</li>
                <li>Review the response, make your decision, and click <strong>Submit</strong></li>
            </ol>
        </div>
    </div>

    <!-- INFO TABLE -->
    <table cellpadding="8" cellspacing="0" width="100%" style="border-collapse:collapse; margin-top:10px;">
        <tr>
            <td colspan="2" style="background:#f2f2f2; font-weight:bold; border:1px solid #ddd;">
                Item Information
            </td>
        </tr>

        <tr>
            <td style="width:160px; border:1px solid #ddd; font-weight:bold;">Type</td>
            <td style="border:1px solid #ddd;">{item['type']}
        </tr>

        <tr>
            <td style="width:160px; border:1px solid #ddd; font-weight:bold;">Identifier</td>
            <td style="border:1px solid #ddd;">{item['identifier']}</td>
        </tr>

        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Title</td>
            <td style="border:1px solid #ddd;">{item['title'] or 'N/A'}</td>
        </tr>

        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Date Received</td>
            <td style="border:1px solid #ddd;">{item['date_received'] or 'N/A'}</td>
        </tr>

        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Priority</td>
            <td style="border:1px solid #ddd; color:{priority_color}; font-weight:bold;">{item['priority'] or 'Normal'}</td>
        </tr>

        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Initial Reviewer</td>
            <td style="border:1px solid #ddd;">{item['reviewer_name'] or 'N/A'}</td>
        </tr>

        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">QC Reviewer</td>
            <td style="border:1px solid #ddd;">{item['qcr_name'] or 'N/A'}</td>
        </tr>

        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">QC Due Date</td>
            <td style="border:1px solid #ddd; color:#2980b9; font-weight:bold;">{qcr_due_date_email or 'N/A'}</td>
        </tr>

        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Contractor Due Date</td>
            <td style="border:1px solid #ddd; color:#c0392b; font-weight:bold;">{item['due_date'] or 'N/A'}</td>
        </tr>
    </table>

    <!-- REVIEWER RESPONSE SECTION -->
    <table cellpadding="8" cellspacing="0" width="100%" style="border-collapse:collapse; margin-top:16px;">
        <tr>
            <td colspan="2" style="background:#d5f5e3; font-weight:bold; border:1px solid #82e0aa; color:#1e8449;">
                ✅ Reviewer's Submitted Response (v{current_version})
            </td>
        </tr>

        <tr>
            <td style="width:160px; border:1px solid #ddd; font-weight:bold;">Category</td>
            <td style="border:1px solid #ddd; color:#1e8449; font-weight:bold;">{item['reviewer_response_category'] or 'Not specified'}</td>
        </tr>

        <tr>
            <td style="border:1px solid #ddd; font-weight:bold; vertical-align:top;">Selected Files</td>
            <td style="border:1px solid #ddd;">{reviewer_files_display}</td>
        </tr>

        <tr>
            <td style="border:1px solid #ddd; font-weight:bold; vertical-align:top;">Description</td>
            <td style="border:1px solid #ddd;">{reviewer_notes_html}</td>
        </tr>

        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Responded At</td>
            <td style="border:1px solid #ddd;">{reviewer_response_time}</td>
        </tr>
    </table>

    {f'''<!-- INTERNAL NOTES SECTION (Team Only) -->
    <div style="margin-top:16px; padding:12px; background:#fff8e6; border:1px solid #ffd966; border-radius:6px;">
        <div style="font-weight:bold; color:#b7791f; margin-bottom:6px;">🔒 Reviewer's Internal Notes (Team Only)</div>
        <div style="color:#744210; font-size:13px;">{reviewer_internal_notes_html}</div>
    </div>''' if reviewer_internal_notes_html else ''}

    {version_history_html}

    <!-- FILE PATH SECTION -->
    <div style="margin-top:18px;">
        <div style="font-weight:bold; margin-bottom:4px;">📁 Designated Folder:</div>
        <div style="padding:10px; border:1px solid #ddd; background:#fafafa; font-family:Consolas, monospace; font-size:12px; border-radius:4px;">
            {folder_link_html}
        </div>
    </div>

    <!-- CC NOTE -->
    <p style="margin-top:16px; font-size:13px; color:#555;">
        <strong>{item['reviewer_name'] or 'The Initial Reviewer'}</strong> has been CC'd on this email for visibility.
    </p>

    <!-- FOOTER -->
    <p style="margin-top:20px; font-size:12px; color:#777;">
        <em>This message was automatically generated. If you believe you received this by mistake, please contact the project administrator.</em>
    </p>

</div>"""
        else:
            # Fall back to server URL if form generation fails
            use_file_form = False
            print(f"Warning: Could not generate QCR form file: {form_result.get('error')}")
    
    if not use_file_form:
        # Use server-based form (original behavior)
        server_url = f"{get_app_host()}/respond/qcr?token={token}"
        
        if is_local_mode() and HAS_AIRTABLE:
            airtable_url = get_airtable_form_url('qcr', dict(item), token)
            if airtable_url:
                respond_url = airtable_url
                qcr_button_text = "OPEN QC REVIEW FORM"
            else:
                respond_url = server_url
                qcr_button_text = "OPEN QC REVIEW FORM"
        else:
            respond_url = server_url
            qcr_button_text = "OPEN QC REVIEW FORM"
        
        html_body = f"""<div style="font-family:Segoe UI, Helvetica, Arial, sans-serif; color:#333; font-size:14px; line-height:1.5;">

    <!-- HEADER -->
    <h2 style="color:#444; margin-bottom:6px;">
        [LEB] {item['identifier']} – Ready for Your Review {f'(v{current_version})' if is_revision else ''}
    </h2>

    <p style="margin-top:0; font-size:13px; color:#666;">
        {intro_text}
    </p>

    <!-- ACTION BUTTON - AT TOP -->
    <div style="margin:20px 0; text-align:center;">
        <table cellpadding="0" cellspacing="0" border="0" style="margin:0;">
            <tr>
                <td align="center" bgcolor="#27ae60" style="background:#27ae60; border-radius:8px; padding:0;">
                    <a href="{respond_url}" target="_blank"
                       style="background:#27ae60; color:#ffffff; display:inline-block; font-family:Segoe UI,Arial,sans-serif;
                              font-size:16px; font-weight:bold; line-height:50px; text-align:center; text-decoration:none;
                              width:280px; -webkit-text-size-adjust:none; border-radius:8px;">
                        {qcr_button_text}
                    </a>
                </td>
            </tr>
        </table>
    </div>

    <!-- INFO TABLE -->
    <table cellpadding="8" cellspacing="0" width="100%" style="border-collapse:collapse; margin-top:10px;">
        <tr>
            <td colspan="2" style="background:#f2f2f2; font-weight:bold; border:1px solid #ddd;">
                Item Information
            </td>
        </tr>

        <tr>
            <td style="width:160px; border:1px solid #ddd; font-weight:bold;">Type</td>
            <td style="border:1px solid #ddd;">{item['type']}
        </tr>

        <tr>
            <td style="width:160px; border:1px solid #ddd; font-weight:bold;">Identifier</td>
            <td style="border:1px solid #ddd;">{item['identifier']}</td>
        </tr>

        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Title</td>
            <td style="border:1px solid #ddd;">{item['title'] or 'N/A'}</td>
        </tr>

        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Date Received</td>
            <td style="border:1px solid #ddd;">{item['date_received'] or 'N/A'}</td>
        </tr>

        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Priority</td>
            <td style="border:1px solid #ddd; color:{priority_color}; font-weight:bold;">{item['priority'] or 'Normal'}</td>
        </tr>

        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Initial Reviewer</td>
            <td style="border:1px solid #ddd;">{item['reviewer_name'] or 'N/A'}</td>
        </tr>

        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">QC Reviewer</td>
            <td style="border:1px solid #ddd;">{item['qcr_name'] or 'N/A'}</td>
        </tr>

        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">QC Due Date</td>
            <td style="border:1px solid #ddd; color:#2980b9; font-weight:bold;">{qcr_due_date_email or 'N/A'}</td>
        </tr>

        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Contractor Due Date</td>
            <td style="border:1px solid #ddd; color:#c0392b; font-weight:bold;">{item['due_date'] or 'N/A'}</td>
        </tr>
    </table>

    <!-- REVIEWER RESPONSE SECTION -->
    <table cellpadding="8" cellspacing="0" width="100%" style="border-collapse:collapse; margin-top:16px;">
        <tr>
            <td colspan="2" style="background:#d5f5e3; font-weight:bold; border:1px solid #82e0aa; color:#1e8449;">
                ✅ Reviewer's Submitted Response (v{current_version})
            </td>
        </tr>

        <tr>
            <td style="width:160px; border:1px solid #ddd; font-weight:bold;">Category</td>
            <td style="border:1px solid #ddd; color:#1e8449; font-weight:bold;">{item['reviewer_response_category'] or 'Not specified'}</td>
        </tr>

        <tr>
            <td style="border:1px solid #ddd; font-weight:bold; vertical-align:top;">Selected Files</td>
            <td style="border:1px solid #ddd;">{reviewer_files_display}</td>
        </tr>

        <tr>
            <td style="border:1px solid #ddd; font-weight:bold; vertical-align:top;">Description</td>
            <td style="border:1px solid #ddd;">{reviewer_notes_html}</td>
        </tr>

        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Responded At</td>
            <td style="border:1px solid #ddd;">{reviewer_response_time}</td>
        </tr>
    </table>

    {f'''<!-- INTERNAL NOTES SECTION (Team Only) -->
    <div style="margin-top:16px; padding:12px; background:#fff8e6; border:1px solid #ffd966; border-radius:6px;">
        <div style="font-weight:bold; color:#b7791f; margin-bottom:6px;">🔒 Reviewer's Internal Notes (Team Only)</div>
        <div style="color:#744210; font-size:13px;">{reviewer_internal_notes_html}</div>
    </div>''' if reviewer_internal_notes_html else ''}

    {version_history_html}

    <!-- FILE PATH SECTION -->
    <div style="margin-top:18px;">
        <div style="font-weight:bold; margin-bottom:4px;">📁 Designated Folder:</div>
        <div style="padding:10px; border:1px solid #ddd; background:#fafafa; font-family:Consolas, monospace; font-size:12px; border-radius:4px;">
            {folder_link_html}
        </div>
        <p style="font-size:11px; color:#888; margin-top:4px;">Click to open folder location</p>
    </div>

    <!-- INSTRUCTIONS -->
    <div style="margin-top:18px; padding:12px; background:#f7f9fc; border-left:4px solid #27ae60;">
        <strong>In the QC form, you will:</strong>
        <ul style="margin-top:8px; padding-left:20px;">
            <li>Review the Initial Reviewer's response and files</li>
            <li>Choose an action: <strong>Approve</strong>, <strong>Modify</strong>, or <strong>Send Back</strong></li>
            <li>Decide how to handle the response text: Keep as is, Tweak, or Revise</li>
            <li>Confirm or adjust file selection for the official response</li>
            <li>Add any QC notes or conditions</li>
        </ul>
    </div>

    <!-- CC NOTE -->
    <p style="margin-top:16px; font-size:13px; color:#555;">
        <strong>{item['reviewer_name'] or 'The Initial Reviewer'}</strong> has been CC'd on this email for visibility.
    </p>

    <!--QCR_AIRTABLE_FALLBACK_PLACEHOLDER-->

    <!-- FOOTER -->
    <p style="margin-top:20px; font-size:12px; color:#777;">
        <em>This message was automatically generated. If you believe you received this by mistake, please contact the project administrator.</em>
    </p>

</div>"""
        
        # Generate Airtable fallback link only when deployed to web (not in local mode)
        qcr_airtable_fallback = ""
        if not is_local_mode() and HAS_AIRTABLE:
            try:
                airtable_url = get_airtable_form_url('qcr', dict(item), token)
                if airtable_url:
                    qcr_airtable_fallback = f'''
    <!-- FALLBACK OPTION -->
    <div style="margin-top:24px; padding:16px; background:#fff8e6; border:1px solid #ffd666; border-radius:8px;">
        <div style="font-weight:bold; color:#d48806; margin-bottom:8px;">
            🌐 Server Unavailable?
        </div>
        <p style="font-size:13px; color:#666; margin-bottom:12px;">
            If the button above doesn't work, use this alternative form:
        </p>
        <a href="{airtable_url}"
           style="background:#ffd666; color:#333; padding:10px 20px; 
                  font-size:14px; font-weight:600; text-decoration:none; border-radius:6px; display:inline-block;">
            📝 Use Backup Form
        </a>
    </div>
'''
            except:
                pass
        
        html_body = html_body.replace('<!--QCR_AIRTABLE_FALLBACK_PLACEHOLDER-->', qcr_airtable_fallback)
    
    try:
        pythoncom.CoInitialize()
        try:
            outlook = win32com.client.Dispatch("Outlook.Application")
            mail = outlook.CreateItem(0)  # 0 = olMailItem
            mail.To = item['qcr_email']
            if item['reviewer_email']:
                mail.CC = item['reviewer_email']
            mail.Subject = subject
            mail.HTMLBody = html_body
            mail.Send()
            
            # Update database
            cursor.execute('''
                UPDATE item SET 
                    qcr_email_sent_at = ?,
                    qcr_response_status = 'Email Sent',
                    status = 'In QC'
                WHERE id = ?
            ''', (datetime.now().isoformat(), item_id))
            conn.commit()
            
        finally:
            pythoncom.CoUninitialize()
        
        conn.close()
        return {'success': True, 'message': 'QCR assignment email sent'}
        
    except Exception as e:
        conn.close()
        return {'success': False, 'error': str(e)}


def send_qcr_version_update_email(item_id, version):
    """Send an email to QCR notifying them of an updated reviewer response (before QCR has responded)."""
    if not HAS_WIN32COM:
        return {'success': False, 'error': 'Outlook not available'}
    
    conn = get_db()
    cursor = conn.cursor()
    
    # Get item with reviewer info
    cursor.execute('''
        SELECT i.*, 
               ir.email as reviewer_email, ir.display_name as reviewer_name,
               qcr.email as qcr_email, qcr.display_name as qcr_name
        FROM item i
        LEFT JOIN user ir ON i.initial_reviewer_id = ir.id
        LEFT JOIN user qcr ON i.qcr_id = qcr.id
        WHERE i.id = ?
    ''', (item_id,))
    item = cursor.fetchone()
    
    if not item:
        conn.close()
        return {'success': False, 'error': 'Item not found'}
    
    if not item['qcr_email']:
        conn.close()
        return {'success': False, 'error': 'No QCR assigned'}
    
    # Get existing QCR token
    token = item['email_token_qcr']
    if not token:
        token = generate_token()
        cursor.execute('UPDATE item SET email_token_qcr = ? WHERE id = ?', (token, item_id))
        conn.commit()
    
    # Build the form URL - use Airtable in local mode, server URL when deployed
    server_url = f"{get_app_host()}/respond/qcr?token={token}"
    
    if is_local_mode() and HAS_AIRTABLE:
        airtable_url = get_airtable_form_url('qcr', dict(item), token)
        if airtable_url:
            respond_url = airtable_url
        else:
            respond_url = server_url
    else:
        respond_url = server_url
    
    # Get version history
    cursor.execute('''
        SELECT version, submitted_at 
        FROM reviewer_response_history 
        WHERE item_id = ? 
        ORDER BY version DESC
        LIMIT 5
    ''', (item_id,))
    history = cursor.fetchall()
    version_history_html = ''
    if history:
        history_parts = [f"v{h['version']} ({h['submitted_at'][:16].replace('T', ' ')})" for h in history]
        version_history_html = f"<p><strong>Previous versions:</strong> {', '.join(history_parts)}</p>"
    
    conn.close()
    
    # Format reviewer notes
    reviewer_notes_html = (item['reviewer_notes'] or 'No notes provided').replace('\n', '<br>')
    
    # Format reviewer internal notes (for team visibility only)
    reviewer_internal_notes_html = ''
    if item['reviewer_internal_notes']:
        reviewer_internal_notes_html = item['reviewer_internal_notes'].replace('\n', '<br>')
    
    # Format files
    reviewer_files_display = 'None selected'
    if item['reviewer_selected_files']:
        try:
            files = json.loads(item['reviewer_selected_files'])
            if files:
                reviewer_files_display = '<br>'.join([f"• {f}" for f in files])
        except:
            pass
    
    subject = f"[LEB] {item['identifier']} – Reviewer response updated (v{version})"
    
    html_body = f"""<div style="font-family:Segoe UI, Helvetica, Arial, sans-serif; color:#333; font-size:14px; line-height:1.5;">

    <h2 style="color:#444; margin-bottom:6px;">
        [LEB] {item['identifier']} – Reviewer Response Updated
    </h2>

    <div style="background: #fef3c7; border: 1px solid #f59e0b; border-radius: 8px; padding: 15px; margin: 15px 0;">
        <p style="margin: 0; color: #92400e;">
            <strong>📝 Update Notice:</strong> The Initial Reviewer has updated their response to <strong>version {version}</strong>.
        </p>
    </div>

    <table cellpadding="8" cellspacing="0" width="100%" style="border-collapse:collapse; margin-top:10px;">
        <tr>
            <td colspan="2" style="background:#d5f5e3; font-weight:bold; border:1px solid #82e0aa; color:#1e8449;">
                ✅ Updated Reviewer Response (v{version})
            </td>
        </tr>

        <tr>
            <td style="width:160px; border:1px solid #ddd; font-weight:bold;">Category</td>
            <td style="border:1px solid #ddd; color:#1e8449; font-weight:bold;">{item['reviewer_response_category'] or 'Not specified'}</td>
        </tr>

        <tr>
            <td style="border:1px solid #ddd; font-weight:bold; vertical-align:top;">Selected Files</td>
            <td style="border:1px solid #ddd;">{reviewer_files_display}</td>
        </tr>

        <tr>
            <td style="border:1px solid #ddd; font-weight:bold; vertical-align:top;">Description</td>
            <td style="border:1px solid #ddd;">{reviewer_notes_html}</td>
        </tr>
    </table>

    {f'''<!-- INTERNAL NOTES SECTION (Team Only) -->
    <div style="margin-top:16px; padding:12px; background:#fff8e6; border:1px solid #ffd966; border-radius:6px;">
        <div style="font-weight:bold; color:#b7791f; margin-bottom:6px;">🔒 Reviewer's Internal Notes (Team Only)</div>
        <div style="color:#744210; font-size:13px;">{reviewer_internal_notes_html}</div>
    </div>''' if reviewer_internal_notes_html else ''}

    {version_history_html}

    <p style="margin-top: 15px;">The QC Review form will always show the <strong>latest version</strong> of the reviewer's response.</p>

    <div style="margin-top:22px; text-align:left;">
        <a href="{respond_url}"
           style="background:#27ae60; color:white; padding:10px 18px; 
                  font-size:15px; text-decoration:none; border-radius:4px; display:inline-block;">
            Open QC Review Form
        </a>
    </div>

    <p style="margin-top:20px; font-size:12px; color:#777;">
        <em>This message was automatically generated. If you believe you received this by mistake, please contact the project administrator.</em>
    </p>

</div>"""
    
    try:
        pythoncom.CoInitialize()
        try:
            outlook = win32com.client.Dispatch("Outlook.Application")
            mail = outlook.CreateItem(0)
            mail.To = item['qcr_email']
            if item['reviewer_email']:
                mail.CC = item['reviewer_email']
            mail.Subject = subject
            mail.HTMLBody = html_body
            mail.Send()
        finally:
            pythoncom.CoUninitialize()
        
        return {'success': True, 'message': 'Version update email sent'}
        
    except Exception as e:
        return {'success': False, 'error': str(e)}


def send_reviewer_notification_email(item_id, qc_action, qcr_notes, final_category=None, final_text=None):
    """Send notification email to reviewer based on QCR action."""
    if not HAS_WIN32COM:
        return {'success': False, 'error': 'Outlook not available'}
    
    conn = get_db()
    cursor = conn.cursor()
    
    # Get item with reviewer info
    cursor.execute('''
        SELECT i.*, 
               ir.email as reviewer_email, ir.display_name as reviewer_name,
               qcr.email as qcr_email, qcr.display_name as qcr_name
        FROM item i
        LEFT JOIN user ir ON i.initial_reviewer_id = ir.id
        LEFT JOIN user qcr ON i.qcr_id = qcr.id
        WHERE i.id = ?
    ''', (item_id,))
    item = cursor.fetchone()
    
    if not item:
        conn.close()
        return {'success': False, 'error': 'Item not found'}
    
    if not item['reviewer_email']:
        conn.close()
        return {'success': False, 'error': 'No reviewer email'}
    
    # Get version info
    version = item['reviewer_response_version'] if item['reviewer_response_version'] is not None else 0
    
    # Build email based on action
    if qc_action == 'Approve':
        subject = f"[LEB] {item['identifier']} – Your response (v{version}) was approved"
        body_intro = f"""<p>Good news! Your response <strong>(v{version})</strong> for the following item has been <strong style="color: #059669;">approved</strong> by QC.</p>"""
        action_color = '#059669'
        action_icon = '✅'
    elif qc_action == 'Modify':
        subject = f"[LEB] {item['identifier']} – Your response (v{version}) was modified by QC"
        body_intro = f"""<p>Your response <strong>(v{version})</strong> for the following item has been <strong style="color: #2563eb;">modified</strong> by QC and is now finalized.</p>"""
        action_color = '#2563eb'
        action_icon = '✏️'
    else:  # Send Back
        subject = f"[LEB] {item['identifier']} – Revisions requested on v{version}"
        body_intro = f"""<p>Your response <strong>(v{version})</strong> for the following item has been <strong style="color: #dc2626;">returned</strong> for revision by QC.</p>"""
        action_color = '#dc2626'
        action_icon = '↩️'
    
    # Format QCR notes
    qcr_notes_html = (qcr_notes or 'No notes provided').replace('\n', '<br>')
    
    # Format QCR internal notes (for team visibility only)
    qcr_internal_notes_html = ''
    if item['qcr_internal_notes']:
        qcr_internal_notes_html = item['qcr_internal_notes'].replace('\n', '<br>')
    
    # Build the response comparison section for Modify action
    response_comparison = ''
    if qc_action == 'Modify' and final_text:
        original_text = item['reviewer_response_text'] or item['reviewer_notes'] or 'No original text'
        response_comparison = f"""
        <div style="margin: 20px 0;">
            <h4 style="margin: 0 0 10px 0;">Your Original Response (v{version}):</h4>
            <div style="background: #fef2f2; border: 1px solid #fecaca; border-radius: 6px; padding: 12px; margin-bottom: 15px;">
                {original_text.replace(chr(10), '<br>')}
            </div>
            
            <h4 style="margin: 0 0 10px 0;">Final Response (QC Modified):</h4>
            <div style="background: #f0fdf4; border: 1px solid #86efac; border-radius: 6px; padding: 12px;">
                {final_text.replace(chr(10), '<br>')}
            </div>
        </div>
        """
    
    # Build revision link for Send Back action
    revision_link = ''
    if qc_action == 'Send Back':
        # Generate new token for revision
        new_token = generate_token()
        cursor.execute('UPDATE item SET email_token_reviewer = ? WHERE id = ?', (new_token, item_id))
        conn.commit()
        
        # Build the form URL - use Airtable in local mode, server URL when deployed
        server_url = f"{get_app_host()}/respond/reviewer?token={new_token}"
        
        if is_local_mode() and HAS_AIRTABLE:
            airtable_url = get_airtable_form_url('reviewer', dict(item), new_token)
            if airtable_url:
                respond_url = airtable_url
            else:
                respond_url = server_url
        else:
            respond_url = server_url
        
        revision_link = f"""
        <p style="margin: 20px 0;">
            <a href="{respond_url}" style="display: inline-block; background: #dc2626; color: white; padding: 12px 24px; text-decoration: none; border-radius: 8px; font-weight: bold;">📝 Revise Your Response (Submit v{version + 1})</a>
        </p>
        """
    
    conn.close()
    
    html_body = f"""<html><body style="font-family: Arial, sans-serif; font-size: 14px; color: #333;">
<p>Hello {item['reviewer_name'] or 'Reviewer'},</p>

{body_intro}

<table style="border-collapse: collapse; margin: 15px 0;">
<tr><td style="padding: 5px 15px 5px 0; font-weight: bold;">Type:</td><td>{item['type']}</td></tr>
<tr><td style="padding: 5px 15px 5px 0; font-weight: bold;">Identifier:</td><td>{item['identifier']}</td></tr>
<tr><td style="padding: 5px 15px 5px 0; font-weight: bold;">Title:</td><td>{item['title'] or 'N/A'}</td></tr>
<tr><td style="padding: 5px 15px 5px 0; font-weight: bold;">Due Date:</td><td>{item['due_date'] or 'N/A'}</td></tr>
<tr><td style="padding: 5px 15px 5px 0; font-weight: bold;">Priority:</td><td>{item['priority'] or 'Normal'}</td></tr>
<tr><td style="padding: 5px 15px 5px 0; font-weight: bold;">Response Version:</td><td>v{version}</td></tr>
</table>

<div style="background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px; padding: 15px; margin: 20px 0;">
<h3 style="margin: 0 0 10px 0; color: {action_color};">{action_icon} QC Decision: {qc_action}</h3>
<p style="font-size: 12px; color: #666; margin-bottom: 10px;">Feedback on your v{version} response</p>
{f'<p><strong>Final Category:</strong> {final_category}</p>' if final_category else ''}
<p><strong>QC Notes:</strong></p>
<div style="background: white; border-radius: 4px; padding: 10px; margin-top: 8px;">
{qcr_notes_html}
</div>
</div>

{f'''<!-- QCR INTERNAL NOTES SECTION (Team Only) -->
<div style="margin: 20px 0; padding:12px; background:#fff8e6; border:1px solid #ffd966; border-radius:6px;">
    <div style="font-weight:bold; color:#b7791f; margin-bottom:6px;">🔒 QCR's Internal Notes (Team Only)</div>
    <div style="color:#744210; font-size:13px;">{qcr_internal_notes_html}</div>
</div>''' if qcr_internal_notes_html else ''}

{response_comparison}
{revision_link}

<p>Thank you.</p>
</body></html>"""
    
    try:
        pythoncom.CoInitialize()
        try:
            outlook = win32com.client.Dispatch("Outlook.Application")
            mail = outlook.CreateItem(0)
            mail.To = item['reviewer_email']
            if item['qcr_email']:
                mail.CC = item['qcr_email']
            mail.Subject = subject
            mail.HTMLBody = html_body
            mail.Send()
        finally:
            pythoncom.CoUninitialize()
        
        return {'success': True, 'message': 'Reviewer notification sent'}
        
    except Exception as e:
        return {'success': False, 'error': str(e)}


def send_qcr_completion_confirmation_email(item_id, qc_action, qcr_notes, final_category=None, final_text=None):
    """Send confirmation email to both QCR and reviewer after QCR completes review."""
    if not HAS_WIN32COM:
        return {'success': False, 'error': 'Outlook not available'}
    
    conn = get_db()
    cursor = conn.cursor()
    
    # Get item with reviewer and QCR info
    cursor.execute('''
        SELECT i.*, 
               ir.email as reviewer_email, ir.display_name as reviewer_name,
               qcr.email as qcr_email, qcr.display_name as qcr_name
        FROM item i
        LEFT JOIN user ir ON i.initial_reviewer_id = ir.id
        LEFT JOIN user qcr ON i.qcr_id = qcr.id
        WHERE i.id = ?
    ''', (item_id,))
    item = cursor.fetchone()
    conn.close()
    
    if not item:
        return {'success': False, 'error': 'Item not found'}
    
    # Get version info
    version = item['reviewer_response_version'] if item['reviewer_response_version'] is not None else 1
    
    # Build comparison section
    original_category = item['reviewer_response_category'] or 'Not specified'
    original_text = item['reviewer_response_text'] or item['reviewer_notes'] or 'No notes provided'
    final_category_display = final_category or item['final_response_category'] or original_category
    final_text_display = final_text or item['final_response_text'] or original_text
    
    # Parse selected files
    original_files = 'None selected'
    if item['reviewer_selected_files']:
        try:
            files = json.loads(item['reviewer_selected_files'])
            if files:
                original_files = '; '.join(files)
        except:
            original_files = item['reviewer_selected_files']
    
    final_files = original_files
    if item['final_response_files']:
        try:
            files = json.loads(item['final_response_files'])
            if files:
                final_files = '; '.join(files)
        except:
            final_files = item['final_response_files']
    
    # Determine if there were changes
    category_changed = original_category != final_category_display
    text_changed = original_text.strip() != final_text_display.strip()
    files_changed = original_files != final_files
    has_changes = category_changed or text_changed or files_changed
    
    # Build the changes summary
    changes_list = []
    if category_changed:
        changes_list.append(f'Category: "{original_category}" → "{final_category_display}"')
    if text_changed:
        changes_list.append('Response text was modified')
    if files_changed:
        changes_list.append('Selected files were updated')
    
    changes_summary = '<br>'.join(changes_list) if changes_list else 'No changes were made'
    
    # Determine action display
    if qc_action == 'Approve':
        action_display = '✅ Approved'
        action_color = '#059669'
        action_description = 'The initial reviewer\'s response was accepted without changes.'
    else:  # Modify
        action_display = '✏️ Modified'
        action_color = '#2563eb'
        action_description = 'The QC reviewer made modifications to the response.'
    
    # Format QCR notes
    qcr_notes_html = (qcr_notes or 'No notes provided').replace('\n', '<br>')
    
    # Format QCR internal notes (for team visibility)
    qcr_internal_notes_html = ''
    if item['qcr_internal_notes']:
        qcr_internal_notes_html = item['qcr_internal_notes'].replace('\n', '<br>')
    
    # Build comparison HTML
    comparison_html = f"""
    <div style="margin: 20px 0;">
        <h4 style="margin: 0 0 15px 0; border-bottom: 1px solid #e5e7eb; padding-bottom: 8px;">📊 Response Comparison</h4>
        
        <table style="width: 100%; border-collapse: collapse; margin-bottom: 15px;">
            <tr style="background: #f8fafc;">
                <th style="text-align: left; padding: 10px; border: 1px solid #e5e7eb; width: 140px;"></th>
                <th style="text-align: left; padding: 10px; border: 1px solid #e5e7eb;">Original (Reviewer v{version})</th>
                <th style="text-align: left; padding: 10px; border: 1px solid #e5e7eb;">Final (QC {qc_action})</th>
            </tr>
            <tr>
                <td style="padding: 10px; border: 1px solid #e5e7eb; font-weight: bold;">Category</td>
                <td style="padding: 10px; border: 1px solid #e5e7eb; {'background: #fef2f2;' if category_changed else ''}">{original_category}</td>
                <td style="padding: 10px; border: 1px solid #e5e7eb; {'background: #f0fdf4;' if category_changed else ''}">{final_category_display}</td>
            </tr>
            <tr>
                <td style="padding: 10px; border: 1px solid #e5e7eb; font-weight: bold; vertical-align: top;">Response Text</td>
                <td style="padding: 10px; border: 1px solid #e5e7eb; vertical-align: top; {'background: #fef2f2;' if text_changed else ''}">{original_text.replace(chr(10), '<br>')}</td>
                <td style="padding: 10px; border: 1px solid #e5e7eb; vertical-align: top; {'background: #f0fdf4;' if text_changed else ''}">{final_text_display.replace(chr(10), '<br>')}</td>
            </tr>
            <tr>
                <td style="padding: 10px; border: 1px solid #e5e7eb; font-weight: bold; vertical-align: top;">Selected Files</td>
                <td style="padding: 10px; border: 1px solid #e5e7eb; vertical-align: top; font-size: 12px; {'background: #fef2f2;' if files_changed else ''}">{original_files}</td>
                <td style="padding: 10px; border: 1px solid #e5e7eb; vertical-align: top; font-size: 12px; {'background: #f0fdf4;' if files_changed else ''}">{final_files}</td>
            </tr>
        </table>
        
        <div style="background: {'#fef3c7' if has_changes else '#f0fdf4'}; border: 1px solid {'#fbbf24' if has_changes else '#86efac'}; border-radius: 6px; padding: 12px;">
            <strong>{'⚠️ Changes Made:' if has_changes else '✅ No Changes:'}</strong><br>
            {changes_summary}
        </div>
    </div>
    """
    
    subject = f"[LEB] {item['identifier']} – QC Review Complete ({qc_action})"
    
    html_body = f"""<html><body style="font-family: Arial, sans-serif; font-size: 14px; color: #333;">
<p style="font-size: 16px; font-weight: 600; color: #166534; margin-bottom: 15px;">✅ Final Response Recorded – {item['identifier']}</p>

<p>The QC review for the item below has been completed and the final response has been recorded. This item is now ready for closeout.</p>

<table style="border-collapse: collapse; margin: 15px 0;">
<tr><td style="padding: 5px 15px 5px 0; font-weight: bold;">Type:</td><td>{item['type']}</td></tr>
<tr><td style="padding: 5px 15px 5px 0; font-weight: bold;">Identifier:</td><td>{item['identifier']}</td></tr>
<tr><td style="padding: 5px 15px 5px 0; font-weight: bold;">Title:</td><td>{item['title'] or 'N/A'}</td></tr>
<tr><td style="padding: 5px 15px 5px 0; font-weight: bold;">Contractor Due Date:</td><td>{item['due_date'] or 'N/A'}</td></tr>
<tr><td style="padding: 5px 15px 5px 0; font-weight: bold;">Initial Reviewer:</td><td>{item['reviewer_name'] or 'N/A'}</td></tr>
<tr><td style="padding: 5px 15px 5px 0; font-weight: bold;">QC Reviewer:</td><td>{item['qcr_name'] or 'N/A'}</td></tr>
</table>

<div style="background: #f0fdf4; border: 1px solid #86efac; border-radius: 8px; padding: 15px; margin: 20px 0;">
<h3 style="margin: 0 0 10px 0; color: {action_color};">{action_display}</h3>
<p>{action_description}</p>
<p><strong>Final Category:</strong> {final_category_display}</p>
</div>

{comparison_html}

<div style="background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px; padding: 15px; margin: 20px 0;">
<p style="margin: 0 0 8px 0;"><strong>QC Reviewer Notes:</strong></p>
<div style="background: white; border-radius: 4px; padding: 10px;">
{qcr_notes_html}
</div>
</div>

{f'''<!-- QCR INTERNAL NOTES (Team Only) -->
<div style="margin: 20px 0; padding:12px; background:#fff8e6; border:1px solid #ffd966; border-radius:6px;">
    <div style="font-weight:bold; color:#b7791f; margin-bottom:6px;">🔒 QCR's Internal Notes (Team Only)</div>
    <div style="color:#744210; font-size:13px;">{qcr_internal_notes_html}</div>
</div>''' if qcr_internal_notes_html else ''}

<p style="color: #666; font-size: 13px;">This item is now ready for final response to the contractor.</p>

<p>Thank you.</p>
</body></html>"""
    
    try:
        pythoncom.CoInitialize()
        try:
            outlook = win32com.client.Dispatch("Outlook.Application")
            mail = outlook.CreateItem(0)
            
            # Send to both QCR and reviewer
            recipients = []
            if item['qcr_email']:
                recipients.append(item['qcr_email'])
            if item['reviewer_email']:
                recipients.append(item['reviewer_email'])
            
            if not recipients:
                return {'success': False, 'error': 'No recipients found'}
            
            mail.To = '; '.join(recipients)
            mail.Subject = subject
            mail.HTMLBody = html_body
            mail.Send()
        finally:
            pythoncom.CoUninitialize()
        
        return {'success': True, 'message': 'QCR completion confirmation sent'}
        
    except Exception as e:
        return {'success': False, 'error': str(e)}


# =============================================================================
# MULTI-REVIEWER EMAIL FUNCTIONS
# =============================================================================

def generate_multi_reviewer_form(item_id, reviewer_record):
    """Generate an HTA form for a specific reviewer in multi-reviewer mode.
    
    Args:
        item_id: The item ID
        reviewer_record: The reviewer record from item_reviewers table
    
    Returns:
        dict with 'success', 'path' or 'error'
    """
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT i.*, qcr.display_name as qcr_name, qcr.email as qcr_email
        FROM item i
        LEFT JOIN user qcr ON i.qcr_id = qcr.id
        WHERE i.id = ?
    ''', (item_id,))
    item = cursor.fetchone()
    
    if not item:
        conn.close()
        return {'success': False, 'error': 'Item not found'}
    
    # Convert to dict for easier handling
    item = dict(item)
    
    if not item['folder_link']:
        conn.close()
        return {'success': False, 'error': 'Item has no folder assigned'}
    
    # Get all reviewers for this item to show status
    cursor.execute('SELECT * FROM item_reviewers WHERE item_id = ?', (item_id,))
    all_reviewers = cursor.fetchall()
    conn.close()
    
    # Calculate due dates
    reviewer_due = item['initial_reviewer_due_date'] or 'N/A'
    qcr_due = item['qcr_due_date'] or 'N/A'
    
    # Check if this is truly multi-reviewer (more than 1 reviewer)
    is_multi_reviewer = len(all_reviewers) > 1
    
    # Build other reviewers status section (only for multi-reviewer)
    other_reviewers_html = ""
    if is_multi_reviewer:
        other_reviewers_rows = ""
        for r in all_reviewers:
            is_current = r['reviewer_email'] == reviewer_record['reviewer_email']
            if is_current:
                status_badge = '<span style="background:#3b82f6; color:white; padding:2px 8px; border-radius:10px; font-size:11px;">You</span>'
            elif r['response_at']:
                status_badge = f'<span style="background:#10b981; color:white; padding:2px 8px; border-radius:10px; font-size:11px;">Responded - {r["response_category"] or "N/A"}</span>'
            else:
                status_badge = '<span style="background:#f59e0b; color:white; padding:2px 8px; border-radius:10px; font-size:11px;">Pending</span>'
            other_reviewers_rows += f'''
            <div style="display:flex; justify-content:space-between; align-items:center; padding:8px; background:{"#eff6ff" if is_current else "#f8f9fa"}; border-radius:6px; margin-bottom:6px;">
                <span style="font-weight:{"600" if is_current else "normal"};">{r["reviewer_name"]}</span>
                {status_badge}
            </div>'''
        
        other_reviewers_html = f'''
        <div style="background:#f0f9ff; border:1px solid #bae6fd; border-radius:8px; padding:16px; margin-bottom:24px;">
            <div style="font-size:14px; font-weight:600; color:#0369a1; margin-bottom:12px;">Other Reviewers ({len(all_reviewers)} total)</div>
            {other_reviewers_rows}
        </div>'''
    
    # Load appropriate template based on item type and reviewer count
    item_type = (item['type'] or '').upper()
    
    if is_multi_reviewer:
        # Multi-reviewer: use multi-reviewer specific template (no file selection, Bluebeam focused)
        template_path = TEMPLATES_DIR / "_MULTI_REVIEWER_RESPONSE_TEMPLATE.hta"
        if not template_path.exists():
            # Fallback to regular template
            template_path = TEMPLATES_DIR / "_RESPONSE_FORM_TEMPLATE_v3.hta"
            if not template_path.exists():
                template_path = TEMPLATES_DIR / "_RESPONSE_FORM_TEMPLATE_v3.html"
                if not template_path.exists():
                    return {'success': False, 'error': 'Multi-reviewer form template not found'}
    else:
        # Single reviewer: check if RFI or Submittal
        if item_type == 'RFI':
            # RFI: use RFI-specific template (no response category, different labels)
            template_path = TEMPLATES_DIR / "_RFI_RESPONSE_FORM_TEMPLATE.hta"
            if not template_path.exists():
                return {'success': False, 'error': 'RFI form template not found'}
        else:
            # Submittal: use regular single-reviewer template (with file selection and response category)
            template_path = TEMPLATES_DIR / "_RESPONSE_FORM_TEMPLATE_v3.hta"
            if not template_path.exists():
                template_path = TEMPLATES_DIR / "_RESPONSE_FORM_TEMPLATE_v3.html"
                if not template_path.exists():
                    return {'success': False, 'error': 'Reviewer form template not found'}
    
    with open(template_path, 'r', encoding='utf-8') as f:
        template = f.read()
    
    # Escape special characters for JavaScript embedding
    def js_escape(s):
        if not s:
            return ''
        return s.replace('\\', '\\\\').replace('"', '\\"').replace("'", "\\'").replace('\n', '\\n').replace('\r', '')
    
    # Replace placeholders - use reviewer info from item_reviewers record
    html = template.replace('{{ITEM_ID}}', str(item['id']))
    html = html.replace('{{ITEM_TYPE}}', item['type'] or '')
    html = html.replace('{{ITEM_IDENTIFIER}}', item['identifier'] or '')
    html = html.replace('{{ITEM_TITLE}}', js_escape(item['title']) or 'N/A')
    html = html.replace('{{DATE_RECEIVED}}', item['date_received'] or 'N/A')
    html = html.replace('{{REVIEWER_DUE_DATE}}', reviewer_due)
    html = html.replace('{{QCR_DUE_DATE}}', qcr_due)
    html = html.replace('{{CONTRACTOR_DUE_DATE}}', item['due_date'] or 'N/A')
    html = html.replace('{{REVIEWER_NAME}}', js_escape(reviewer_record['reviewer_name']) or 'N/A')
    html = html.replace('{{REVIEWER_EMAIL}}', reviewer_record['reviewer_email'] or '')
    html = html.replace('{{TOKEN}}', reviewer_record['email_token'] or '')
    html = html.replace('{{FOLDER_PATH}}', js_escape(item['folder_link']) or '')
    html = html.replace('{{FOLDER_PATH_RAW}}', item['folder_link'] or '')
    html = html.replace('{{RFI_QUESTION}}', js_escape(item.get('rfi_question', '') or 'N/A'))
    html = html.replace('{{OTHER_REVIEWERS_SECTION}}', other_reviewers_html)
    
    # Save to Responses subfolder with reviewer-specific name
    folder_path = Path(item['folder_link'])
    responses_folder = folder_path / "Responses"
    
    # Create Responses subfolder if it doesn't exist
    try:
        responses_folder.mkdir(exist_ok=True)
    except Exception as e:
        return {'success': False, 'error': f'Failed to create Responses folder: {e}'}
    
    # Create safe filename from reviewer name
    safe_name = "".join(c for c in reviewer_record['reviewer_name'] if c.isalnum() or c in (' ', '-', '_')).strip()
    safe_name = safe_name.replace(' ', '_')
    
    form_path = responses_folder / f"_RESPONSE_FORM_{safe_name}.hta"
    
    try:
        with open(form_path, 'w', encoding='utf-8') as f:
            f.write(html)
        return {'success': True, 'path': str(form_path)}
    except Exception as e:
        return {'success': False, 'error': f'Failed to save form: {e}'}


def send_multi_reviewer_assignment_emails(item_id):
    """Send assignment emails to all reviewers in multi-reviewer mode.
    
    Uses the same email template style as single-reviewer mode:
    - Each reviewer gets their own email with the item details
    - QCR is CC'd on each email
    - In local mode, generates .hta forms for each reviewer
    - Instructions reference Bluebeam session for markups
    """
    if not HAS_WIN32COM:
        return {'success': False, 'error': 'Outlook not available'}
    
    conn = get_db()
    cursor = conn.cursor()
    
    # Get item info with QCR
    cursor.execute('''
        SELECT i.*, qcr.email as qcr_email, qcr.display_name as qcr_name
        FROM item i
        LEFT JOIN user qcr ON i.qcr_id = qcr.id
        WHERE i.id = ?
    ''', (item_id,))
    item = cursor.fetchone()
    
    if not item:
        conn.close()
        return {'success': False, 'error': 'Item not found'}
    
    # Get all reviewers
    cursor.execute('SELECT * FROM item_reviewers WHERE item_id = ?', (item_id,))
    reviewers = cursor.fetchall()
    
    if not reviewers:
        conn.close()
        return {'success': False, 'error': 'No reviewers assigned'}
    
    # Check if single reviewer mode (for email language)
    is_single_reviewer = len(reviewers) == 1
    
    # Ensure multi_reviewer_mode flag is correct based on actual count
    correct_mode = 0 if is_single_reviewer else 1
    if item['multi_reviewer_mode'] != correct_mode:
        cursor.execute('UPDATE item SET multi_reviewer_mode = ? WHERE id = ?', (correct_mode, item_id))
        conn.commit()
        print(f"  [Send Emails] Corrected multi_reviewer_mode to {correct_mode} for item {item_id} ({len(reviewers)} reviewers)")
    
    # Calculate due dates if needed
    reviewer_due_date = item['initial_reviewer_due_date']
    qcr_due_date = item['qcr_due_date']
    
    if item['date_received'] and item['due_date']:
        calculated = calculate_review_due_dates(
            item['date_received'],
            item['due_date'],
            item['priority']
        )
        if not reviewer_due_date:
            reviewer_due_date = calculated['initial_reviewer_due_date']
        if not qcr_due_date:
            qcr_due_date = calculated['qcr_due_date']
        
        # Update database if values were missing
        if not item['initial_reviewer_due_date'] or not item['qcr_due_date']:
            cursor.execute('''
                UPDATE item SET 
                    initial_reviewer_due_date = COALESCE(initial_reviewer_due_date, ?),
                    qcr_due_date = COALESCE(qcr_due_date, ?)
                WHERE id = ?
            ''', (calculated['initial_reviewer_due_date'], calculated['qcr_due_date'], item_id))
            conn.commit()
    
    # Priority color
    priority_color = '#e67e22' if item['priority'] == 'Medium' else '#c0392b' if item['priority'] == 'High' else '#27ae60'
    
    # Build list of all reviewer names for email display (each on new line) - only needed for multi-reviewer
    all_reviewer_names = "<br>".join([r['reviewer_name'] for r in reviewers]) if not is_single_reviewer else ""
    
    # Build the reviewer row HTML - only show when multiple reviewers
    reviewer_row_html = f'''<tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Initial Reviewers</td>
            <td style="border:1px solid #ddd;">{all_reviewer_names}</td>
        </tr>''' if not is_single_reviewer else ""
    
    # QCR notification text - singular vs plural
    qcr_note_text = f"<strong>{item['qcr_name'] or 'The QC Reviewer'}</strong> has been assigned and will be notified once {'you have submitted your response' if is_single_reviewer else 'all reviewers have submitted their responses'}."
    
    # Determine if using file-based forms (local mode) - still need folder for form generation
    folder_path = item['folder_link'] or 'Not set'
    use_file_form = is_local_mode() and folder_path != 'Not set'
    
    sent_count = 0
    errors = []
    
    for reviewer in reviewers:
        try:
            # Generate token if not exists
            token = reviewer['email_token']
            if not token:
                token = generate_token()
                cursor.execute('UPDATE item_reviewers SET email_token = ? WHERE id = ?', (token, reviewer['id']))
                conn.commit()
                # Update reviewer record with token
                reviewer = dict(reviewer)
                reviewer['email_token'] = token
            
            subject = f"[LEB] {item['identifier']} – Assigned to You"
            
            # Build RFI Question section if applicable
            rfi_question_html = ""
            if (item['type'] or '').upper() == 'RFI' and item.get('rfi_question'):
                rfi_question_html = f"""
    <!-- RFI QUESTION -->
    <div style="margin:15px 0; padding:15px; background:#fff8e1; border:2px solid #ffc107; border-radius:8px;">
        <div style="font-size:14px; color:#f57c00; font-weight:bold; margin-bottom:8px;">❓ RFI Question</div>
        <div style="font-size:13px; color:#333; line-height:1.6; white-space:pre-wrap;">{item.get('rfi_question', '')}</div>
    </div>"""
            
            if use_file_form:
                # Generate the HTA form file for this reviewer
                form_result = generate_multi_reviewer_form(item_id, dict(reviewer))
                if form_result['success']:
                    form_file_path = form_result['path']
                    form_file_link = f'file:///{form_file_path.replace(chr(92), "/")}'
                    
                    html_body = f"""<div style="font-family:Segoe UI, Helvetica, Arial, sans-serif; color:#333; font-size:14px; line-height:1.5;">

    <!-- HEADER -->
    <h2 style="color:#444; margin-bottom:6px;">
        [LEB] {item['identifier']} - Assigned to You
    </h2>

    <p style="margin-top:0; font-size:13px; color:#666;">
        You have been assigned a new review task. Please review the details below.
    </p>

    {rfi_question_html}

    <!-- BLUEBEAM INSTRUCTIONS -->
    <div style="margin:15px 0; padding:15px; background:#dbeafe; border:1px solid #3b82f6; border-radius:8px;">
        <div style="font-size:14px; color:#1e40af; font-weight:bold;">📐 Markups Instructions</div>
        <div style="font-size:13px; color:#1e40af; margin-top:8px;">
            <strong>Provide markups in the corresponding Bluebeam session.</strong><br/>
            Do not attach files to your response. All markups should be completed in the shared Bluebeam Studio session.
        </div>
    </div>

    <!-- DIRECT LINK TO FORM - PROMINENT BUTTON -->
    <div style="margin:20px 0; text-align:center;">
        <table cellpadding="0" cellspacing="0" border="0" style="margin:0 auto;">
            <tr>
                <td align="center" bgcolor="#27ae60" style="background:#27ae60; border-radius:8px; padding:0;">
                    <a href="{form_file_link}" target="_blank"
                       style="background:#27ae60; color:#ffffff; display:inline-block; font-family:Segoe UI,Arial,sans-serif;
                              font-size:16px; font-weight:bold; line-height:50px; text-align:center; text-decoration:none;
                              width:320px; -webkit-text-size-adjust:none; border-radius:8px;">
                        OPEN RESPONSE FORM
                    </a>
                </td>
            </tr>
        </table>
    </div>

    <!-- INSTRUCTIONS FOR HTA -->
    <div style="margin:20px 0; padding:15px; background:#e8f5e9; border:1px solid #4caf50; border-radius:8px;">
        <div style="font-size:14px; color:#2e7d32;">
            <strong>Instructions:</strong>
            <ol style="margin:8px 0 0 0; padding-left:20px;">
                <li>Click the green button above to open the response form</li>
                <li>If prompted, select <strong>"Microsoft (R) HTML Application host"</strong> - choose <strong>Open</strong>, do NOT save the file</li>
                <li>Select your response category</li>
                <li>Click <strong>Submit Response</strong> - your response will be saved automatically</li>
            </ol>
        </div>
    </div>

    <!-- INFO TABLE -->
    <table cellpadding="8" cellspacing="0" width="100%" style="border-collapse:collapse; margin-top:10px;">
        <tr>
            <td colspan="2" style="background:#f2f2f2; font-weight:bold; border:1px solid #ddd;">
                Item Information
            </td>
        </tr>
        <tr>
            <td style="width:160px; border:1px solid #ddd; font-weight:bold;">Type</td>
            <td style="border:1px solid #ddd;">{item['type']}</td>
        </tr>
        <tr>
            <td style="width:160px; border:1px solid #ddd; font-weight:bold;">Identifier</td>
            <td style="border:1px solid #ddd;">{item['identifier']}</td>
        </tr>
        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Title</td>
            <td style="border:1px solid #ddd;">{item['title'] or 'N/A'}</td>
        </tr>
        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Date Received</td>
            <td style="border:1px solid #ddd;">{item['date_received'] or 'N/A'}</td>
        </tr>
        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Priority</td>
            <td style="border:1px solid #ddd; color:{priority_color}; font-weight:bold;">{item['priority'] or 'Normal'}</td>
        </tr>
        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Initial Review Due Date</td>
            <td style="border:1px solid #ddd; color:#2980b9; font-weight:bold;">{reviewer_due_date or 'N/A'}</td>
        </tr>
        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">QCR Due Date</td>
            <td style="border:1px solid #ddd; color:#27ae60; font-weight:bold;">{qcr_due_date or 'N/A'}</td>
        </tr>
        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Contractor Due Date</td>
            <td style="border:1px solid #ddd; color:#c0392b; font-weight:bold;">{item['due_date'] or 'N/A'}</td>
        </tr>
        {reviewer_row_html}
        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">QC Reviewer</td>
            <td style="border:1px solid #ddd;">{item['qcr_name'] or 'Not assigned'}</td>
        </tr>
    </table>

    <!-- QCR NOTE -->
    <p style="margin-top:16px; font-size:13px; color:#555;">
        {qcr_note_text}
    </p>

    <!-- FOOTER -->
    <p style="margin-top:20px; font-size:12px; color:#777;">
        <em>This message was automatically generated. If you believe you received this by mistake, please contact the project administrator.</em>
    </p>

</div>"""
                else:
                    # Form generation failed, log and skip to server-based
                    print(f"Warning: Could not generate form file for {reviewer['reviewer_name']}: {form_result.get('error')}")
                    use_file_form = False
            
            if not use_file_form:
                # Use server-based form (fallback)
                app_host = get_app_host()
                review_url = f"{app_host}/respond/multi-reviewer?token={token}"
                
                html_body = f"""<div style="font-family:Segoe UI, Helvetica, Arial, sans-serif; color:#333; font-size:14px; line-height:1.5;">

    <!-- HEADER -->
    <h2 style="color:#444; margin-bottom:6px;">
        [LEB] {item['identifier']} - Assigned to You
    </h2>

    <p style="margin-top:0; font-size:13px; color:#666;">
        You have been assigned a new review task. Please review the details below.
    </p>

    <!-- BLUEBEAM INSTRUCTIONS -->
    <div style="margin:15px 0; padding:15px; background:#dbeafe; border:1px solid #3b82f6; border-radius:8px;">
        <div style="font-size:14px; color:#1e40af; font-weight:bold;">📐 Markups Instructions</div>
        <div style="font-size:13px; color:#1e40af; margin-top:8px;">
            <strong>Provide markups in the corresponding Bluebeam session.</strong><br/>
            Do not attach files to your response. All markups should be completed in the shared Bluebeam Studio session.
        </div>
    </div>

    <!-- ACTION BUTTON -->
    <div style="margin:20px 0; text-align:center;">
        <table cellpadding="0" cellspacing="0" border="0" style="margin:0 auto;">
            <tr>
                <td align="center" bgcolor="#27ae60" style="background:#27ae60; border-radius:8px; padding:0;">
                    <a href="{review_url}" target="_blank"
                       style="background:#27ae60; color:#ffffff; display:inline-block; font-family:Segoe UI,Arial,sans-serif;
                              font-size:16px; font-weight:bold; line-height:50px; text-align:center; text-decoration:none;
                              width:280px; -webkit-text-size-adjust:none; border-radius:8px;">
                        OPEN RESPONSE FORM
                    </a>
                </td>
            </tr>
        </table>
    </div>

    <!-- INFO TABLE -->
    <table cellpadding="8" cellspacing="0" width="100%" style="border-collapse:collapse; margin-top:10px;">
        <tr>
            <td colspan="2" style="background:#f2f2f2; font-weight:bold; border:1px solid #ddd;">
                Item Information
            </td>
        </tr>
        <tr>
            <td style="width:160px; border:1px solid #ddd; font-weight:bold;">Type</td>
            <td style="border:1px solid #ddd;">{item['type']}</td>
        </tr>
        <tr>
            <td style="width:160px; border:1px solid #ddd; font-weight:bold;">Identifier</td>
            <td style="border:1px solid #ddd;">{item['identifier']}</td>
        </tr>
        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Title</td>
            <td style="border:1px solid #ddd;">{item['title'] or 'N/A'}</td>
        </tr>
        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Date Received</td>
            <td style="border:1px solid #ddd;">{item['date_received'] or 'N/A'}</td>
        </tr>
        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Priority</td>
            <td style="border:1px solid #ddd; color:{priority_color}; font-weight:bold;">{item['priority'] or 'Normal'}</td>
        </tr>
        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Initial Review Due Date</td>
            <td style="border:1px solid #ddd; color:#2980b9; font-weight:bold;">{reviewer_due_date or 'N/A'}</td>
        </tr>
        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">QCR Due Date</td>
            <td style="border:1px solid #ddd; color:#27ae60; font-weight:bold;">{qcr_due_date or 'N/A'}</td>
        </tr>
        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Contractor Due Date</td>
            <td style="border:1px solid #ddd; color:#c0392b; font-weight:bold;">{item['due_date'] or 'N/A'}</td>
        </tr>
        {reviewer_row_html}
        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">QC Reviewer</td>
            <td style="border:1px solid #ddd;">{item['qcr_name'] or 'Not assigned'}</td>
        </tr>
    </table>

    <!-- QCR NOTE -->
    <p style="margin-top:16px; font-size:13px; color:#555;">
        {qcr_note_text}
    </p>

    <!-- FOOTER -->
    <p style="margin-top:20px; font-size:12px; color:#777;">
        <em>This message was automatically generated. If you believe you received this by mistake, please contact the project administrator.</em>
    </p>

</div>"""
            
            # Send via Outlook (runs for BOTH HTA and server-based forms)
            # This code is OUTSIDE the 'if not use_file_form:' block
            pythoncom.CoInitialize()
            try:
                outlook = win32com.client.Dispatch("Outlook.Application")
                mail = outlook.CreateItem(0)
                mail.To = reviewer['reviewer_email']
                # CC the QCR only for single reviewer (multi-reviewer gets separate notification)
                if item['qcr_email'] and is_single_reviewer:
                    mail.CC = item['qcr_email']
                mail.Subject = subject
                mail.HTMLBody = html_body
                mail.Send()
                
                # Update sent timestamp
                cursor.execute('''
                    UPDATE item_reviewers SET email_sent_at = ? WHERE id = ?
                ''', (datetime.now().isoformat(), reviewer['id']))
                conn.commit()
                
                sent_count += 1
            finally:
                pythoncom.CoUninitialize()
                
        except Exception as e:
            errors.append(f"{reviewer['reviewer_name']}: {str(e)}")
    
    # Send separate notification email to QCR (only for multi-reviewer)
    qcr_email_sent = False
    if item['qcr_email'] and sent_count > 0 and not is_single_reviewer:
        try:
            reviewer_names_list = "<br>".join([f"• {r['reviewer_name']}" for r in reviewers])
            
            qcr_subject = f"[LEB] {item['identifier']} – Assigned to You for QC Review"
            
            qcr_html_body = f"""<div style="font-family:Segoe UI, Helvetica, Arial, sans-serif; color:#333; font-size:14px; line-height:1.6;">

    <h2 style="color:#2563eb; margin-bottom:10px;">[LEB] {item['identifier']} – Assigned to You for QC Review</h2>
    
    <div style="background:#dbeafe; border:1px solid #3b82f6; border-radius:8px; padding:15px; margin:15px 0;">
        <p style="margin:0; font-size:14px; color:#1e40af;">
            <strong>📋 You will be notified once all reviewers have submitted their responses.</strong>
        </p>
    </div>
    
    <p>You have been assigned as the <strong>QC Reviewer</strong> for the following item. The initial reviewers listed below have been notified and are currently working on their responses.</p>
    
    <div style="background:#f8fafc; border-radius:8px; padding:15px; margin:15px 0; border:1px solid #e5e7eb;">
        <h4 style="margin:0 0 10px 0; color:#374151;">👥 Assigned Reviewers ({len(reviewers)})</h4>
        <div style="color:#4b5563;">
            {reviewer_names_list}
        </div>
    </div>
    
    <table cellpadding="8" cellspacing="0" width="100%" style="border-collapse:collapse; margin-top:15px;">
        <tr>
            <td colspan="2" style="background:#f2f2f2; font-weight:bold; border:1px solid #ddd;">
                Item Information
            </td>
        </tr>
        <tr>
            <td style="width:160px; border:1px solid #ddd; font-weight:bold;">Type</td>
            <td style="border:1px solid #ddd;">{item['type']}</td>
        </tr>
        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Identifier</td>
            <td style="border:1px solid #ddd;">{item['identifier']}</td>
        </tr>
        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Title</td>
            <td style="border:1px solid #ddd;">{item['title'] or 'N/A'}</td>
        </tr>
        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Date Received</td>
            <td style="border:1px solid #ddd;">{item['date_received'] or 'N/A'}</td>
        </tr>
        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Priority</td>
            <td style="border:1px solid #ddd; color:{priority_color}; font-weight:bold;">{item['priority'] or 'Normal'}</td>
        </tr>
        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">QCR Due Date</td>
            <td style="border:1px solid #ddd; color:#27ae60; font-weight:bold;">{qcr_due_date or 'N/A'}</td>
        </tr>
        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Contractor Due Date</td>
            <td style="border:1px solid #ddd; color:#c0392b; font-weight:bold;">{item['due_date'] or 'N/A'}</td>
        </tr>
    </table>
    
    <p style="margin-top:20px; font-size:13px; color:#666;">
        Once all reviewers have submitted their responses, you will receive another email with a link to complete your QC review.
    </p>
    
    <p style="margin-top:20px; font-size:12px; color:#777;">
        <em>This message was automatically generated.</em>
    </p>

</div>"""
            
            pythoncom.CoInitialize()
            try:
                outlook = win32com.client.Dispatch("Outlook.Application")
                mail = outlook.CreateItem(0)
                mail.To = item['qcr_email']
                mail.Subject = qcr_subject
                mail.HTMLBody = qcr_html_body
                mail.Send()
                qcr_email_sent = True
            finally:
                pythoncom.CoUninitialize()
        except Exception as e:
            errors.append(f"QCR ({item['qcr_name']}): {str(e)}")
    
    # Update item status
    if sent_count > 0:
        cursor.execute('''
            UPDATE item SET 
                status = 'Assigned',
                reviewer_response_status = 'Emails Sent'
            WHERE id = ?
        ''', (item_id,))
        conn.commit()
    
    conn.close()
    
    if errors:
        return {
            'success': sent_count > 0,
            'sent_count': sent_count,
            'qcr_notified': qcr_email_sent if not is_single_reviewer else True,
            'errors': errors,
            'message': f'Sent {sent_count} reviewer emails with {len(errors)} errors'
        }
    
    if is_single_reviewer:
        return {'success': True, 'sent_count': sent_count, 'qcr_notified': True, 'message': f'Sent {sent_count} reviewer email (QCR CC\'d)'}
    else:
        return {'success': True, 'sent_count': sent_count, 'qcr_notified': qcr_email_sent, 'message': f'Sent {sent_count} reviewer emails + separate QCR notification'}


def generate_multi_reviewer_qcr_form(item_id):
    """Generate an HTA form for the QCR - uses appropriate template based on reviewer count.
    
    Args:
        item_id: The item ID
    
    Returns:
        dict with 'success', 'path' or 'error'
    """
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT i.*, qcr.display_name as qcr_name, qcr.email as qcr_email
        FROM item i
        LEFT JOIN user qcr ON i.qcr_id = qcr.id
        WHERE i.id = ?
    ''', (item_id,))
    item = cursor.fetchone()
    
    if not item:
        conn.close()
        return {'success': False, 'error': 'Item not found'}
    
    # Convert to dict for easier handling
    item = dict(item)
    
    if not item['folder_link']:
        conn.close()
        return {'success': False, 'error': 'Item has no folder assigned'}
    
    # Get all reviewer responses
    cursor.execute('SELECT * FROM item_reviewers WHERE item_id = ?', (item_id,))
    reviewers = cursor.fetchall()
    conn.close()
    
    # Check if truly multi-reviewer (more than 1 reviewer)
    is_multi_reviewer = len(reviewers) > 1
    
    # Load appropriate template based on reviewer count
    if is_multi_reviewer:
        # Multi-reviewer: use multi-reviewer QCR template
        template_path = TEMPLATES_DIR / "_MULTI_REVIEWER_QCR_TEMPLATE.hta"
        if not template_path.exists():
            return {'success': False, 'error': 'Multi-reviewer QCR form template not found'}
    else:
        # Single reviewer: use regular QCR template
        template_path = TEMPLATES_DIR / "_QCR_FORM_TEMPLATE_v3.hta"
        if not template_path.exists():
            return {'success': False, 'error': 'QCR form template not found'}
    
    with open(template_path, 'r', encoding='utf-8') as f:
        template = f.read()
    
    # Escape special characters for JavaScript embedding
    def js_escape(s):
        if not s:
            return ''
        return s.replace('\\', '\\\\').replace('"', '\\"').replace("'", "\\'").replace('\n', '\\n').replace('\r', '')
    
    # Start with the template
    html = template
    
    if is_multi_reviewer:
        # Build multi-reviewer specific HTML sections
        reviewer_html = ""
        reviewer_checkboxes_html = ""
        reviewers_json_list = []
        
        for idx, r in enumerate(reviewers, 1):
            category = r['response_category'] or 'N/A'
            notes = r['internal_notes'] or ''
            
            notes_section = ""
            if notes:
                notes_section = f'''
            <div class="internal-notes-box">
                <h5>Suggested Response for QC Reviewer (Team Only):</h5>
                <div class="internal-notes-content">{notes}</div>
            </div>'''
            else:
                notes_section = '<p class="no-notes">No suggested response provided.</p>'
            
            reviewer_html += f'''
        <div class="reviewer-response-box">
            <div class="reviewer-header">
                <span class="reviewer-badge">{idx}</span>
                <span class="reviewer-name">{r['reviewer_name']}</span>
                <span class="category-chip">{category}</span>
            </div>
            {notes_section}
        </div>'''
            
            # Build checkbox for send-back selection
            reviewer_checkboxes_html += f'''
        <label style="display:flex; align-items:center; padding:8px; background:white; border-radius:6px; margin-bottom:6px; cursor:pointer;">
            <input type="checkbox" name="sendback_reviewers" value="{r['id']}" checked style="width:18px; height:18px; margin-right:10px;">
            <span style="flex:1;">{r['reviewer_name']}</span>
            <span style="background:#e0e7ff; color:#3730a3; padding:2px 8px; border-radius:10px; font-size:11px;">{category}</span>
        </label>'''
            
            # Build JSON data for JavaScript
            reviewers_json_list.append({
                'id': r['id'],
                'name': r['reviewer_name'],
                'email': r['reviewer_email'],
                'category': category
            })
        
        # Convert reviewers list to JSON for JavaScript
        reviewers_json = json.dumps(reviewers_json_list)
        
        # Replace multi-reviewer specific placeholders
        html = html.replace('{{REVIEWER_COUNT}}', str(len(reviewers)))
        html = html.replace('{{REVIEWER_RESPONSES_HTML}}', reviewer_html)
        html = html.replace('{{REVIEWER_CHECKBOXES_HTML}}', reviewer_checkboxes_html)
        html = html.replace('{{REVIEWERS_JSON}}', reviewers_json)
    else:
        # Single reviewer from item_reviewers table - populate single-reviewer template fields
        r = reviewers[0]  # Get the single reviewer
        reviewer_notes = r['internal_notes'] or 'No comments provided'
        reviewer_category = r['response_category'] or 'N/A'
        
        # For single reviewer using item_reviewers, we don't have selected_files 
        # (Bluebeam-based workflow), so indicate that
        reviewer_selected_files_text = 'Files selected in Bluebeam session'
        reviewer_selected_files_js = '[]'  # Empty array for JS
        
        # Replace single-reviewer specific placeholders
        html = template.replace('{{RESPONSE_VERSION}}', '1')
        html = template.replace('{{REVIEWER_NAME}}', r['reviewer_name'] or 'N/A')
        html = template.replace('{{REVIEWER_RESPONSE_CATEGORY}}', reviewer_category)
        html = html.replace('{{REVIEWER_NOTES}}', js_escape(reviewer_notes))
        html = html.replace('{{REVIEWER_SELECTED_FILES_TEXT}}', reviewer_selected_files_text)
        html = html.replace('{{REVIEWER_SELECTED_FILES_JS}}', reviewer_selected_files_js)
        html = html.replace('{{REVIEWER_INTERNAL_NOTES}}', js_escape(reviewer_notes))
        html = html.replace('{{REVIEWER_INTERNAL_NOTES_DISPLAY}}', 'block' if reviewer_notes else 'none')
    
    # Replace common placeholders (used by both templates)
    html = html.replace('{{ITEM_ID}}', str(item['id']))
    html = html.replace('{{ITEM_TYPE}}', item['type'] or '')
    html = html.replace('{{ITEM_IDENTIFIER}}', item['identifier'] or '')
    html = html.replace('{{ITEM_TITLE}}', js_escape(item['title']) or 'N/A')
    html = html.replace('{{DATE_RECEIVED}}', item['date_received'] or 'N/A')
    html = html.replace('{{PRIORITY}}', item['priority'] or 'Normal')
    html = html.replace('{{QCR_NAME}}', item['qcr_name'] or 'N/A')
    html = html.replace('{{QCR_EMAIL}}', item['qcr_email'] or '')
    html = html.replace('{{QCR_DUE_DATE}}', item['qcr_due_date'] or 'N/A')
    html = html.replace('{{CONTRACTOR_DUE_DATE}}', item['due_date'] or 'N/A')
    html = html.replace('{{FOLDER_PATH}}', js_escape(item['folder_link']))
    html = html.replace('{{FOLDER_PATH_RAW}}', item['folder_link'] or '')
    html = html.replace('{{RFI_QUESTION}}', js_escape(item.get('rfi_question', '') or 'N/A'))
    html = html.replace('{{TOKEN}}', item['email_token_qcr'] or '')
    
    # Save to item folder
    folder_path = Path(item['folder_link'])
    try:
        folder_path.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        return {'success': False, 'error': f'Failed to create folder: {e}'}
    
    # Create Responses folder
    responses_folder = folder_path / "Responses"
    try:
        responses_folder.mkdir(exist_ok=True)
    except Exception as e:
        return {'success': False, 'error': f'Failed to create Responses folder: {e}'}
    
    # Generate filename
    safe_name = "".join(c for c in (item['qcr_name'] or 'QCR') if c.isalnum() or c in (' ', '-', '_')).strip()
    file_name = f"_QCR_Review_Form_{safe_name}.hta"
    file_path = folder_path / file_name
    
    try:
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(html)
        return {'success': True, 'path': str(file_path)}
    except Exception as e:
        return {'success': False, 'error': f'Failed to write form: {e}'}


def send_multi_reviewer_qcr_email(item_id):
    """Send QCR assignment email for multi-reviewer items."""
    if not HAS_WIN32COM:
        return {'success': False, 'error': 'Outlook not available'}
    
    conn = get_db()
    cursor = conn.cursor()
    
    # ATOMIC CHECK-AND-CLAIM: Update qcr_email_sent_at ONLY if it's NULL
    # This prevents race conditions where multiple watcher cycles try to send
    cursor.execute('''
        UPDATE item SET qcr_email_sent_at = ? 
        WHERE id = ? AND qcr_email_sent_at IS NULL
    ''', (datetime.now().isoformat(), item_id))
    
    if cursor.rowcount == 0:
        # Another process already claimed this - qcr_email_sent_at was not NULL
        conn.close()
        print(f"  [QCR Email] Skipping duplicate - another process already sending/sent for item {item_id}")
        return {'success': True, 'message': 'QCR email already being sent', 'skipped': True}
    
    conn.commit()  # Commit the claim
    
    # Get item with QCR info
    cursor.execute('''
        SELECT i.*, qcr.email as qcr_email, qcr.display_name as qcr_name
        FROM item i
        LEFT JOIN user qcr ON i.qcr_id = qcr.id
        WHERE i.id = ?
    ''', (item_id,))
    item = cursor.fetchone()
    
    if not item:
        conn.close()
        return {'success': False, 'error': 'Item not found'}
    
    if not item['qcr_email']:
        conn.close()
        return {'success': False, 'error': 'No QCR assigned'}
    
    # Generate token if not exists
    token = item['email_token_qcr']
    if not token:
        token = generate_token()
        cursor.execute('UPDATE item SET email_token_qcr = ? WHERE id = ?', (token, item_id))
        conn.commit()
    
    # Get reviewer responses
    cursor.execute('SELECT * FROM item_reviewers WHERE item_id = ?', (item_id,))
    reviewers = cursor.fetchall()
    
    app_host = get_app_host()
    
    # Check if using local mode with file-based forms
    folder_path = item['folder_link'] or 'Not set'
    use_file_form = is_local_mode() and folder_path != 'Not set'
    
    # Generate HTA form if in local mode
    hta_form_path = None
    if use_file_form:
        form_result = generate_multi_reviewer_qcr_form(item_id)
        if form_result.get('success'):
            hta_form_path = form_result['path']
        else:
            print(f"  Warning: Could not generate QCR HTA form: {form_result.get('error')}")
    
    # Build reviewer names for display (each on new line)
    reviewer_names_html = "<br>".join([r['reviewer_name'] for r in reviewers])
    
    # Build reviewer summary table with full comments (not truncated)
    reviewer_summary = ""
    for r in reviewers:
        category = r['response_category'] or 'N/A'
        # Show full internal notes without truncation
        internal_notes = r['internal_notes'] or ''
        notes_html = internal_notes.replace('\n', '<br>') if internal_notes else '<span style="color:#999;">No comments</span>'
        reviewer_summary += f"""
        <tr>
            <td style="padding:10px; border-bottom:1px solid #e5e7eb; vertical-align:top;">{r['reviewer_name']}</td>
            <td style="padding:10px; border-bottom:1px solid #e5e7eb; vertical-align:top;"><span style="background:#e0e7ff; color:#3730a3; padding:2px 8px; border-radius:10px; font-size:12px;">{category}</span></td>
            <td style="padding:10px; border-bottom:1px solid #e5e7eb; font-size:13px; color:#444; vertical-align:top;">{notes_html}</td>
        </tr>
"""
    
    # Build action section based on mode
    if use_file_form and hta_form_path:
        form_file_link = f'file:///{hta_form_path.replace(chr(92), "/")}'
        action_button_html = f"""
    <!-- ACTION BUTTON AT TOP - HTA FORM -->
    <div style="margin:20px 0; text-align:center;">
        <table cellpadding="0" cellspacing="0" border="0" style="margin:0 auto;">
            <tr>
                <td align="center" bgcolor="#10b981" style="background:#10b981; border-radius:8px; padding:0;">
                    <a href="{form_file_link}" target="_blank"
                       style="background:#10b981; color:#ffffff; display:inline-block; font-family:Segoe UI,Arial,sans-serif;
                              font-size:16px; font-weight:bold; line-height:50px; text-align:center; text-decoration:none;
                              width:320px; -webkit-text-size-adjust:none; border-radius:8px;">
                        OPEN QC REVIEW FORM
                    </a>
                </td>
            </tr>
        </table>
    </div>

    <!-- INSTRUCTIONS FOR HTA -->
    <div style="margin:20px 0; padding:15px; background:#e8f5e9; border:1px solid #4caf50; border-radius:8px;">
        <div style="font-size:14px; color:#2e7d32;">
            <strong>Instructions:</strong>
            <ol style="margin:8px 0 0 0; padding-left:20px;">
                <li>Click the green button above to open the QC review form</li>
                <li>If prompted, select <strong>"Microsoft (R) HTML Application host"</strong> - choose <strong>Open</strong>, do NOT save the file</li>
                <li>Review all responses and select your final action (Complete or Send Back)</li>
                <li>Click <strong>Submit Decision</strong> - your decision will be saved automatically</li>
            </ol>
        </div>
    </div>"""
    else:
        review_url = f"{app_host}/respond/multi-qcr?token={token}"
        action_button_html = f"""
    <!-- ACTION BUTTON AT TOP - SERVER FORM -->
    <div style="margin:20px 0; text-align:center;">
        <table cellpadding="0" cellspacing="0" border="0" style="margin:0 auto;">
            <tr>
                <td align="center" bgcolor="#10b981" style="background:#10b981; border-radius:8px; padding:0;">
                    <a href="{review_url}" target="_blank"
                       style="background:#10b981; color:#ffffff; display:inline-block; font-family:Segoe UI,Arial,sans-serif;
                              font-size:16px; font-weight:bold; line-height:50px; text-align:center; text-decoration:none;
                              width:320px; -webkit-text-size-adjust:none; border-radius:8px;">
                        COMPLETE QC REVIEW
                    </a>
                </td>
            </tr>
        </table>
    </div>"""
    
    # Priority color
    priority = item['priority'] or 'Normal'
    priority_color = '#e67e22' if priority == 'Medium' else '#c0392b' if priority == 'High' else '#27ae60'
    
    html_body = f"""<div style="font-family:Segoe UI, Helvetica, Arial, sans-serif; color:#333; font-size:14px; line-height:1.5;">
    <h2 style="color:#444; margin-bottom:6px;">[LEB] {item['identifier']} - QC Review Ready</h2>
    <p style="color:#666; margin-top:0;">All {len(reviewers)} reviewers have submitted their responses. Please complete the QC review.</p>
    
    {action_button_html}
    
    <!-- Bluebeam Notice -->
    <div style="background:#dbeafe; border:1px solid #3b82f6; border-radius:8px; padding:15px; margin:15px 0;">
        <div style="font-size:13px; color:#1e40af;">
            <strong>Markups Location:</strong> All reviewer markups are in the <strong>Bluebeam Studio session</strong> for this item.
        </div>
    </div>
    
    <!-- Item Info -->
    <table cellpadding="8" cellspacing="0" width="100%" style="border-collapse:collapse; margin-top:10px;">
        <tr>
            <td colspan="2" style="background:#f2f2f2; font-weight:bold; border:1px solid #ddd;">
                Item Information
            </td>
        </tr>
        <tr>
            <td style="width:160px; border:1px solid #ddd; font-weight:bold;">Type</td>
            <td style="border:1px solid #ddd;">{item['type']}</td>
        </tr>
        <tr>
            <td style="width:160px; border:1px solid #ddd; font-weight:bold;">Identifier</td>
            <td style="border:1px solid #ddd;">{item['identifier']}</td>
        </tr>
        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Title</td>
            <td style="border:1px solid #ddd;">{item['title'] or 'N/A'}</td>
        </tr>
        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Date Received</td>
            <td style="border:1px solid #ddd;">{item['date_received'] or 'N/A'}</td>
        </tr>
        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Priority</td>
            <td style="border:1px solid #ddd; color:{priority_color}; font-weight:bold;">{priority}</td>
        </tr>
        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Initial Review Due Date</td>
            <td style="border:1px solid #ddd; color:#2980b9; font-weight:bold;">{item['initial_reviewer_due_date'] or 'N/A'}</td>
        </tr>
        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">QCR Due Date</td>
            <td style="border:1px solid #ddd; color:#27ae60; font-weight:bold;">{item['qcr_due_date'] or 'N/A'}</td>
        </tr>
        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Contractor Due Date</td>
            <td style="border:1px solid #ddd; color:#c0392b; font-weight:bold;">{item['due_date'] or 'N/A'}</td>
        </tr>
        <tr>
            <td style="border:1px solid #ddd; font-weight:bold;">Initial Reviewer(s)</td>
            <td style="border:1px solid #ddd;">{reviewer_names_html}</td>
        </tr>
    </table>
    
    <!-- Reviewer Responses Table -->
    <div style="background:#f0fdf4; border:1px solid #86efac; border-radius:8px; padding:15px; margin:15px 0;">
        <div style="font-size:14px; color:#166534; font-weight:bold; margin-bottom:10px;">Reviewer Responses ({len(reviewers)})</div>
        <table style="width:100%; border-collapse:collapse; font-size:13px;">
            <tr style="background:#ecfdf5;">
                <th style="padding:8px; text-align:left; border-bottom:2px solid #86efac;">Reviewer</th>
                <th style="padding:8px; text-align:left; border-bottom:2px solid #86efac;">Category</th>
                <th style="padding:8px; text-align:left; border-bottom:2px solid #86efac;">Suggested Response</th>
            </tr>
            {reviewer_summary}
        </table>
    </div>
    
    <p style="font-size:12px; color:#888;">As QC Reviewer, you will write the final response to be sent to the contractor.</p>
    
    <!-- CC Note -->
    <p style="margin-top:16px; font-size:13px; color:#555;">
        All reviewers have been CC'd on this email.
    </p>
</div>"""
    
    try:
        pythoncom.CoInitialize()
        try:
            outlook = win32com.client.Dispatch("Outlook.Application")
            mail = outlook.CreateItem(0)
            mail.To = item['qcr_email']
            
            # CC all the reviewers
            reviewer_emails = [r['reviewer_email'] for r in reviewers if r['reviewer_email']]
            if reviewer_emails:
                mail.CC = "; ".join(reviewer_emails)
            
            mail.Subject = f"[LEB] {item['identifier']} - QC Review Ready ({len(reviewers)} Reviewers)"
            mail.HTMLBody = html_body
            
            # Note: HTA files are not attached as they get blocked by email providers (Gmail, etc.)
            # The email body contains a link/button to access the form directly from the shared folder
            
            mail.Send()
            
            # Just update qcr_response_status (qcr_email_sent_at was already set atomically at start)
            cursor.execute('''
                UPDATE item SET qcr_response_status = 'Email Sent' WHERE id = ?
            ''', (item_id,))
            conn.commit()
        finally:
            pythoncom.CoUninitialize()
        
        conn.close()
        return {'success': True, 'message': 'QCR email sent'}
        
    except Exception as e:
        # Reset qcr_email_sent_at since we failed to send
        try:
            conn2 = get_db()
            cursor2 = conn2.cursor()
            cursor2.execute('UPDATE item SET qcr_email_sent_at = NULL WHERE id = ?', (item_id,))
            conn2.commit()
            conn2.close()
            print(f"  [QCR Email] Failed to send, reset qcr_email_sent_at for item {item_id}")
        except:
            pass  # Best effort reset
        conn.close()
        return {'success': False, 'error': str(e)}

def send_multi_reviewer_sendback_emails(item_id, feedback, reviewer_ids=None):
    """Send emails to selected reviewers when QCR sends back for revision.
    
    Args:
        item_id: The item ID
        feedback: The QCR's feedback message
        reviewer_ids: List of specific reviewer IDs to send to (None = all reviewers)
    """
    if not HAS_WIN32COM:
        return {'success': False, 'error': 'Outlook not available'}
    
    conn = get_db()
    cursor = conn.cursor()
    
    # Get item info with QCR email
    cursor.execute('''
        SELECT i.*, qcr.display_name as qcr_name, qcr.email as qcr_email
        FROM item i
        LEFT JOIN user qcr ON i.qcr_id = qcr.id
        WHERE i.id = ?
    ''', (item_id,))
    item = cursor.fetchone()
    
    if not item:
        conn.close()
        return {'success': False, 'error': 'Item not found'}
    
    qcr_email = item['qcr_email']  # Store QCR email for CC
    
    # Get selected reviewers or all reviewers
    if reviewer_ids:
        placeholders = ','.join(['?' for _ in reviewer_ids])
        cursor.execute(f'SELECT * FROM item_reviewers WHERE item_id = ? AND id IN ({placeholders})', 
                      [item_id] + list(reviewer_ids))
    else:
        cursor.execute('SELECT * FROM item_reviewers WHERE item_id = ?', (item_id,))
    reviewers = cursor.fetchall()
    
    if not reviewers:
        conn.close()
        return {'success': False, 'error': 'No reviewers found'}
    
    # Mark only selected reviewers as needing response, clear their response
    # Other reviewers keep their current response
    selected_ids = [r['id'] for r in reviewers]
    for rid in selected_ids:
        cursor.execute('''
            UPDATE item_reviewers SET 
                needs_response = 1,
                response_at = NULL,
                response_category = NULL,
                internal_notes = NULL
            WHERE id = ?
        ''', (rid,))
    
    # Mark non-selected reviewers as NOT needing response (they keep their existing response)
    if reviewer_ids:
        cursor.execute('''
            UPDATE item_reviewers SET needs_response = 0
            WHERE item_id = ? AND id NOT IN ({})
        '''.format(','.join(['?' for _ in selected_ids])), [item_id] + selected_ids)
    
    # Update item status (use 'In Review' as 'Sent Back' is not in status constraint)
    cursor.execute('''
        UPDATE item SET 
            status = 'In Review',
            reviewer_response_status = 'Revision Requested',
            qcr_email_sent_at = NULL,
            qcr_response_status = NULL,
            qcr_action = 'Send Back',
            qcr_notes = ?
        WHERE id = ?
    ''', (feedback, item_id))
    
    conn.commit()
    
    # Determine if using local mode
    folder_path = item['folder_link'] or 'Not set'
    use_file_form = is_local_mode() and folder_path != 'Not set'
    
    app_host = get_app_host()
    sent_count = 0
    errors = []
    
    for reviewer in reviewers:
        try:
            # Generate new token for revision
            token = generate_token()
            cursor.execute('UPDATE item_reviewers SET email_token = ?, response_version = response_version + 1 WHERE id = ?', 
                          (token, reviewer['id']))
            conn.commit()
            
            subject = f"[LEB] {item['identifier']} – REVISION REQUESTED"
            
            if use_file_form:
                # Generate HTA form for this reviewer
                reviewer_record = dict(reviewer)
                reviewer_record['email_token'] = token
                form_result = generate_multi_reviewer_form(item_id, reviewer_record)
                
                if form_result.get('success'):
                    form_file_path = form_result['path']
                    form_file_link = f'file:///{form_file_path.replace(chr(92), "/")}'
                    
                    html_body = f"""<div style="font-family:Segoe UI, Helvetica, Arial, sans-serif; color:#333; font-size:14px; line-height:1.5;">
    <h2 style="color:#c0392b; margin-bottom:6px;">[LEB] {item['identifier']} - REVISION REQUESTED</h2>
    <p style="color:#666; margin-top:0;">The QC Reviewer has requested revisions to your response.</p>
    
    <!-- ACTION BUTTON -->
    <div style="margin:20px 0; text-align:center;">
        <table cellpadding="0" cellspacing="0" border="0" style="margin:0 auto;">
            <tr>
                <td align="center" bgcolor="#f59e0b" style="background:#f59e0b; border-radius:8px; padding:0;">
                    <a href="{form_file_link}" target="_blank"
                       style="background:#f59e0b; color:#ffffff; display:inline-block; font-family:Segoe UI,Arial,sans-serif;
                              font-size:16px; font-weight:bold; line-height:50px; text-align:center; text-decoration:none;
                              width:320px; -webkit-text-size-adjust:none; border-radius:8px;">
                        📝 SUBMIT REVISED RESPONSE
                    </a>
                </td>
            </tr>
        </table>
    </div>
    
    <!-- INSTRUCTIONS FOR HTA -->
    <div style="margin:20px 0; padding:15px; background:#fef3c7; border:1px solid #f59e0b; border-radius:8px;">
        <div style="font-size:14px; color:#92400e;">
            <strong>Instructions:</strong>
            <ol style="margin:8px 0 0 0; padding-left:20px;">
                <li>Click the orange button above to open the response form</li>
                <li>If prompted, select <strong>"Microsoft (R) HTML Application host"</strong> - choose <strong>Open</strong></li>
                <li>Review the feedback and select your updated response category</li>
                <li>Click <strong>Submit Response</strong></li>
            </ol>
        </div>
    </div>
    
    <div style="background:#fef2f2; border:1px solid #fecaca; border-radius:8px; padding:15px; margin:15px 0;">
        <div style="font-size:14px; color:#991b1b; font-weight:bold;">↩️ Feedback from {item['qcr_name'] or 'QC Reviewer'}</div>
        <div style="margin-top:10px; padding:10px; background:white; border-radius:4px; color:#991b1b;">
            {feedback}
        </div>
    </div>
    
    <div style="background:#f8f9fa; border-radius:8px; padding:15px; margin:15px 0;">
        <table style="width:100%; border-collapse:collapse; font-size:13px;">
            <tr><td style="padding:5px 0; color:#666; width:120px;">Type:</td><td style="font-weight:600;">{item['type']}</td></tr>
            <tr><td style="padding:5px 0; color:#666;">Identifier:</td><td style="font-weight:600;">{item['identifier']}</td></tr>
            <tr><td style="padding:5px 0; color:#666;">Title:</td><td>{item['title'] or 'N/A'}</td></tr>
        </table>
    </div>
    
    <p style="font-size:12px; color:#888;">Please review the feedback and submit an updated response.</p>
</div>"""
                else:
                    # Fallback to server form if HTA generation fails
                    use_file_form = False
            
            if not use_file_form:
                review_url = f"{app_host}/respond/multi-reviewer?token={token}"
                
                html_body = f"""<div style="font-family:Segoe UI, Helvetica, Arial, sans-serif; color:#333; font-size:14px; line-height:1.5;">
    <h2 style="color:#c0392b; margin-bottom:6px;">[LEB] {item['identifier']} - REVISION REQUESTED</h2>
    <p style="color:#666; margin-top:0;">The QC Reviewer has requested revisions to your response.</p>
    
    <div style="background:#fef2f2; border:1px solid #fecaca; border-radius:8px; padding:15px; margin:15px 0;">
        <div style="font-size:14px; color:#991b1b; font-weight:bold;">↩️ Feedback from {item['qcr_name'] or 'QC Reviewer'}</div>
        <div style="margin-top:10px; padding:10px; background:white; border-radius:4px; color:#991b1b;">
            {feedback}
        </div>
    </div>
    
    <div style="background:#f8f9fa; border-radius:8px; padding:15px; margin:15px 0;">
        <table style="width:100%; border-collapse:collapse; font-size:13px;">
            <tr><td style="padding:5px 0; color:#666; width:120px;">Type:</td><td style="font-weight:600;">{item['type']}</td></tr>
            <tr><td style="padding:5px 0; color:#666;">Identifier:</td><td style="font-weight:600;">{item['identifier']}</td></tr>
            <tr><td style="padding:5px 0; color:#666;">Title:</td><td>{item['title'] or 'N/A'}</td></tr>
        </table>
    </div>
    
    <div style="margin:20px 0;">
        <a href="{review_url}" style="display:inline-block; padding:12px 24px; background:#f59e0b; color:white; text-decoration:none; border-radius:8px; font-weight:600;">
            📝 Submit Revised Response
        </a>
    </div>
    
    <p style="font-size:12px; color:#888;">Please review the feedback and submit an updated response.</p>
</div>"""
            
            pythoncom.CoInitialize()
            try:
                outlook = win32com.client.Dispatch("Outlook.Application")
                mail = outlook.CreateItem(0)
                mail.To = reviewer['reviewer_email']
                # CC the QCR so they know revision request was sent
                if qcr_email:
                    mail.CC = qcr_email
                mail.Subject = subject
                mail.HTMLBody = html_body
                mail.Send()
                sent_count += 1
            finally:
                pythoncom.CoUninitialize()
                
        except Exception as e:
            errors.append(f"{reviewer['reviewer_name']}: {str(e)}")
    
    conn.close()
    
    return {
        'success': sent_count > 0,
        'sent_count': sent_count,
        'total_selected': len(reviewers),
        'errors': errors if errors else None
    }


def send_multi_reviewer_completion_email(item_id, final_category, final_text):
    """Send completion confirmation email to QCR and ALL reviewers when QCR completes multi-reviewer item.
    
    This goes to everyone (QCR + all reviewers) regardless of who was sent back for revision.
    """
    if not HAS_WIN32COM:
        return {'success': False, 'error': 'Outlook not available'}
    
    conn = get_db()
    cursor = conn.cursor()
    
    # Get item with QCR info
    cursor.execute('''
        SELECT i.*, qcr.display_name as qcr_name, qcr.email as qcr_email
        FROM item i
        LEFT JOIN user qcr ON i.qcr_id = qcr.id
        WHERE i.id = ?
    ''', (item_id,))
    item = cursor.fetchone()
    
    if not item:
        conn.close()
        return {'success': False, 'error': 'Item not found'}
    
    # Get ALL reviewers (not just those who needed response)
    cursor.execute('SELECT * FROM item_reviewers WHERE item_id = ?', (item_id,))
    reviewers = cursor.fetchall()
    conn.close()
    
    if not reviewers:
        return {'success': False, 'error': 'No reviewers found'}
    
    # Build list of initial reviewer names
    initial_reviewers_list = ', '.join([r['reviewer_name'] for r in reviewers])
    
    # Build reviewer summary HTML with full comments
    reviewer_summary = ""
    for idx, r in enumerate(reviewers, 1):
        category = r['response_category'] or 'N/A'
        # Show full internal notes without truncation
        internal_notes = r['internal_notes'] or ''
        notes_html = internal_notes.replace('\n', '<br>') if internal_notes else '<span style="color:#999;">No comments</span>'
        reviewer_summary += f"""
        <tr>
            <td style="padding:10px; border:1px solid #e5e7eb; vertical-align:top;">{idx}</td>
            <td style="padding:10px; border:1px solid #e5e7eb; vertical-align:top;">{r['reviewer_name']}</td>
            <td style="padding:10px; border:1px solid #e5e7eb; vertical-align:top;"><span style="background:#e0e7ff; color:#3730a3; padding:2px 8px; border-radius:10px; font-size:12px;">{category}</span></td>
            <td style="padding:10px; border:1px solid #e5e7eb; vertical-align:top;">{notes_html}</td>
        </tr>"""
    
    subject = f"[LEB] {item['identifier']} – QC Review Complete ✅"
    
    html_body = f"""<div style="font-family:Segoe UI, Helvetica, Arial, sans-serif; color:#333; font-size:14px; line-height:1.6;">
    <div style="background:#059669; color:white; padding:20px; border-radius:8px 8px 0 0;">
        <h2 style="margin:0;">✅ QC Review Complete</h2>
        <p style="margin:8px 0 0 0; opacity:0.9;">{item['identifier']}</p>
    </div>
    
    <div style="background:#f8fafc; padding:20px; border:1px solid #e5e7eb; border-top:none;">
        <p>The QC review for this item has been completed and is ready for final response to the contractor.</p>
        
        <div style="background:white; border-radius:8px; padding:15px; margin:15px 0; border:1px solid #e5e7eb;">
            <table style="width:100%; border-collapse:collapse; font-size:13px;">
                <tr><td style="padding:5px 0; color:#666; width:140px;">Type:</td><td style="font-weight:600;">{item['type']}</td></tr>
                <tr><td style="padding:5px 0; color:#666;">Identifier:</td><td style="font-weight:600;">{item['identifier']}</td></tr>
                <tr><td style="padding:5px 0; color:#666;">Title:</td><td>{item['title'] or 'N/A'}</td></tr>
                <tr><td style="padding:5px 0; color:#666;">Initial Reviewer(s):</td><td>{initial_reviewers_list}</td></tr>
                <tr><td style="padding:5px 0; color:#666;">QC Reviewer:</td><td>{item['qcr_name'] or 'N/A'}</td></tr>
            </table>
        </div>
        
        <div style="background:#f0fdf4; border:1px solid #86efac; border-radius:8px; padding:15px; margin:15px 0;">
            <h4 style="margin:0 0 10px 0; color:#166534;">📋 Final Response</h4>
            <table style="width:100%; border-collapse:collapse; font-size:13px;">
                <tr><td style="padding:5px 0; color:#666; width:140px;">Category:</td><td style="font-weight:600; color:#166534;">{final_category}</td></tr>
                <tr><td style="padding:5px 0; color:#666; vertical-align:top;">Response:</td><td>{final_text.replace(chr(10), '<br>') if final_text else 'N/A'}</td></tr>
            </table>
        </div>
        
        <div style="margin:15px 0;">
            <h4 style="margin:0 0 10px 0; color:#374151;">👥 Reviewer Responses</h4>
            <table style="width:100%; border-collapse:collapse; font-size:13px; background:white;">
                <tr style="background:#f1f5f9;">
                    <th style="padding:10px; border:1px solid #e5e7eb; text-align:left; width:40px;">#</th>
                    <th style="padding:10px; border:1px solid #e5e7eb; text-align:left; width:140px;">Reviewer</th>
                    <th style="padding:10px; border:1px solid #e5e7eb; text-align:left; width:120px;">Category</th>
                    <th style="padding:10px; border:1px solid #e5e7eb; text-align:left;">Comments</th>
                </tr>
                {reviewer_summary}
            </table>
        </div>
        
        <p style="font-size:12px; color:#888; margin-top:20px;">This item is now ready for final response to the contractor.</p>
    </div>
</div>"""
    
    try:
        pythoncom.CoInitialize()
        try:
            outlook = win32com.client.Dispatch("Outlook.Application")
            mail = outlook.CreateItem(0)
            
            # Send to QCR
            mail.To = item['qcr_email'] or ''
            
            # CC all reviewers
            reviewer_emails = [r['reviewer_email'] for r in reviewers if r['reviewer_email']]
            if reviewer_emails:
                mail.CC = "; ".join(reviewer_emails)
            
            mail.Subject = subject
            mail.HTMLBody = html_body
            mail.Send()
        finally:
            pythoncom.CoUninitialize()
        
        return {'success': True, 'message': 'Completion email sent to QCR and all reviewers'}
    except Exception as e:
        return {'success': False, 'error': str(e)}


# =============================================================================
# AUTHENTICATION DECORATOR
# =============================================================================

def login_required(f):
    """Decorator to require login for API endpoints."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': 'Authentication required'}), 401
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    """Decorator to require admin role."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': 'Authentication required'}), 401
        if session.get('role') != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        return f(*args, **kwargs)
    return decorated_function

# Global error handler for API routes - return JSON instead of HTML
@app.errorhandler(500)
def internal_error(error):
    return jsonify({'error': 'Internal server error', 'details': str(error)}), 500

@app.errorhandler(404)
def not_found_error(error):
    # Only return JSON for API routes
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Not found'}), 404
    return send_from_directory('static', 'index.html')

# =============================================================================
# API ROUTES - AUTHENTICATION
# =============================================================================

@app.route('/api/auth/login', methods=['POST'])
def api_login():
    """Login endpoint."""
    data = request.get_json()
    email = data.get('email', '').strip()
    password = data.get('password', '')
    
    if not email or not password:
        return jsonify({'error': 'Email and password required'}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM user WHERE email = ?', (email,))
    user = cursor.fetchone()
    conn.close()
    
    if not user:
        return jsonify({'error': 'Invalid credentials'}), 401
    
    if not bcrypt.checkpw(password.encode('utf-8'), user['password_hash'].encode('utf-8')):
        return jsonify({'error': 'Invalid credentials'}), 401
    
    session.permanent = True
    session['user_id'] = user['id']
    session['email'] = user['email']
    session['display_name'] = user['display_name']
    session['role'] = user['role']
    
    return jsonify({
        'id': user['id'],
        'email': user['email'],
        'display_name': user['display_name'],
        'role': user['role']
    })

@app.route('/api/auth/logout', methods=['POST'])
def api_logout():
    """Logout endpoint."""
    session.clear()
    return jsonify({'success': True})

@app.route('/api/auth/me', methods=['GET'])
def api_me():
    """Get current user info."""
    if 'user_id' not in session:
        return jsonify({'error': 'Not authenticated'}), 401
    
    return jsonify({
        'id': session.get('user_id'),
        'email': session.get('email'),
        'display_name': session.get('display_name'),
        'role': session.get('role')
    })

# =============================================================================
# API ROUTES - USERS
# =============================================================================

@app.route('/api/users', methods=['GET'])
@login_required
def api_get_users():
    """Get all users."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT id, email, display_name, role, created_at FROM user')
    users = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify(users)

@app.route('/api/users', methods=['POST'])
@admin_required
def api_create_user():
    """Create a new user (admin only). Password is optional for team members."""
    data = request.get_json()
    email = (data.get('email') or '').strip()
    password = (data.get('password') or '').strip()
    display_name = (data.get('display_name') or '').strip()
    role = data.get('role', 'user')
    
    if not email:
        return jsonify({'error': 'Email is required'}), 400
    
    if role not in ['admin', 'user']:
        role = 'user'
    
    # Password is optional - if not provided, user cannot log in but can be assigned to items
    password_hash = None
    if password:
        password_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute('''
            INSERT INTO user (email, password_hash, display_name, role)
            VALUES (?, ?, ?, ?)
        ''', (email, password_hash, display_name or email, role))
        conn.commit()
        user_id = cursor.lastrowid
    except sqlite3.IntegrityError as e:
        conn.close()
        error_msg = str(e).lower()
        if 'email' in error_msg or 'unique' in error_msg:
            return jsonify({'error': 'Email already exists'}), 400
        return jsonify({'error': f'Database error: {e}'}), 400
    except Exception as e:
        conn.close()
        print(f"User creation error: {e}")
        return jsonify({'error': f'Error creating user: {e}'}), 500
    
    conn.close()
    return jsonify({'id': user_id, 'email': email, 'display_name': display_name or email, 'role': role})

@app.route('/api/outlook/contacts', methods=['GET'])
@login_required
def api_search_outlook_contacts():
    """Search Outlook contacts/address book for autocomplete."""
    query = request.args.get('q', '').strip()
    
    if not query or len(query) < 2:
        return jsonify([])
    
    if not HAS_WIN32COM:
        return jsonify({'error': 'Outlook not available'}), 503
    
    contacts = []
    
    try:
        pythoncom.CoInitialize()
        try:
            outlook = win32com.client.Dispatch("Outlook.Application")
            namespace = outlook.GetNamespace("MAPI")
            
            # Search using Outlook's address book resolution
            # This searches the Global Address List (GAL), contacts, and recent recipients
            recipient = namespace.CreateRecipient(query)
            recipient.Resolve()
            
            # If exact match found
            if recipient.Resolved:
                try:
                    addr_entry = recipient.AddressEntry
                    email = ''
                    display_name = addr_entry.Name or ''
                    
                    # Try to get SMTP address
                    if addr_entry.Type == 'EX':
                        # Exchange user - get SMTP address
                        try:
                            exch_user = addr_entry.GetExchangeUser()
                            if exch_user:
                                email = exch_user.PrimarySmtpAddress
                                display_name = exch_user.Name or display_name
                        except:
                            pass
                    else:
                        email = addr_entry.Address or ''
                    
                    if email:
                        contacts.append({
                            'email': email,
                            'display_name': display_name,
                            'type': 'resolved'
                        })
                except Exception as e:
                    print(f"Error getting resolved contact: {e}")
            
            # Also search the Global Address List for partial matches
            try:
                gal = namespace.AddressLists.Item("Global Address List")
                if gal:
                    query_lower = query.lower()
                    count = 0
                    for entry in gal.AddressEntries:
                        if count >= 10:  # Limit results
                            break
                        try:
                            name = entry.Name or ''
                            if query_lower in name.lower():
                                email = ''
                                if entry.Type == 'EX':
                                    try:
                                        exch_user = entry.GetExchangeUser()
                                        if exch_user:
                                            email = exch_user.PrimarySmtpAddress
                                    except:
                                        pass
                                else:
                                    email = entry.Address or ''
                                
                                if email and not any(c['email'].lower() == email.lower() for c in contacts):
                                    contacts.append({
                                        'email': email,
                                        'display_name': name,
                                        'type': 'gal'
                                    })
                                    count += 1
                        except:
                            continue
            except Exception as e:
                print(f"GAL search error: {e}")
            
            # Also search personal contacts folder
            try:
                contacts_folder = namespace.GetDefaultFolder(10)  # olFolderContacts
                query_lower = query.lower()
                count = 0
                for item in contacts_folder.Items:
                    if count >= 5:
                        break
                    try:
                        full_name = getattr(item, 'FullName', '') or ''
                        email = getattr(item, 'Email1Address', '') or ''
                        
                        if query_lower in full_name.lower() or query_lower in email.lower():
                            if email and not any(c['email'].lower() == email.lower() for c in contacts):
                                contacts.append({
                                    'email': email,
                                    'display_name': full_name or email,
                                    'type': 'contact'
                                })
                                count += 1
                    except:
                        continue
            except Exception as e:
                print(f"Contacts folder search error: {e}")
                
        finally:
            pythoncom.CoUninitialize()
            
    except Exception as e:
        print(f"Outlook contact search error: {e}")
        return jsonify({'error': str(e)}), 500
    
    return jsonify(contacts[:15])  # Limit to 15 results

# =============================================================================
# API ROUTES - ITEMS
# =============================================================================

@app.route('/api/items', methods=['GET'])
@login_required
def api_get_items():
    """Get items with optional filtering."""
    bucket = request.args.get('bucket')
    item_type = request.args.get('type')
    status = request.args.get('status')
    assigned_to = request.args.get('assigned_to')
    
    conn = get_db()
    cursor = conn.cursor()
    
    query = '''
        SELECT i.*, 
               u.display_name as assigned_to_name,
               CASE 
                   WHEN EXISTS (SELECT 1 FROM item_reviewers WHERE item_id = i.id) THEN (
                       SELECT GROUP_CONCAT(reviewer_name, ', ') 
                       FROM item_reviewers 
                       WHERE item_id = i.id
                   )
                   ELSE ir.display_name 
               END as initial_reviewer_name,
               qcr.display_name as qcr_name
        FROM item i
        LEFT JOIN user u ON i.assigned_to_user_id = u.id
        LEFT JOIN user ir ON i.initial_reviewer_id = ir.id
        LEFT JOIN user qcr ON i.qcr_id = qcr.id
        WHERE 1=1
    '''
    params = []
    
    if bucket and bucket != 'ALL':
        query += ' AND i.bucket = ?'
        params.append(bucket)
    
    if item_type:
        query += ' AND i.type = ?'
        params.append(item_type)
    
    if status:
        query += ' AND i.status = ?'
        params.append(status)
    
    if assigned_to:
        query += ' AND i.assigned_to_user_id = ?'
        params.append(assigned_to)
    
    # Handle show_closed filter
    show_closed = request.args.get('show_closed', 'true').lower()
    if show_closed != 'true':
        query += " AND i.closed_at IS NULL"
    
    query += ' ORDER BY i.date_received DESC, i.created_at DESC'
    
    cursor.execute(query, params)
    items = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    return jsonify(items)

@app.route('/api/item/<int:item_id>', methods=['GET'])
@login_required
def api_get_item(item_id):
    """Get a single item with all reviewer info."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT i.*, 
               u.display_name as assigned_to_name,
               CASE 
                   WHEN EXISTS (SELECT 1 FROM item_reviewers WHERE item_id = i.id) THEN (
                       SELECT GROUP_CONCAT(reviewer_name, ', ') 
                       FROM item_reviewers 
                       WHERE item_id = i.id
                   )
                   ELSE ir.display_name 
               END as initial_reviewer_name,
               qcr.display_name as qcr_name
        FROM item i
        LEFT JOIN user u ON i.assigned_to_user_id = u.id
        LEFT JOIN user ir ON i.initial_reviewer_id = ir.id
        LEFT JOIN user qcr ON i.qcr_id = qcr.id
        WHERE i.id = ?
    ''', (item_id,))
    item = cursor.fetchone()
    conn.close()
    
    if not item:
        return jsonify({'error': 'Item not found'}), 404
    
    item_dict = dict(item)
    
    # Add due date status colors
    item_dict['initial_reviewer_due_status'] = get_due_date_status(item_dict.get('initial_reviewer_due_date'))
    item_dict['qcr_due_status'] = get_due_date_status(item_dict.get('qcr_due_date'))
    item_dict['contractor_due_status'] = get_due_date_status(item_dict.get('due_date'))
    
    return jsonify(item_dict)

@app.route('/api/item/<int:item_id>', methods=['POST', 'PUT'])
@login_required
def api_update_item(item_id):
    """Update an item with validation for reviewers."""
    data = request.get_json()
    
    # Validate that initial_reviewer_id and qcr_id are not the same
    initial_reviewer_id = data.get('initial_reviewer_id')
    qcr_id = data.get('qcr_id')
    
    if initial_reviewer_id and qcr_id and str(initial_reviewer_id) == str(qcr_id):
        return jsonify({'error': 'Initial Reviewer and QCR must be different users'}), 400
    
    # Check if user is manually setting due dates
    manual_initial_due = 'initial_reviewer_due_date' in data
    manual_qcr_due = 'qcr_due_date' in data
    
    # Fields that can be updated
    allowed_fields = [
        'title', 'due_date', 'priority', 'status', 'assigned_to_user_id', 'notes', 'folder_link',
        'initial_reviewer_id', 'qcr_id', 'date_received',
        'initial_reviewer_due_date', 'qcr_due_date', 'rfi_question'
    ]
    updates = []
    params = []
    
    for field in allowed_fields:
        if field in data:
            updates.append(f'{field} = ?')
            value = data[field] if data[field] != '' else None
            params.append(value)
    
    if not updates:
        return jsonify({'error': 'No fields to update'}), 400
    
    params.append(item_id)
    
    conn = get_db()
    cursor = conn.cursor()
    
    # Get current item data to check for recalculation needs
    cursor.execute('SELECT date_received, due_date, priority FROM item WHERE id = ?', (item_id,))
    current = cursor.fetchone()
    
    cursor.execute(f'''
        UPDATE item SET {', '.join(updates)} WHERE id = ?
    ''', params)
    
    # Only recalculate due dates if priority, due_date, or date_received changed
    # AND the user didn't manually provide the due dates
    recalc_fields = ['priority', 'due_date', 'date_received']
    needs_recalc = any(field in data for field in recalc_fields) and not manual_initial_due and not manual_qcr_due
    
    if needs_recalc:
        # Get updated values
        cursor.execute('SELECT date_received, due_date, priority FROM item WHERE id = ?', (item_id,))
        updated = cursor.fetchone()
        
        if updated['date_received'] and updated['due_date']:
            due_dates = calculate_review_due_dates(
                updated['date_received'],
                updated['due_date'],
                updated['priority']
            )
            
            cursor.execute('''
                UPDATE item SET 
                    initial_reviewer_due_date = ?,
                    qcr_due_date = ?,
                    is_contractor_window_insufficient = ?
                WHERE id = ?
            ''', (
                due_dates['initial_reviewer_due_date'],
                due_dates['qcr_due_date'],
                1 if due_dates['is_contractor_window_insufficient'] else 0,
                item_id
            ))
    
    conn.commit()
    
    # Return updated item with all joins
    cursor.execute('''
        SELECT i.*, 
               u.display_name as assigned_to_name,
               ir.display_name as initial_reviewer_name,
               qcr.display_name as qcr_name
        FROM item i
        LEFT JOIN user u ON i.assigned_to_user_id = u.id
        LEFT JOIN user ir ON i.initial_reviewer_id = ir.id
        LEFT JOIN user qcr ON i.qcr_id = qcr.id
        WHERE i.id = ?
    ''', (item_id,))
    item = cursor.fetchone()
    conn.close()
    
    item_dict = dict(item)
    item_dict['initial_reviewer_due_status'] = get_due_date_status(item_dict.get('initial_reviewer_due_date'))
    item_dict['qcr_due_status'] = get_due_date_status(item_dict.get('qcr_due_date'))
    item_dict['contractor_due_status'] = get_due_date_status(item_dict.get('due_date'))
    
    return jsonify(item_dict)

@app.route('/api/items', methods=['POST'])
@login_required
def api_create_item():
    """Manually create a new item."""
    data = request.get_json()
    
    item_type = data.get('type')
    identifier = data.get('identifier', '').strip()
    bucket = data.get('bucket', 'ALL')
    
    if not item_type or item_type not in ['RFI', 'Submittal']:
        return jsonify({'error': 'Valid type (RFI or Submittal) required'}), 400
    
    if not identifier:
        return jsonify({'error': 'Identifier required'}), 400
    
    # Create folder
    title = data.get('title')
    folder_link = create_item_folder(item_type, identifier, bucket, title)
    
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute('''
            INSERT INTO item (type, bucket, identifier, title, due_date, priority, status, folder_link)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            item_type, bucket, identifier,
            data.get('title'), data.get('due_date'), data.get('priority'),
            data.get('status', 'Unassigned'), folder_link
        ))
        conn.commit()
        item_id = cursor.lastrowid
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({'error': 'Item with this identifier and bucket already exists'}), 400
    
    cursor.execute('SELECT * FROM item WHERE id = ?', (item_id,))
    item = dict(cursor.fetchone())
    conn.close()
    
    return jsonify(item), 201


@app.route('/api/items/<int:item_id>', methods=['DELETE'])
@login_required
def api_delete_item(item_id):
    """Delete an item and optionally its folder."""
    conn = get_db()
    cursor = conn.cursor()
    
    # Get the item first to check if it exists and get folder path
    cursor.execute('SELECT * FROM item WHERE id = ?', (item_id,))
    item = cursor.fetchone()
    
    if not item:
        conn.close()
        return jsonify({'error': 'Item not found'}), 404
    
    folder_path = item['folder_link']
    
    # Delete related records first (history, notifications, etc.)
    cursor.execute('DELETE FROM reviewer_response_history WHERE item_id = ?', (item_id,))
    cursor.execute('DELETE FROM notification WHERE item_id = ?', (item_id,))
    
    # Delete the item
    cursor.execute('DELETE FROM item WHERE id = ?', (item_id,))
    conn.commit()
    conn.close()
    
    # Optionally delete the folder (check query param)
    delete_folder = request.args.get('delete_folder', 'false').lower() == 'true'
    folder_deleted = False
    
    if delete_folder and folder_path:
        try:
            import shutil
            folder = Path(folder_path)
            if folder.exists():
                shutil.rmtree(folder)
                folder_deleted = True
        except Exception as e:
            print(f"Warning: Could not delete folder {folder_path}: {e}")
    
    return jsonify({
        'success': True,
        'message': f'Item {item_id} deleted',
        'folder_deleted': folder_deleted
    })

# =============================================================================
# API ROUTES - INBOX
# =============================================================================

@app.route('/api/inbox', methods=['GET'])
@login_required
def api_get_inbox():
    """Get items where user is initial reviewer, QCR, or assigned, sorted by relevant due date."""
    user_id = session['user_id']
    show_closed = request.args.get('show_closed', 'false').lower() == 'true'
    
    conn = get_db()
    cursor = conn.cursor()
    
    # Get items where user is initial reviewer, QCR, or assigned to
    query = '''
        SELECT i.*, 
               u.display_name as assigned_to_name,
               ir.display_name as initial_reviewer_name,
               qcr.display_name as qcr_name,
               CASE 
                   WHEN i.initial_reviewer_id = ? THEN 'initial_reviewer'
                   WHEN i.qcr_id = ? THEN 'qcr'
                   ELSE 'assigned'
               END as user_role
        FROM item i
        LEFT JOIN user u ON i.assigned_to_user_id = u.id
        LEFT JOIN user ir ON i.initial_reviewer_id = ir.id
        LEFT JOIN user qcr ON i.qcr_id = qcr.id
        WHERE (i.initial_reviewer_id = ? OR i.qcr_id = ? OR i.assigned_to_user_id = ?)
    '''
    params = [user_id, user_id, user_id, user_id, user_id]
    
    if not show_closed:
        query += " AND i.closed_at IS NULL"
    
    # Sort by role-specific due date, then priority, then date_received
    query += '''
        ORDER BY 
            CASE 
                WHEN i.initial_reviewer_id = ? THEN COALESCE(i.initial_reviewer_due_date, i.due_date)
                WHEN i.qcr_id = ? THEN COALESCE(i.qcr_due_date, i.due_date)
                ELSE i.due_date
            END ASC NULLS LAST,
            CASE i.priority 
                WHEN 'High' THEN 1 
                WHEN 'Medium' THEN 2 
                WHEN 'Low' THEN 3 
                ELSE 4 
            END,
            i.date_received ASC
    '''
    params.extend([user_id, user_id])
    
    cursor.execute(query, params)
    items = [dict(row) for row in cursor.fetchall()]
    
    # Check read status and add due date status for each item
    user_id_str = str(user_id)
    for item in items:
        read_by = item.get('read_by') or ''
        item['is_unread'] = user_id_str not in read_by.split(',')
        
        # Add role-specific due date status
        role = item.get('user_role')
        if role == 'initial_reviewer':
            item['my_due_date'] = item.get('initial_reviewer_due_date')
            item['my_due_status'] = get_due_date_status(item.get('initial_reviewer_due_date'))
        elif role == 'qcr':
            item['my_due_date'] = item.get('qcr_due_date')
            item['my_due_status'] = get_due_date_status(item.get('qcr_due_date'))
        else:
            item['my_due_date'] = item.get('due_date')
            item['my_due_status'] = get_due_date_status(item.get('due_date'))
    
    conn.close()
    return jsonify(items)

@app.route('/api/item/<int:item_id>/mark-read', methods=['POST'])
@login_required
def api_mark_read(item_id):
    """Mark an item as read by the current user."""
    user_id = str(session['user_id'])
    
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('SELECT read_by FROM item WHERE id = ?', (item_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Item not found'}), 404
    
    read_by = row['read_by'] or ''
    read_list = [x for x in read_by.split(',') if x]
    
    if user_id not in read_list:
        read_list.append(user_id)
        cursor.execute('UPDATE item SET read_by = ? WHERE id = ?', (','.join(read_list), item_id))
        conn.commit()
    
    conn.close()
    return jsonify({'success': True})

# =============================================================================
# API ROUTES - RESPONSE & CLOSEOUT
# =============================================================================

@app.route('/api/item/<int:item_id>/response', methods=['POST'])
@login_required
def api_save_response(item_id):
    """Save response details for an item."""
    data = request.get_json()
    user_id = session['user_id']
    user_role = session.get('role')
    
    conn = get_db()
    cursor = conn.cursor()
    
    # Check if user can edit (admin or assigned to item)
    cursor.execute('SELECT assigned_to_user_id, status FROM item WHERE id = ?', (item_id,))
    item = cursor.fetchone()
    
    if not item:
        conn.close()
        return jsonify({'error': 'Item not found'}), 404
    
    if item['status'] == 'Closed':
        conn.close()
        return jsonify({'error': 'Cannot modify closed item'}), 400
    
    if user_role != 'admin' and item['assigned_to_user_id'] != user_id:
        conn.close()
        return jsonify({'error': 'Not authorized to edit this item'}), 403
    
    # Update response fields (both regular and final response fields)
    updates = []
    params = []
    
    if 'response_category' in data:
        updates.append('response_category = ?')
        params.append(data['response_category'] or None)
        # Also update final response category
        updates.append('final_response_category = ?')
        params.append(data['response_category'] or None)
    
    if 'response_text' in data:
        updates.append('response_text = ?')
        params.append(data['response_text'] or None)
        # Also update final response text
        updates.append('final_response_text = ?')
        params.append(data['response_text'] or None)
    
    if 'response_files' in data:
        import json as json_module
        files_json = json_module.dumps(data['response_files']) if data['response_files'] else None
        updates.append('response_files = ?')
        params.append(files_json)
        # Also update final response files
        updates.append('final_response_files = ?')
        params.append(files_json)
    
    if updates:
        params.append(item_id)
        cursor.execute(f"UPDATE item SET {', '.join(updates)} WHERE id = ?", params)
        conn.commit()
    
    conn.close()
    return jsonify({'success': True})

@app.route('/api/item/<int:item_id>/close', methods=['POST'])
@login_required
def api_close_item(item_id):
    """Close out an item."""
    user_id = session['user_id']
    user_role = session.get('role')
    
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('SELECT * FROM item WHERE id = ?', (item_id,))
    item = cursor.fetchone()
    
    if not item:
        conn.close()
        return jsonify({'error': 'Item not found'}), 404
    
    if item['status'] == 'Closed':
        conn.close()
        return jsonify({'error': 'Item is already closed'}), 400
    
    if user_role != 'admin' and item['assigned_to_user_id'] != user_id:
        conn.close()
        return jsonify({'error': 'Not authorized to close this item'}), 403
    
    # Validate closeout requirements
    if not item['response_category']:
        conn.close()
        return jsonify({'error': 'Response category is required before closing'}), 400
    
    if not item['response_text']:
        conn.close()
        return jsonify({'error': 'Response text is required before closing'}), 400
    
    # Close the item
    cursor.execute('''
        UPDATE item SET status = 'Closed', closed_at = ? WHERE id = ?
    ''', (datetime.now().isoformat(), item_id))
    conn.commit()
    
    cursor.execute('SELECT * FROM item WHERE id = ?', (item_id,))
    updated_item = dict(cursor.fetchone())
    conn.close()
    
    # Update RFI Bulletin Tracker Excel if this is an RFI
    excel_result = update_rfi_tracker_excel(updated_item, action='close')
    if excel_result.get('success'):
        updated_item['excel_update'] = excel_result.get('message', 'Excel updated')
    elif excel_result.get('error'):
        updated_item['excel_update_error'] = excel_result.get('error')
    
    return jsonify(updated_item)

@app.route('/api/item/<int:item_id>/reopen', methods=['POST'])
@admin_required
def api_reopen_item(item_id):
    """Reopen a closed item (admin only)."""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('SELECT status FROM item WHERE id = ?', (item_id,))
    item = cursor.fetchone()
    
    if not item:
        conn.close()
        return jsonify({'error': 'Item not found'}), 404
    
    if item['status'] != 'Closed':
        conn.close()
        return jsonify({'error': 'Item is not closed'}), 400
    
    # Reopen - set to Ready for Response
    cursor.execute('''
        UPDATE item SET status = 'Ready for Response', closed_at = NULL WHERE id = ?
    ''', (item_id,))
    conn.commit()
    
    cursor.execute('SELECT * FROM item WHERE id = ?', (item_id,))
    updated_item = dict(cursor.fetchone())
    conn.close()
    
    # Update RFI Bulletin Tracker Excel if this is an RFI
    excel_result = update_rfi_tracker_excel(updated_item, action='reopen')
    if excel_result.get('success'):
        updated_item['excel_update'] = excel_result.get('message', 'Excel updated')
    elif excel_result.get('error'):
        updated_item['excel_update_error'] = excel_result.get('error')
    
    return jsonify(updated_item)

@app.route('/api/item/<int:item_id>/files', methods=['GET'])
@login_required
def api_get_item_files(item_id):
    """List files in the item's folder."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT folder_link FROM item WHERE id = ?', (item_id,))
    item = cursor.fetchone()
    conn.close()
    
    if not item:
        return jsonify({'error': 'Item not found'}), 404
    
    folder_path = item['folder_link']
    if not folder_path:
        return jsonify([])
    
    folder = Path(folder_path)
    if not folder.exists() or not folder.is_dir():
        return jsonify([])
    
    # Allowed file extensions
    allowed_extensions = {'.pdf', '.doc', '.docx', '.xls', '.xlsx', '.dwg', '.dxf', '.png', '.jpg', '.jpeg', '.tif', '.tiff', '.zip'}
    
    files = []
    try:
        for f in folder.iterdir():
            if f.is_file() and f.suffix.lower() in allowed_extensions:
                stat = f.stat()
                files.append({
                    'filename': f.name,
                    'modified': datetime.fromtimestamp(stat.st_mtime).isoformat(),
                    'size': stat.st_size
                })
        # Sort by modified date descending
        files.sort(key=lambda x: x['modified'], reverse=True)
    except Exception as e:
        print(f"Error listing files: {e}")
    
    return jsonify({'files': files, 'folder_path': folder_path})


@app.route('/api/open-folder', methods=['POST'])
@login_required
def api_open_folder():
    """Open a folder in Windows Explorer."""
    data = request.get_json()
    folder_path = data.get('path', '')
    
    if not folder_path:
        return jsonify({'error': 'No folder path provided'}), 400
    
    folder = Path(folder_path)
    
    # Create folder if it doesn't exist
    if not folder.exists():
        try:
            folder.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            return jsonify({'error': f'Could not create folder: {e}'}), 500
    
    # Open folder in Windows Explorer
    try:
        import subprocess
        subprocess.Popen(['explorer', str(folder)])
        return jsonify({'success': True, 'path': str(folder)})
    except Exception as e:
        return jsonify({'error': f'Could not open folder: {e}'}), 500


@app.route('/api/item/<int:item_id>/open-email', methods=['POST'])
@login_required
def api_open_original_email(item_id):
    """Open the original email in Outlook for an item."""
    if not HAS_WIN32COM:
        return jsonify({'error': 'Outlook integration not available'}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    
    # First try to get email_entry_id from item table
    cursor.execute('SELECT email_entry_id FROM item WHERE id = ?', (item_id,))
    item = cursor.fetchone()
    
    entry_id = None
    if item and item['email_entry_id']:
        entry_id = item['email_entry_id']
    else:
        # Fallback: get from email_log table
        cursor.execute('''
            SELECT entry_id FROM email_log 
            WHERE item_id = ? 
            ORDER BY received_at ASC 
            LIMIT 1
        ''', (item_id,))
        email_log = cursor.fetchone()
        if email_log:
            entry_id = email_log['entry_id']
    
    conn.close()
    
    if not entry_id:
        return jsonify({'error': 'No original email found for this item'}), 404
    
    try:
        pythoncom.CoInitialize()
        try:
            outlook = win32com.client.Dispatch("Outlook.Application")
            namespace = outlook.GetNamespace("MAPI")
            
            # Get the email by EntryID
            mail_item = namespace.GetItemFromID(entry_id)
            
            # Get the inspector (window) and activate it to bring to front
            inspector = mail_item.GetInspector
            inspector.Activate()
            mail_item.Display(True)  # True = modal window
            
            return jsonify({'success': True, 'message': 'Email opened in Outlook'})
        finally:
            pythoncom.CoUninitialize()
    except Exception as e:
        print(f"Error opening email: {e}")
        return jsonify({'error': f'Could not open email: {str(e)}'}), 500

# =============================================================================
# API ROUTES - COMMENTS
# =============================================================================

@app.route('/api/comments/<int:item_id>', methods=['GET'])
@login_required
def api_get_comments(item_id):
    """Get comments for an item."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT c.*, u.display_name as author_name, u.email as author_email
        FROM comment c
        JOIN user u ON c.user_id = u.id
        WHERE c.item_id = ?
        ORDER BY c.created_at ASC
    ''', (item_id,))
    comments = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify(comments)

@app.route('/api/comments/<int:item_id>', methods=['POST'])
@login_required
def api_add_comment(item_id):
    """Add a comment to an item."""
    data = request.get_json()
    body = data.get('body', '').strip()
    
    if not body:
        return jsonify({'error': 'Comment body required'}), 400
    
    user_id = session['user_id']
    
    conn = get_db()
    cursor = conn.cursor()
    
    # Verify item exists
    cursor.execute('SELECT id FROM item WHERE id = ?', (item_id,))
    if not cursor.fetchone():
        conn.close()
        return jsonify({'error': 'Item not found'}), 404
    
    cursor.execute('''
        INSERT INTO comment (item_id, user_id, body)
        VALUES (?, ?, ?)
    ''', (item_id, user_id, body))
    conn.commit()
    comment_id = cursor.lastrowid
    
    cursor.execute('''
        SELECT c.*, u.display_name as author_name, u.email as author_email
        FROM comment c
        JOIN user u ON c.user_id = u.id
        WHERE c.id = ?
    ''', (comment_id,))
    comment = dict(cursor.fetchone())
    conn.close()
    
    return jsonify(comment), 201

# =============================================================================
# API ROUTES - STATS
# =============================================================================

@app.route('/api/stats', methods=['GET'])
@login_required
def api_get_stats():
    """Get dashboard statistics."""
    user_id = session.get('user_id')
    conn = get_db()
    cursor = conn.cursor()
    
    # Total counts
    cursor.execute('SELECT COUNT(*) FROM item')
    total_items = cursor.fetchone()[0]
    
    # Counts by type
    cursor.execute('SELECT type, COUNT(*) FROM item GROUP BY type')
    by_type = {row[0]: row[1] for row in cursor.fetchall()}
    
    # Counts by bucket
    cursor.execute('SELECT bucket, COUNT(*) FROM item GROUP BY bucket')
    by_bucket = {row[0]: row[1] for row in cursor.fetchall()}
    
    # Counts by status
    cursor.execute('SELECT status, COUNT(*) FROM item GROUP BY status')
    by_status = {row[0]: row[1] for row in cursor.fetchall()}
    
    # Open items (not closed)
    cursor.execute("SELECT COUNT(*) FROM item WHERE closed_at IS NULL")
    open_items = cursor.fetchone()[0]
    
    # Overdue items
    today = datetime.now().strftime('%Y-%m-%d')
    cursor.execute('''
        SELECT COUNT(*) FROM item 
        WHERE due_date < ? AND closed_at IS NULL
    ''', (today,))
    overdue_items = cursor.fetchone()[0]
    
    # Items due this week
    week_end = (datetime.now() + timedelta(days=7)).strftime('%Y-%m-%d')
    cursor.execute('''
        SELECT COUNT(*) FROM item 
        WHERE due_date BETWEEN ? AND ? AND closed_at IS NULL
    ''', (today, week_end))
    due_this_week = cursor.fetchone()[0]
    
    # Unassigned items
    cursor.execute("SELECT COUNT(*) FROM item WHERE assigned_to_user_id IS NULL AND closed_at IS NULL")
    unassigned = cursor.fetchone()[0]
    
    # Inbox count for current user
    inbox_count = 0
    if user_id:
        cursor.execute("SELECT COUNT(*) FROM item WHERE assigned_to_user_id = ? AND closed_at IS NULL", (user_id,))
        inbox_count = cursor.fetchone()[0]
    
    conn.close()
    
    return jsonify({
        'total_items': total_items,
        'by_type': by_type,
        'by_bucket': by_bucket,
        'by_status': by_status,
        'open_items': open_items,
        'overdue_items': overdue_items,
        'due_this_week': due_this_week,
        'unassigned': unassigned,
        'inbox_count': inbox_count
    })

# =============================================================================
# API ROUTES - SYSTEM
# =============================================================================

@app.route('/api/poll-status', methods=['GET'])
@login_required
def api_poll_status():
    """Get email polling status."""
    return jsonify(email_poller.get_status())

@app.route('/api/poll-now', methods=['POST'])
@admin_required
def api_poll_now():
    """Trigger immediate email poll."""
    if HAS_WIN32COM:
        try:
            threading.Thread(target=email_poller._poll_emails, daemon=True).start()
            return jsonify({'success': True, 'message': 'Poll triggered'})
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    return jsonify({'error': 'Outlook integration not available'}), 400

@app.route('/api/config', methods=['GET'])
@admin_required
def api_get_config():
    """Get current configuration."""
    return jsonify(CONFIG)

@app.route('/api/config', methods=['POST'])
@admin_required
def api_update_config():
    """Update configuration."""
    global CONFIG
    data = request.get_json()
    
    for key in ['base_folder_path', 'outlook_folder', 'poll_interval_minutes', 'project_name']:
        if key in data:
            CONFIG[key] = data[key]
    
    with open(CONFIG_PATH, 'w') as f:
        json.dump(CONFIG, f, indent=2)
    
    return jsonify(CONFIG)

# =============================================================================
# API ROUTES - AIRTABLE SYNC
# =============================================================================

@app.route('/api/airtable/sync', methods=['POST'])
@admin_required
def api_airtable_sync():
    """Manually trigger sync of Airtable responses."""
    if not HAS_AIRTABLE:
        return jsonify({'error': 'Airtable integration not available'}), 400
    
    try:
        result = sync_airtable_responses()
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/airtable/status', methods=['GET'])
@login_required
def api_airtable_status():
    """Get Airtable configuration status."""
    if not HAS_AIRTABLE:
        return jsonify({'configured': False, 'available': False})
    
    config = load_airtable_config()
    is_configured = bool(config.get('api_key') and config.get('base_id'))
    
    return jsonify({
        'available': True,
        'configured': is_configured,
        'has_reviewer_form': bool(config.get('reviewer_form_id')),
        'has_qcr_form': bool(config.get('qcr_form_id'))
    })

# =============================================================================
# MAGIC-LINK RESPONSE ROUTES
# =============================================================================

@app.route('/respond/reviewer', methods=['GET'])
def respond_reviewer_form():
    """Show reviewer response form via magic link."""
    token = request.args.get('token')
    if not token:
        return render_template_string(ERROR_PAGE_TEMPLATE, error='Missing token'), 400
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT i.*, 
               ir.display_name as reviewer_name, ir.email as reviewer_email,
               qcr.display_name as qcr_name
        FROM item i
        LEFT JOIN user ir ON i.initial_reviewer_id = ir.id
        LEFT JOIN user qcr ON i.qcr_id = qcr.id
        WHERE i.email_token_reviewer = ?
    ''', (token,))
    item = cursor.fetchone()
    
    if not item:
        conn.close()
        return render_template_string(ERROR_PAGE_TEMPLATE, error='Invalid or expired token'), 404
    
    item_dict = dict(item)
    
    # Check if item is closed
    is_closed = item['status'] == 'Closed'
    
    # Determine if submission is allowed
    # Allowed when: NOT closed AND (QCR hasn't responded yet OR QCR sent it back)
    can_submit = False
    is_resubmit = False
    qcr_feedback = None
    
    if not is_closed:
        qcr_action = item['qcr_action']
        qcr_response_at = item['qcr_response_at']
        
        # Case 1: QCR hasn't responded yet - always allow
        if not qcr_response_at or item['qcr_response_status'] in ['Not Sent', 'Email Sent', 'Waiting for Revision']:
            can_submit = True
            # It's a resubmit if reviewer already responded before
            if item['reviewer_response_at']:
                is_resubmit = True
        
        # Case 2: QCR sent it back - allow resubmit
        if qcr_action == 'Send Back':
            can_submit = True
            is_resubmit = True
            qcr_feedback = item['qcr_notes']
        
        # Case 3: QCR approved/modified - don't allow (finalized)
        if qcr_action in ['Approve', 'Modify'] and item['status'] == 'Ready for Response':
            can_submit = False
    
    # Get current version info
    # First submission is v0, revisions are v1, v2, etc.
    current_version = item['reviewer_response_version'] or 0
    if is_resubmit:
        next_version = current_version + 1
    else:
        next_version = 0  # First submission is always v0
    
    # Get previous response for pre-fill
    previous_response = None
    previous_files = []
    if is_resubmit and item['reviewer_response_at']:
        previous_response = {
            'category': item['reviewer_response_category'],
            'text': item['reviewer_notes'] or item['reviewer_response_text'],
            'files': item['reviewer_selected_files']
        }
        if item['reviewer_selected_files']:
            try:
                previous_files = json.loads(item['reviewer_selected_files'])
            except:
                pass
    
    # Get version history
    version_history = ''
    cursor.execute('''
        SELECT version, submitted_at 
        FROM reviewer_response_history 
        WHERE item_id = ? 
        ORDER BY version DESC
    ''', (item['id'],))
    history = cursor.fetchall()
    if history:
        version_parts = [f"v{h['version']} ({h['submitted_at'][:16].replace('T', ' ')})" for h in history]
        version_history = ', '.join(version_parts)
    
    conn.close()
    
    # Get files in folder
    files = []
    if item['folder_link']:
        try:
            folder_path = Path(item['folder_link'])
            if folder_path.exists():
                files = [f.name for f in folder_path.iterdir() if f.is_file()]
        except:
            pass
    
    return render_template_string(REVIEWER_RESPONSE_TEMPLATE, 
        item=item_dict,
        files=files,
        token=token,
        version=next_version,
        is_closed=is_closed,
        can_submit=can_submit,
        is_resubmit=is_resubmit,
        qcr_feedback=qcr_feedback,
        previous_response=previous_response,
        previous_files=previous_files,
        version_history=version_history
    )

@app.route('/respond/reviewer', methods=['POST'])
def respond_reviewer_submit():
    """Handle reviewer response submission with version tracking."""
    token = request.form.get('token')
    if not token:
        return render_template_string(ERROR_PAGE_TEMPLATE, error='Missing token'), 400
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT i.*, ir.id as reviewer_user_id
        FROM item i
        LEFT JOIN user ir ON i.initial_reviewer_id = ir.id
        WHERE i.email_token_reviewer = ?
    ''', (token,))
    item = cursor.fetchone()
    
    if not item:
        conn.close()
        return render_template_string(ERROR_PAGE_TEMPLATE, error='Invalid or expired token'), 404
    
    item_id = item['id']
    
    # Check if item is closed - block all submissions
    if item['status'] == 'Closed':
        conn.close()
        return render_template_string(ERROR_PAGE_TEMPLATE, 
            error='This item has been closed. No further changes can be submitted. Contact the project administrator if this is unexpected.'), 403
    
    # Check if submission is allowed
    qcr_action = item['qcr_action']
    qcr_response_at = item['qcr_response_at']
    
    # Determine if this is a valid submission scenario
    can_submit = False
    is_resubmit = False
    
    # Case 1: QCR hasn't responded yet - always allow
    if not qcr_response_at or item['qcr_response_status'] in ['Not Sent', 'Email Sent', 'Waiting for Revision']:
        can_submit = True
        if item['reviewer_response_at']:
            is_resubmit = True
    
    # Case 2: QCR sent it back - allow resubmit
    if qcr_action == 'Send Back':
        can_submit = True
        is_resubmit = True
    
    # Case 3: QCR approved/modified - don't allow (finalized)
    if qcr_action in ['Approve', 'Modify'] and item['status'] == 'Ready for Response':
        can_submit = False
    
    if not can_submit:
        conn.close()
        return render_template_string(ERROR_PAGE_TEMPLATE, 
            error='This item has already been finalized in QC. Please contact the project admin if additional changes are required.'), 403
    
    # Get form data
    response_category = request.form.get('response_category')
    notes = request.form.get('notes', '')
    internal_notes = request.form.get('internal_notes', '')
    selected_files = request.form.getlist('selected_files')
    
    # Calculate new version
    # First submission is v0, revisions are v1, v2, etc.
    current_version = item['reviewer_response_version']
    new_version = (current_version + 1) if current_version is not None else 0
    
    # Track if this was a send-back scenario (before we reset qcr_action)
    was_sent_back = qcr_action == 'Send Back'
    
    # Store current response in history before updating (if this is a resubmit)
    if is_resubmit and item['reviewer_response_at']:
        cursor.execute('''
            INSERT INTO reviewer_response_history 
            (item_id, version, submitted_at, response_category, response_text, response_files, submitted_by_user_id)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (
            item_id,
            current_version,
            item['reviewer_response_at'],
            item['reviewer_response_category'],
            item['reviewer_notes'] or item['reviewer_response_text'],
            item['reviewer_selected_files'],
            item['reviewer_user_id']
        ))
    
    # Determine new status based on whether QCR sent it back
    if was_sent_back:
        # Reset QCR state for new review cycle
        cursor.execute('''
            UPDATE item SET
                reviewer_response_at = ?,
                reviewer_response_status = 'Responded',
                reviewer_response_category = ?,
                reviewer_notes = ?,
                reviewer_internal_notes = ?,
                reviewer_selected_files = ?,
                reviewer_response_version = ?,
                status = 'In QC',
                qcr_action = NULL,
                qcr_response_at = NULL,
                qcr_response_status = 'Not Sent',
                qcr_notes = NULL,
                qcr_internal_notes = NULL,
                qcr_response_category = NULL,
                qcr_response_text = NULL,
                qcr_response_mode = NULL
            WHERE id = ?
        ''', (
            datetime.now().isoformat(),
            response_category,
            notes,
            internal_notes,
            json.dumps(selected_files),
            new_version,
            item_id
        ))
    else:
        # Standard update
        cursor.execute('''
            UPDATE item SET
                reviewer_response_at = ?,
                reviewer_response_status = 'Responded',
                reviewer_response_category = ?,
                reviewer_notes = ?,
                reviewer_internal_notes = ?,
                reviewer_selected_files = ?,
                reviewer_response_version = ?,
                status = 'In QC'
            WHERE id = ?
        ''', (
            datetime.now().isoformat(),
            response_category,
            notes,
            internal_notes,
            json.dumps(selected_files),
            new_version,
            item_id
        ))
    
    conn.commit()
    conn.close()
    
    # Send appropriate notifications
    if is_resubmit:
        # Send version update notification to QCR
        if was_sent_back:
            # Full new QCR assignment email for revision after send-back
            send_qcr_assignment_email(item_id, is_revision=True, version=new_version)
        else:
            # Just an update notification (QCR hasn't responded yet)
            send_qcr_version_update_email(item_id, new_version)
    else:
        # First submission - send QCR assignment
        send_qcr_assignment_email(item_id)
    
    if is_resubmit:
        return render_template_string(SUCCESS_TEMPLATE, 
            message=f'Your revised response (v{new_version}) has been submitted!',
            details='The QC Reviewer has been notified of your updated response.'
        )
    else:
        return render_template_string(SUCCESS_TEMPLATE, 
            message='Your review has been submitted successfully!',
            details='The QC Reviewer has been notified and will complete the final review.'
        )

@app.route('/respond/qcr', methods=['GET'])
def respond_qcr_form():
    """Show QCR response form via magic link."""
    token = request.args.get('token')
    if not token:
        return render_template_string(ERROR_PAGE_TEMPLATE, error='Missing token'), 400
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT i.*, 
               ir.display_name as reviewer_name, ir.email as reviewer_email,
               qcr.display_name as qcr_name, qcr.email as qcr_email
        FROM item i
        LEFT JOIN user ir ON i.initial_reviewer_id = ir.id
        LEFT JOIN user qcr ON i.qcr_id = qcr.id
        WHERE i.email_token_qcr = ?
    ''', (token,))
    item = cursor.fetchone()
    
    if not item:
        conn.close()
        return render_template_string(ERROR_PAGE_TEMPLATE, error='Invalid or expired token'), 404
    
    # Check if item is closed
    if item['status'] == 'Closed':
        conn.close()
        return render_template_string(ERROR_PAGE_TEMPLATE, 
            error='This item has been closed. No further changes can be submitted. Contact the project administrator if this is unexpected.'), 403
    
    # Check if already responded
    if item['qcr_response_at']:
        conn.close()
        return render_template_string(ALREADY_RESPONDED_TEMPLATE, 
            item=dict(item),
            response_type='qcr'
        )
    
    # Get version history
    cursor.execute('''
        SELECT version, submitted_at, response_category
        FROM reviewer_response_history 
        WHERE item_id = ? 
        ORDER BY version DESC
        LIMIT 5
    ''', (item['id'],))
    version_history = cursor.fetchall()
    conn.close()
    
    # Get files in folder
    files = []
    if item['folder_link']:
        try:
            folder_path = Path(item['folder_link'])
            if folder_path.exists():
                files = [f.name for f in folder_path.iterdir() if f.is_file()]
        except:
            pass
    
    # Parse reviewer selected files
    reviewer_files = []
    if item['reviewer_selected_files']:
        try:
            reviewer_files = json.loads(item['reviewer_selected_files'])
        except:
            pass
    
    # Get version info
    current_version = item['reviewer_response_version'] if item['reviewer_response_version'] is not None else 0
    
    return render_template_string(QCR_RESPONSE_TEMPLATE, 
        item=dict(item),
        files=files,
        reviewer_files=reviewer_files,
        token=token,
        version=current_version,
        version_history=[dict(v) for v in version_history] if version_history else []
    )

@app.route('/respond/qcr', methods=['POST'])
def respond_qcr_submit():
    """Handle QCR response submission."""
    token = request.form.get('token')
    if not token:
        return render_template_string(ERROR_PAGE_TEMPLATE, error='Missing token'), 400
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM item WHERE email_token_qcr = ?', (token,))
    item = cursor.fetchone()
    
    if not item:
        conn.close()
        return render_template_string(ERROR_PAGE_TEMPLATE, error='Invalid or expired token'), 404
    
    # Check if item is closed - block all submissions
    if item['status'] == 'Closed':
        conn.close()
        return render_template_string(ERROR_PAGE_TEMPLATE, 
            error='This item has been closed. No further changes can be submitted. Contact the project administrator if this is unexpected.'), 403
    
    # Check if already responded
    if item['qcr_response_at']:
        conn.close()
        return render_template_string(ALREADY_RESPONDED_TEMPLATE, 
            item=dict(item),
            response_type='qcr'
        )
    
    # Get form data
    qc_action = request.form.get('qc_action')  # Approve, Modify, Send Back
    response_mode = request.form.get('response_mode', 'Keep')  # Keep, Tweak, Revise
    response_category = request.form.get('response_category')
    response_text = request.form.get('response_text', '')
    qcr_internal_notes = request.form.get('qcr_internal_notes', '')
    selected_files = request.form.getlist('selected_files')
    
    # Use response_text as qcr_notes (description)
    qcr_notes = response_text
    
    item_id = item['id']
    item_dict = dict(item)
    
    # Determine final values based on QC action
    if qc_action == 'Send Back':
        # Item goes back to reviewer - set status to indicate revision needed
        new_status = 'In Review'  # Back to review stage
        final_category = None
        final_text = None
        final_files = None
        qcr_response_category = None  # No final category for send back
        qcr_selected_files = None
    else:
        # Approve or Modify - finalize the response
        new_status = 'Ready for Response'
        
        # Determine final response text based on mode
        if response_mode == 'Keep':
            final_text = item['reviewer_response_text'] or item['reviewer_notes'] or ''
        elif response_mode == 'Tweak' or response_mode == 'Revise':
            final_text = response_text
        else:
            final_text = item['reviewer_response_text'] or item['reviewer_notes'] or ''
        
        final_category = response_category
        final_files = json.dumps(selected_files) if selected_files else item['reviewer_selected_files']
        qcr_response_category = response_category
        qcr_selected_files = json.dumps(selected_files) if selected_files else None
    
    # Update item with all QCR response data
    cursor.execute('''
        UPDATE item SET
            qcr_response_at = ?,
            qcr_response_status = 'Responded',
            qcr_action = ?,
            qcr_response_mode = ?,
            qcr_response_category = ?,
            qcr_response_text = ?,
            qcr_notes = ?,
            qcr_internal_notes = ?,
            qcr_selected_files = ?,
            final_response_category = ?,
            final_response_text = ?,
            final_response_files = ?,
            status = ?
        WHERE id = ?
    ''', (
        datetime.now().isoformat(),
        qc_action,
        response_mode if qc_action != 'Send Back' else None,
        qcr_response_category,
        response_text if qc_action != 'Send Back' else None,
        qcr_notes,
        qcr_internal_notes,
        qcr_selected_files,
        final_category,
        final_text,
        final_files,
        new_status,
        item_id
    ))
    
    # If sending back, also clear the reviewer's response so they can resubmit
    if qc_action == 'Send Back':
        cursor.execute('''
            UPDATE item SET
                reviewer_response_at = NULL,
                reviewer_response_status = 'Email Sent',
                qcr_response_status = 'Waiting for Revision'
            WHERE id = ?
        ''', (item_id,))
    
    conn.commit()
    
    # Get item details for notification
    cursor.execute('SELECT type, identifier, title FROM item WHERE id = ?', (item_id,))
    item_info = cursor.fetchone()
    conn.close()
    
    # Send notification email to reviewer
    send_reviewer_notification_email(
        item_id, 
        qc_action, 
        qcr_notes, 
        final_category=final_category,
        final_text=final_text
    )
    
    # Create system notifications based on QC action
    if qc_action == 'Approve' or qc_action == 'Modify':
        # Notification that response is ready to send to contractor
        create_notification(
            'response_ready',
            f'✅ Response Ready: {item_info["type"]} {item_info["identifier"]}',
            f'QC review complete. The response for "{item_info["title"] or item_info["identifier"]}" is ready to be sent to the contractor. Final category: {final_category}',
            item_id=item_id,
            action_url=f'/api/items/{item_id}/complete',
            action_label='Mark Complete'
        )
        
        # Send completion confirmation email to both QCR and reviewer with summary
        send_qcr_completion_confirmation_email(
            item_id, 
            qc_action, 
            qcr_notes, 
            final_category=final_category, 
            final_text=final_text
        )
    elif qc_action == 'Send Back':
        # Notification that item was sent back
        create_notification(
            'sent_back',
            f'↩️ Sent Back: {item_info["type"]} {item_info["identifier"]}',
            f'The item "{item_info["title"] or item_info["identifier"]}" has been sent back to the reviewer for revisions.',
            item_id=item_id
        )
    
    # Return appropriate success message
    if qc_action == 'Approve':
        return render_template_string(SUCCESS_TEMPLATE, 
            message='Response Approved!',
            details='The reviewer has been notified. The item is now ready for closeout.'
        )
    elif qc_action == 'Modify':
        return render_template_string(SUCCESS_TEMPLATE, 
            message='Response Modified and Finalized!',
            details='The reviewer has been notified of your modifications. The item is now ready for closeout.'
        )
    else:  # Send Back
        return render_template_string(SUCCESS_TEMPLATE, 
            message='Item Sent Back to Reviewer',
            details='The reviewer has been notified and will receive a link to revise their response.'
        )

# =============================================================================
# ADMIN WORKFLOW API
# =============================================================================

@app.route('/api/admin/workflow', methods=['GET'])
@admin_required
def api_admin_workflow():
    """Get workflow status for all items."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT i.id, i.type, i.identifier, i.title, i.status, i.priority,
               i.initial_reviewer_due_date, i.qcr_due_date, i.folder_link,
               i.reviewer_email_sent_at, i.reviewer_response_at, i.reviewer_response_status,
               i.reviewer_response_category, i.reviewer_notes, i.reviewer_selected_files,
               i.reviewer_response_text,
               i.qcr_email_sent_at, i.qcr_response_at, i.qcr_response_status,
               i.qcr_response_category, i.qcr_notes, i.qcr_selected_files,
               i.qcr_action, i.qcr_response_mode, i.qcr_response_text,
               i.final_response_category, i.final_response_text, i.final_response_files,
               ir.display_name as reviewer_name, ir.email as reviewer_email,
               qcr.display_name as qcr_name, qcr.email as qcr_email
        FROM item i
        LEFT JOIN user ir ON i.initial_reviewer_id = ir.id
        LEFT JOIN user qcr ON i.qcr_id = qcr.id
        WHERE i.initial_reviewer_id IS NOT NULL OR i.qcr_id IS NOT NULL
        ORDER BY i.created_at DESC
    ''')
    items = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify(items)

@app.route('/api/admin/send_reviewer_email/<int:item_id>', methods=['POST'])
@admin_required
def api_send_reviewer_email(item_id):
    """Send or resend reviewer assignment email."""
    try:
        result = send_reviewer_assignment_email(item_id)
        if result['success']:
            return jsonify(result)
        else:
            return jsonify(result), 400
    except Exception as e:
        return jsonify({'success': False, 'error': f'Server error: {str(e)}'}), 500

@app.route('/api/admin/send_qcr_email/<int:item_id>', methods=['POST'])
@admin_required
def api_send_qcr_email(item_id):
    """Send or resend QCR assignment email."""
    try:
        result = send_qcr_assignment_email(item_id)
        if result['success']:
            return jsonify(result)
        else:
            return jsonify(result), 400
    except Exception as e:
        return jsonify({'success': False, 'error': f'Server error: {str(e)}'}), 500

@app.route('/api/admin/send_multi_reviewer_qcr_email/<int:item_id>', methods=['POST'])
@admin_required
def api_send_multi_reviewer_qcr_email(item_id):
    """Send or resend multi-reviewer QCR assignment email."""
    try:
        result = send_multi_reviewer_qcr_email(item_id)
        if result['success']:
            return jsonify(result)
        else:
            return jsonify(result), 400
    except Exception as e:
        return jsonify({'success': False, 'error': f'Server error: {str(e)}'}), 500

# =============================================================================
# CONTRACTOR UPDATE REVIEW API ENDPOINTS
# =============================================================================

@app.route('/api/pending-updates', methods=['GET'])
@admin_required
def api_get_pending_updates():
    """Get all items with pending contractor updates that need admin review."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT i.id, i.type, i.identifier, i.title, i.status, i.due_date,
               i.has_pending_update, i.update_type, i.update_detected_at,
               i.previous_due_date, i.previous_title, i.previous_priority,
               i.update_email_body, i.reopened_from_closed, i.status_before_update,
               i.reviewer_response_status, i.qcr_response_status,
               ir.display_name as reviewer_name,
               qcr.display_name as qcr_name
        FROM item i
        LEFT JOIN user ir ON i.initial_reviewer_id = ir.id
        LEFT JOIN user qcr ON i.qcr_id = qcr.id
        WHERE i.has_pending_update = 1
        ORDER BY i.update_detected_at DESC
    ''')
    items = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify(items)

@app.route('/api/item/<int:item_id>/update-history', methods=['GET'])
@login_required
def api_get_item_update_history(item_id):
    """Get the update history for an item."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT uh.*, u.display_name as reviewed_by_name
        FROM item_update_history uh
        LEFT JOIN user u ON uh.admin_reviewed_by = u.id
        WHERE uh.item_id = ?
        ORDER BY uh.detected_at DESC
    ''', (item_id,))
    history = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify(history)

@app.route('/api/item/<int:item_id>/review-update', methods=['POST'])
@admin_required
def api_review_update(item_id):
    """Admin reviews a contractor update and decides on action.
    
    Actions:
    - 'accept_due_date': Just update the due dates, notify appropriate parties
    - 'restart_workflow': Content changed, restart to reviewer with note
    - 'dismiss': Dismiss the update without action
    """
    data = request.get_json()
    action = data.get('action')  # 'accept_due_date', 'restart_workflow', 'dismiss'
    admin_note = data.get('admin_note', '')
    apply_new_values = data.get('apply_new_values', True)  # Whether to apply the new due_date/title/priority
    
    if action not in ('accept_due_date', 'restart_workflow', 'dismiss'):
        return jsonify({'error': 'Invalid action'}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    
    # Get item details
    cursor.execute('''
        SELECT i.*, 
               ir.email as reviewer_email, ir.display_name as reviewer_name,
               qcr.email as qcr_email, qcr.display_name as qcr_name
        FROM item i
        LEFT JOIN user ir ON i.initial_reviewer_id = ir.id
        LEFT JOIN user qcr ON i.qcr_id = qcr.id
        WHERE i.id = ?
    ''', (item_id,))
    item = cursor.fetchone()
    
    if not item:
        conn.close()
        return jsonify({'error': 'Item not found'}), 404
    
    if not item['has_pending_update']:
        conn.close()
        return jsonify({'error': 'No pending update for this item'}), 400
    
    # Get the latest update history entry
    cursor.execute('''
        SELECT * FROM item_update_history
        WHERE item_id = ? AND admin_reviewed_at IS NULL
        ORDER BY detected_at DESC LIMIT 1
    ''', (item_id,))
    update_history = cursor.fetchone()
    
    now = datetime.now().isoformat()
    admin_user_id = session.get('user_id')
    result = {'success': True, 'action': action, 'emails_sent': []}
    
    if action == 'dismiss':
        # Just clear the update flag without changing anything
        cursor.execute('''
            UPDATE item SET
                has_pending_update = 0,
                update_type = NULL,
                update_reviewed_at = ?,
                update_admin_note = ?
            WHERE id = ?
        ''', (now, admin_note, item_id))
        
        if update_history:
            cursor.execute('''
                UPDATE item_update_history SET
                    admin_reviewed_at = ?,
                    admin_reviewed_by = ?,
                    admin_note = ?,
                    action_taken = 'dismissed'
                WHERE id = ?
            ''', (now, admin_user_id, admin_note, update_history['id']))
        
    elif action == 'accept_due_date':
        # Apply new due date and notify appropriate parties
        updates = ['has_pending_update = 0', 'update_type = NULL', 
                   'update_reviewed_at = ?', 'update_admin_note = ?']
        params = [now, admin_note]
        
        if apply_new_values and update_history:
            if update_history['new_due_date']:
                updates.append('due_date = ?')
                params.append(update_history['new_due_date'])
                
                # Recalculate review due dates
                date_received = item['date_received']
                priority = update_history['new_priority'] or item['priority']
                if date_received:
                    due_dates = calculate_review_due_dates(
                        date_received, update_history['new_due_date'], priority
                    )
                    updates.extend([
                        'initial_reviewer_due_date = ?',
                        'qcr_due_date = ?',
                        'is_contractor_window_insufficient = ?'
                    ])
                    params.extend([
                        due_dates['initial_reviewer_due_date'],
                        due_dates['qcr_due_date'],
                        1 if due_dates['is_contractor_window_insufficient'] else 0
                    ])
            
            if update_history['new_priority']:
                updates.append('priority = ?')
                params.append(update_history['new_priority'])
        
        params.append(item_id)
        cursor.execute(f'''
            UPDATE item SET {', '.join(updates)} WHERE id = ?
        ''', params)
        
        if update_history:
            cursor.execute('''
                UPDATE item_update_history SET
                    admin_reviewed_at = ?,
                    admin_reviewed_by = ?,
                    admin_note = ?,
                    action_taken = 'due_date_accepted'
                WHERE id = ?
            ''', (now, admin_user_id, admin_note, update_history['id']))
        
        # Send due date update notification based on workflow status
        reviewer_status = item['reviewer_response_status']
        qcr_status = item['qcr_response_status']
        
        new_due_date = update_history['new_due_date'] if update_history else item['due_date']
        was_reopened = item['reopened_from_closed']
        
        if qcr_status == 'Email Sent' and item['qcr_email']:
            # QCR has the ball - notify only QCR
            email_result = send_due_date_update_email(
                item_id, 'qcr', new_due_date, admin_note, was_reopened
            )
            result['emails_sent'].append({'to': 'qcr', 'result': email_result})
        elif reviewer_status == 'Email Sent' and item['reviewer_email']:
            # Reviewer has the ball - notify reviewer
            email_result = send_due_date_update_email(
                item_id, 'reviewer', new_due_date, admin_note, was_reopened
            )
            result['emails_sent'].append({'to': 'reviewer', 'result': email_result})
        
    elif action == 'restart_workflow':
        # Content changed - restart to reviewer
        # First, if item was closed, reopen it
        updates = [
            'has_pending_update = 0', 'update_type = NULL',
            'update_reviewed_at = ?', 'update_admin_note = ?',
            'status = ?',
            'reviewer_response_status = NULL', 'reviewer_response_at = NULL',
            'reviewer_response_category = NULL', 'reviewer_notes = NULL',
            'reviewer_selected_files = NULL', 'reviewer_response_text = NULL',
            'qcr_response_status = NULL', 'qcr_response_at = NULL',
            'qcr_action = NULL', 'qcr_response_mode = NULL',
            'qcr_notes = NULL', 'qcr_selected_files = NULL',
            'final_response_category = NULL', 'final_response_text = NULL',
            'final_response_files = NULL'
        ]
        params = [now, admin_note, 'Assigned']
        
        # Clear closed_at if it was closed
        if item['closed_at']:
            updates.append('closed_at = NULL')
        
        # Apply new values if requested
        if apply_new_values and update_history:
            if update_history['new_due_date']:
                updates.append('due_date = ?')
                params.append(update_history['new_due_date'])
                
                # Recalculate review due dates
                date_received = item['date_received']
                priority = update_history['new_priority'] or item['priority']
                if date_received:
                    due_dates = calculate_review_due_dates(
                        date_received, update_history['new_due_date'], priority
                    )
                    updates.extend([
                        'initial_reviewer_due_date = ?',
                        'qcr_due_date = ?',
                        'is_contractor_window_insufficient = ?'
                    ])
                    params.extend([
                        due_dates['initial_reviewer_due_date'],
                        due_dates['qcr_due_date'],
                        1 if due_dates['is_contractor_window_insufficient'] else 0
                    ])
            
            if update_history['new_title']:
                updates.append('title = ?')
                params.append(update_history['new_title'])
            
            if update_history['new_priority']:
                updates.append('priority = ?')
                params.append(update_history['new_priority'])
        
        params.append(item_id)
        cursor.execute(f'''
            UPDATE item SET {', '.join(updates)} WHERE id = ?
        ''', params)
        
        # Clear multi-reviewer responses if applicable
        cursor.execute('''
            UPDATE item_reviewers SET
                response_at = NULL,
                response_category = NULL,
                internal_notes = NULL,
                needs_response = 1
            WHERE item_id = ?
        ''', (item_id,))
        
        if update_history:
            cursor.execute('''
                UPDATE item_update_history SET
                    admin_reviewed_at = ?,
                    admin_reviewed_by = ?,
                    admin_note = ?,
                    action_taken = 'workflow_restarted'
                WHERE id = ?
            ''', (now, admin_user_id, admin_note, update_history['id']))
        
        # Send restart notification to reviewer(s)
        was_closed = item['reopened_from_closed']
        email_result = send_workflow_restart_email(
            item_id, admin_note, was_closed
        )
        result['emails_sent'].append({'to': 'reviewer', 'result': email_result})
    
    conn.commit()
    conn.close()
    
    return jsonify(result)

@app.route('/api/item/<int:item_id>/clear-update-flag', methods=['POST'])
@admin_required
def api_clear_update_flag(item_id):
    """Manually clear the pending update flag without taking action."""
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        UPDATE item SET
            has_pending_update = 0,
            update_type = NULL,
            update_reviewed_at = ?,
            update_admin_note = 'Manually cleared'
        WHERE id = ?
    ''', (datetime.now().isoformat(), item_id))
    
    conn.commit()
    conn.close()
    
    return jsonify({'success': True})

# =============================================================================
# MULTI-REVIEWER API ENDPOINTS
# =============================================================================

@app.route('/api/item/<int:item_id>/multi-reviewer-mode', methods=['POST'])
@admin_required
def api_toggle_multi_reviewer_mode(item_id):
    """Enable or disable multi-reviewer mode for an item."""
    data = request.get_json()
    enabled = data.get('enabled', False)
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('UPDATE item SET multi_reviewer_mode = ? WHERE id = ?', (1 if enabled else 0, item_id))
    conn.commit()
    conn.close()
    
    return jsonify({'success': True, 'multi_reviewer_mode': enabled})

@app.route('/api/item/<int:item_id>/reviewers', methods=['GET'])
@login_required
def api_get_item_reviewers(item_id):
    """Get all reviewers assigned to an item."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT ir.*, u.display_name as user_display_name
        FROM item_reviewers ir
        LEFT JOIN user u ON ir.user_id = u.id
        WHERE ir.item_id = ?
        ORDER BY ir.created_at ASC
    ''', (item_id,))
    reviewers = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify(reviewers)

@app.route('/api/item/<int:item_id>/reviewers', methods=['POST'])
@admin_required
def api_add_item_reviewer(item_id):
    """Add a reviewer to an item (multi-reviewer mode)."""
    data = request.get_json()
    
    user_id = data.get('user_id')
    reviewer_name = data.get('reviewer_name')
    reviewer_email = data.get('reviewer_email')
    
    if not reviewer_name or not reviewer_email:
        return jsonify({'error': 'Reviewer name and email are required'}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    
    # Enable multi-reviewer mode if not already
    cursor.execute('UPDATE item SET multi_reviewer_mode = 1 WHERE id = ?', (item_id,))
    
    # Generate a unique token for this reviewer
    email_token = generate_token()
    
    cursor.execute('''
        INSERT INTO item_reviewers (item_id, user_id, reviewer_name, reviewer_email, email_token)
        VALUES (?, ?, ?, ?, ?)
    ''', (item_id, user_id, reviewer_name, reviewer_email, email_token))
    
    reviewer_id = cursor.lastrowid
    conn.commit()
    conn.close()
    
    return jsonify({
        'success': True,
        'reviewer_id': reviewer_id,
        'message': f'Reviewer {reviewer_name} added successfully'
    })

@app.route('/api/item/<int:item_id>/reviewers/<int:reviewer_id>', methods=['DELETE'])
@admin_required
def api_remove_item_reviewer(item_id, reviewer_id):
    """Remove a reviewer from an item."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM item_reviewers WHERE id = ? AND item_id = ?', (reviewer_id, item_id))
    
    # Check if any reviewers left, if not disable multi-reviewer mode
    cursor.execute('SELECT COUNT(*) FROM item_reviewers WHERE item_id = ?', (item_id,))
    count = cursor.fetchone()[0]
    if count == 0:
        cursor.execute('UPDATE item SET multi_reviewer_mode = 0 WHERE id = ?', (item_id,))
    
    conn.commit()
    conn.close()
    
    return jsonify({'success': True})

@app.route('/api/item/<int:item_id>/send-multi-reviewer-emails', methods=['POST'])
@admin_required
def api_send_multi_reviewer_emails(item_id):
    """Send assignment emails to all reviewers in multi-reviewer mode."""
    try:
        result = send_multi_reviewer_assignment_emails(item_id)
        if result['success']:
            return jsonify(result)
        else:
            return jsonify(result), 400
    except Exception as e:
        return jsonify({'success': False, 'error': f'Server error: {str(e)}'}), 500


@app.route('/api/item/<int:item_id>/send-back-to-reviewers', methods=['POST'])
@admin_required
def api_send_back_to_reviewers(item_id):
    """Send item back to selected reviewers for revision.
    
    Request body:
    {
        "feedback": "Feedback message for reviewers",
        "reviewer_ids": [1, 2]  // Optional - if not provided, sends to all reviewers
    }
    """
    try:
        data = request.get_json()
        feedback = data.get('feedback', '')
        reviewer_ids = data.get('reviewer_ids')  # None = all reviewers
        
        if not feedback:
            return jsonify({'success': False, 'error': 'Feedback message is required'}), 400
        
        result = send_multi_reviewer_sendback_emails(item_id, feedback, reviewer_ids)
        if result['success']:
            return jsonify(result)
        else:
            return jsonify(result), 400
    except Exception as e:
        return jsonify({'success': False, 'error': f'Server error: {str(e)}'}), 500


def check_all_reviewers_responded(item_id):
    """Check if all reviewers have responded.
    
    Checks ALL reviewers for the item, regardless of needs_response flag.
    """
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT COUNT(*) as total, 
               SUM(CASE WHEN response_at IS NOT NULL THEN 1 ELSE 0 END) as responded
        FROM item_reviewers
        WHERE item_id = ?
    ''', (item_id,))
    result = cursor.fetchone()
    conn.close()
    
    if result['total'] == 0:
        return False
    
    return result['total'] == result['responded']

def get_item_reviewer_responses(item_id):
    """Get all reviewer responses for an item."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT * FROM item_reviewers
        WHERE item_id = ?
        ORDER BY created_at ASC
    ''', (item_id,))
    responses = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return responses

# =============================================================================
# MULTI-REVIEWER MAGIC-LINK ROUTES
# =============================================================================

@app.route('/respond/multi-reviewer', methods=['GET'])
def respond_multi_reviewer_form():
    """Show multi-reviewer response form via magic link."""
    token = request.args.get('token')
    if not token:
        return render_template_string(ERROR_PAGE_TEMPLATE, error='Missing token'), 400
    
    conn = get_db()
    cursor = conn.cursor()
    
    # Find the reviewer by token
    cursor.execute('''
        SELECT ir.*, i.*, 
               qcr.display_name as qcr_name
        FROM item_reviewers ir
        JOIN item i ON ir.item_id = i.id
        LEFT JOIN user qcr ON i.qcr_id = qcr.id
        WHERE ir.email_token = ?
    ''', (token,))
    result = cursor.fetchone()
    
    if not result:
        conn.close()
        return render_template_string(ERROR_PAGE_TEMPLATE, error='Invalid or expired token'), 404
    
    item_dict = dict(result)
    reviewer_id = result['id']
    reviewer_name = result['reviewer_name']
    
    # Check if item is closed
    is_closed = result['status'] == 'Closed'
    
    # Check if QCR has finalized
    can_submit = not is_closed and result['qcr_action'] not in ['Approve', 'Modify', 'Complete']
    
    # Check for resubmit (QCR sent back)
    is_resubmit = result['qcr_action'] == 'Send Back'
    qcr_feedback = result['qcr_notes'] if is_resubmit else None
    
    # Get previous response
    previous_response = None
    if result['response_at']:
        previous_response = {
            'category': result['response_category'],
            'notes': result['internal_notes']
        }
    
    # Version tracking
    version = (result['response_version'] or 0) + 1 if is_resubmit else (result['response_version'] or 0)
    
    # Get all reviewers for this item to show status
    cursor.execute('''
        SELECT reviewer_name, response_at FROM item_reviewers WHERE item_id = ?
    ''', (result['item_id'],))
    all_reviewers = [dict(r) for r in cursor.fetchall()]
    
    pending_reviewers = [r for r in all_reviewers if not r['response_at']]
    
    conn.close()
    
    return render_template_string(MULTI_REVIEWER_RESPONSE_TEMPLATE,
        item=item_dict,
        token=token,
        reviewer_name=reviewer_name,
        version=version,
        is_closed=is_closed,
        can_submit=can_submit,
        is_resubmit=is_resubmit,
        qcr_feedback=qcr_feedback,
        previous_response=previous_response,
        all_reviewers=all_reviewers,
        pending_reviewers=pending_reviewers
    )

@app.route('/respond/multi-reviewer', methods=['POST'])
def respond_multi_reviewer_submit():
    """Handle multi-reviewer response submission."""
    token = request.form.get('token')
    if not token:
        return render_template_string(ERROR_PAGE_TEMPLATE, error='Missing token'), 400
    
    conn = get_db()
    cursor = conn.cursor()
    
    # Find the reviewer by token
    cursor.execute('''
        SELECT ir.*, i.id as item_id, i.status, i.qcr_action, i.qcr_id
        FROM item_reviewers ir
        JOIN item i ON ir.item_id = i.id
        WHERE ir.email_token = ?
    ''', (token,))
    reviewer = cursor.fetchone()
    
    if not reviewer:
        conn.close()
        return render_template_string(ERROR_PAGE_TEMPLATE, error='Invalid or expired token'), 404
    
    # Check if item is closed
    if reviewer['status'] == 'Closed':
        conn.close()
        return render_template_string(ERROR_PAGE_TEMPLATE, 
            error='This item has been closed. No further changes can be submitted.'), 403
    
    # Check if QCR has finalized
    if reviewer['qcr_action'] in ['Approve', 'Modify', 'Complete']:
        conn.close()
        return render_template_string(ERROR_PAGE_TEMPLATE,
            error='This item has been finalized. No further changes can be submitted.'), 403
    
    # Get form data
    response_category = request.form.get('response_category')
    internal_notes = request.form.get('internal_notes', '')
    
    # Track if this is after a send-back
    was_sent_back = reviewer['qcr_action'] == 'Send Back'
    new_version = (reviewer['response_version'] or 0) + 1 if was_sent_back else (reviewer['response_version'] or 0)
    
    # Update the reviewer record
    cursor.execute('''
        UPDATE item_reviewers SET
            response_at = ?,
            response_category = ?,
            internal_notes = ?,
            response_version = ?
        WHERE id = ?
    ''', (
        datetime.now().isoformat(),
        response_category,
        internal_notes,
        new_version,
        reviewer['id']
    ))
    
    item_id = reviewer['item_id']
    
    # Check if all reviewers have now responded
    all_responded = check_all_reviewers_responded(item_id)
    
    if all_responded:
        # Update item status to In QC
        cursor.execute('''
            UPDATE item SET 
                status = 'In QC',
                reviewer_response_status = 'All Responded'
            WHERE id = ?
        ''', (item_id,))
        
        # If QCR sent it back, reset QCR state
        if was_sent_back:
            cursor.execute('''
                UPDATE item SET
                    qcr_action = NULL,
                    qcr_response_at = NULL,
                    qcr_response_status = 'Not Sent'
                WHERE id = ?
            ''', (item_id,))
        
        conn.commit()
        
        # Check if QCR email was already sent (avoid duplicates)
        cursor.execute('SELECT qcr_email_sent_at FROM item WHERE id = ?', (item_id,))
        qcr_check = cursor.fetchone()
        qcr_already_notified = qcr_check and qcr_check['qcr_email_sent_at'] is not None
        
        conn.close()
        
        # Send QCR assignment email now that all reviewers have responded
        # Only send if not already sent (avoid duplicates)
        if reviewer['qcr_id'] and not qcr_already_notified:
            send_multi_reviewer_qcr_email(item_id)
        
        return render_template_string(SUCCESS_TEMPLATE,
            message='Your review has been submitted!',
            details='All reviewers have submitted. The QC Reviewer has been notified.'
        )
    else:
        cursor.execute('''
            UPDATE item SET status = 'In Review' WHERE id = ? AND status = 'Assigned'
        ''', (item_id,))
        conn.commit()
        conn.close()
        
        return render_template_string(SUCCESS_TEMPLATE,
            message='Your review has been submitted!',
            details='Waiting for other reviewers to submit before notifying the QC Reviewer.'
        )

@app.route('/respond/multi-qcr', methods=['GET'])
def respond_multi_qcr_form():
    """Show QCR form for multi-reviewer items."""
    token = request.args.get('token')
    if not token:
        return render_template_string(ERROR_PAGE_TEMPLATE, error='Missing token'), 400
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT i.*, 
               qcr.display_name as qcr_name, qcr.email as qcr_email
        FROM item i
        LEFT JOIN user qcr ON i.qcr_id = qcr.id
        WHERE i.email_token_qcr = ? AND i.multi_reviewer_mode = 1
    ''', (token,))
    item = cursor.fetchone()
    
    if not item:
        conn.close()
        return render_template_string(ERROR_PAGE_TEMPLATE, error='Invalid or expired token'), 404
    
    # Check if item is closed
    if item['status'] == 'Closed':
        conn.close()
        return render_template_string(ERROR_PAGE_TEMPLATE,
            error='This item has been closed.'), 403
    
    # Check if already responded
    if item['qcr_response_at'] and item['qcr_action'] in ['Approve', 'Modify', 'Complete']:
        conn.close()
        return render_template_string(ALREADY_RESPONDED_TEMPLATE,
            item=dict(item),
            response_type='qcr'
        )
    
    # Get all reviewer responses
    reviewer_responses = get_item_reviewer_responses(item['id'])
    
    conn.close()
    
    return render_template_string(MULTI_REVIEWER_QCR_TEMPLATE,
        item=dict(item),
        token=token,
        reviewer_responses=reviewer_responses
    )

@app.route('/respond/multi-qcr', methods=['POST'])
def respond_multi_qcr_submit():
    """Handle QCR response for multi-reviewer items."""
    token = request.form.get('token')
    if not token:
        return render_template_string(ERROR_PAGE_TEMPLATE, error='Missing token'), 400
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT i.*, qcr.email as qcr_email
        FROM item i
        LEFT JOIN user qcr ON i.qcr_id = qcr.id
        WHERE i.email_token_qcr = ? AND i.multi_reviewer_mode = 1
    ''', (token,))
    item = cursor.fetchone()
    
    if not item:
        conn.close()
        return render_template_string(ERROR_PAGE_TEMPLATE, error='Invalid or expired token'), 404
    
    if item['status'] == 'Closed':
        conn.close()
        return render_template_string(ERROR_PAGE_TEMPLATE,
            error='This item has been closed.'), 403
    
    # Get form data
    qc_action = request.form.get('qc_action')
    response_category = request.form.get('response_category')
    response_text = request.form.get('response_text', '')
    sendback_notes = request.form.get('sendback_notes', '')
    qcr_internal_notes = request.form.get('qcr_internal_notes', '')
    
    item_id = item['id']
    
    if qc_action == 'Send Back':
        # Reset all reviewer responses and send emails
        cursor.execute('''
            UPDATE item_reviewers SET
                response_at = NULL,
                response_category = NULL,
                internal_notes = NULL
            WHERE item_id = ?
        ''', (item_id,))
        
        cursor.execute('''
            UPDATE item SET
                status = 'In Review',
                qcr_action = 'Send Back',
                qcr_notes = ?,
                qcr_internal_notes = ?,
                qcr_response_status = 'Waiting for Revision'
            WHERE id = ?
        ''', (sendback_notes, qcr_internal_notes, item_id))
        
        conn.commit()
        conn.close()
        
        # Send emails to all reviewers
        send_multi_reviewer_sendback_emails(item_id, sendback_notes)
        
        return render_template_string(SUCCESS_TEMPLATE,
            message='Sent Back to Reviewers',
            details='All reviewers have been notified to revise their responses.'
        )
    else:
        # Complete - store final response
        cursor.execute('''
            UPDATE item SET
                status = 'Ready for Response',
                qcr_response_at = ?,
                qcr_response_status = 'Responded',
                qcr_action = 'Complete',
                qcr_internal_notes = ?,
                final_response_category = ?,
                final_response_text = ?
            WHERE id = ?
        ''', (
            datetime.now().isoformat(),
            qcr_internal_notes,
            response_category,
            response_text,
            item_id
        ))
        
        conn.commit()
        conn.close()
        
        # Create notification
        create_notification(
            'qc_complete',
            f'QC Review Complete: {item["type"]} {item["identifier"]}',
            f'Final response: {response_category}',
            item_id
        )
        
        return render_template_string(SUCCESS_TEMPLATE,
            message='QC Review Complete!',
            details=f'Final response category: {response_category}. The item is now ready for response.'
        )

# =============================================================================
# NOTIFICATIONS API
# =============================================================================

@app.route('/api/notifications', methods=['GET'])
@login_required
def api_get_notifications():
    """Get all notifications."""
    conn = get_db()
    cursor = conn.cursor()
    
    # Clean up response_ready notifications for items that are now closed
    cursor.execute('''
        DELETE FROM notification 
        WHERE type = 'response_ready' 
        AND item_id IN (SELECT id FROM item WHERE status = 'Closed')
    ''')
    conn.commit()
    
    cursor.execute('''
        SELECT n.*, i.type as item_type, i.identifier as item_identifier
        FROM notification n
        LEFT JOIN item i ON n.item_id = i.id
        ORDER BY n.created_at DESC
        LIMIT 100
    ''')
    notifications = [dict(row) for row in cursor.fetchall()]
    
    # Get unread count
    cursor.execute('SELECT COUNT(*) FROM notification WHERE read_at IS NULL')
    unread_count = cursor.fetchone()[0]
    
    conn.close()
    return jsonify({'notifications': notifications, 'unread_count': unread_count})

@app.route('/api/notifications/<int:notification_id>/read', methods=['POST'])
@login_required
def api_mark_notification_read(notification_id):
    """Mark a notification as read."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE notification SET read_at = ? WHERE id = ?
    ''', (datetime.now().isoformat(), notification_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/notifications/read-all', methods=['POST'])
@login_required
def api_mark_all_notifications_read():
    """Mark all notifications as read."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE notification SET read_at = ? WHERE read_at IS NULL
    ''', (datetime.now().isoformat(),))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/notifications/<int:notification_id>', methods=['DELETE'])
@login_required
def api_delete_notification(notification_id):
    """Delete a notification."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM notification WHERE id = ?', (notification_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/items/<int:item_id>/complete', methods=['POST'])
@admin_required
def api_mark_item_complete(item_id):
    """Mark an item as complete/closed."""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE item SET 
            status = 'Closed',
            closed_at = ?
        WHERE id = ?
    ''', (datetime.now().isoformat(), item_id))
    
    # Delete any response_ready notifications for this item since it's now closed
    cursor.execute('''
        DELETE FROM notification 
        WHERE item_id = ? AND type = 'response_ready'
    ''', (item_id,))
    
    conn.commit()
    conn.close()
    
    return jsonify({'success': True, 'message': 'Item marked as complete'})

# =============================================================================
# FILE-BASED FORM API ENDPOINTS
# =============================================================================

@app.route('/api/items/<int:item_id>/generate-reviewer-form', methods=['POST'])
@admin_required
def api_generate_reviewer_form(item_id):
    """Generate a self-contained HTML reviewer form in the item folder."""
    try:
        result = generate_reviewer_form_html(item_id)
        if result['success']:
            return jsonify(result)
        else:
            return jsonify(result), 400
    except Exception as e:
        return jsonify({'success': False, 'error': f'Server error: {str(e)}'}), 500


@app.route('/api/items/<int:item_id>/generate-qcr-form', methods=['POST'])
@admin_required
def api_generate_qcr_form(item_id):
    """Generate a self-contained HTML QCR form in the item folder."""
    try:
        result = generate_qcr_form_html(item_id)
        if result['success']:
            return jsonify(result)
        else:
            return jsonify(result), 400
    except Exception as e:
        return jsonify({'success': False, 'error': f'Server error: {str(e)}'}), 500


@app.route('/api/scan-folder-responses', methods=['POST'])
@admin_required
def api_scan_folder_responses():
    """Manually trigger a scan for JSON response files."""
    try:
        results = scan_folders_for_responses()
        return jsonify({
            'success': True,
            'reviewer_responses_imported': len(results['reviewer_responses']),
            'qcr_responses_imported': len(results['qcr_responses']),
            'errors': results['errors'],
            'details': results
        })
    except Exception as e:
        return jsonify({'success': False, 'error': f'Server error: {str(e)}'}), 500


@app.route('/api/watcher-status', methods=['GET'])
@login_required
def api_watcher_status():
    """Get the status of the folder response watcher."""
    return jsonify({
        'running': folder_watcher.running,
        'last_scan': folder_watcher.last_scan.isoformat() if folder_watcher.last_scan else None,
        'scan_count': folder_watcher.scan_count,
        'interval_seconds': folder_watcher.interval
    })


@app.route('/api/reminder-status', methods=['GET'])
@login_required
def api_reminder_status():
    """Get the status of the reminder scheduler."""
    return jsonify({
        'running': reminder_scheduler.running,
        'last_check': reminder_scheduler.last_check.isoformat() if reminder_scheduler.last_check else None,
        'last_reminder_date': reminder_scheduler.last_reminder_date.isoformat() if reminder_scheduler.last_reminder_date else None,
        'check_interval_seconds': reminder_scheduler.interval,
        'reminder_time_pst': f"{REMINDER_HOUR_PST}:00 AM PST",
        'is_past_reminder_time': is_past_reminder_time_today()
    })


@app.route('/api/process-reminders', methods=['POST'])
@admin_required
def api_process_reminders():
    """Manually trigger reminder processing."""
    try:
        results = process_all_reminders()
        return jsonify({
            'success': True,
            'results': results
        })
    except Exception as e:
        return jsonify({'success': False, 'error': f'Server error: {str(e)}'}), 500


@app.route('/api/pending-reminders', methods=['GET'])
@login_required
def api_pending_reminders():
    """Get list of items that would receive reminders today."""
    try:
        items = get_items_needing_reminders()
        return jsonify({
            'success': True,
            'single_reviewer': [
                {
                    'item_id': item['id'],
                    'identifier': item['identifier'],
                    'role': role,
                    'due_date': due_date.strftime('%Y-%m-%d'),
                    'reminder_stage': stage,
                    'recipient': item['reviewer_email'] if role == 'reviewer' else item['qcr_email']
                }
                for item, role, due_date, stage in items['single_reviewer']
            ],
            'multi_reviewer': [
                {
                    'item_id': item['id'],
                    'identifier': item['identifier'],
                    'reviewer_name': reviewer['reviewer_name'],
                    'reviewer_email': reviewer['reviewer_email'],
                    'due_date': due_date.strftime('%Y-%m-%d'),
                    'reminder_stage': stage
                }
                for item, reviewer, role, due_date, stage in items['multi_reviewer']
            ],
            'multi_reviewer_qcr': [
                {
                    'item_id': item['id'],
                    'identifier': item['identifier'],
                    'qcr_email': item['qcr_email'],
                    'due_date': due_date.strftime('%Y-%m-%d'),
                    'reminder_stage': stage
                }
                for item, due_date, stage in items['multi_reviewer_qcr']
            ]
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/reminder-history', methods=['GET'])
@login_required
def api_reminder_history():
    """Get history of sent reminders."""
    conn = get_db()
    cursor = conn.cursor()
    
    # Get all reminders with item info, ordered by most recent
    cursor.execute('''
        SELECT r.*, i.identifier, i.type, i.title
        FROM reminder_log r
        LEFT JOIN item i ON r.item_id = i.id
        ORDER BY r.sent_at DESC
        LIMIT 200
    ''')
    reminders = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    return jsonify({
        'success': True,
        'reminders': reminders
    })


@app.route('/api/items/<int:item_id>/send-reminder', methods=['POST'])
@admin_required
def api_send_item_reminder(item_id):
    """Manually send a reminder for a specific item."""
    from datetime import date
    
    conn = get_db()
    cursor = conn.cursor()
    
    # Get item with reviewer and QCR info
    cursor.execute('''
        SELECT i.*, 
               ir.display_name as reviewer_name, ir.email as reviewer_email,
               qcr.display_name as qcr_name, qcr.email as qcr_email
        FROM item i
        LEFT JOIN user ir ON i.initial_reviewer_id = ir.id
        LEFT JOIN user qcr ON i.qcr_id = qcr.id
        WHERE i.id = ?
    ''', (item_id,))
    item = cursor.fetchone()
    
    if not item:
        conn.close()
        return jsonify({'success': False, 'error': 'Item not found'}), 404
    
    # Check if item is closed - don't send reminders for closed items
    if item['closed_at']:
        conn.close()
        return jsonify({'success': False, 'error': 'Item is closed - no reminder needed'})
    
    # Check if status indicates item is not in anyone's court for review
    if item['status'] in ('Closed', 'Ready for Response'):
        conn.close()
        return jsonify({'success': False, 'error': f'Item status is "{item["status"]}" - no reminder needed'})
    
    # Check if multi-reviewer
    cursor.execute('SELECT COUNT(*) FROM item_reviewers WHERE item_id = ?', (item_id,))
    reviewer_count = cursor.fetchone()[0]
    is_multi_reviewer = reviewer_count > 0
    
    results = []
    today = date.today()
    
    if is_multi_reviewer:
        # Only send reviewer reminders if status indicates it's in reviewer's court
        if item['status'] in ('Assigned', 'In Review'):
            # Get reviewers who haven't responded
            cursor.execute('''
                SELECT * FROM item_reviewers 
                WHERE item_id = ? AND responded_at IS NULL
            ''', (item_id,))
            pending_reviewers = cursor.fetchall()
            
            for reviewer in pending_reviewers:
                result = send_multi_reviewer_reminder_email(
                    dict(item), 
                    dict(reviewer), 
                    'reviewer', 
                    today,
                    'manual'
                )
                results.append({
                    'recipient': reviewer['reviewer_email'],
                    'role': 'reviewer',
                    'success': result.get('success', False),
                    'error': result.get('error')
                })
        
        # Only send QCR reminders if status indicates it's in QCR's court
        if item['status'] == 'In QC' and item['qcr_email'] and not item['qcr_response_at']:
            result = send_multi_reviewer_qcr_reminder_email(dict(item), today, 'manual')
            results.append({
                'recipient': item['qcr_email'],
                'role': 'qcr',
                'success': result.get('success', False),
                'error': result.get('error')
            })
        
        conn.close()
    else:
        conn.close()
        # Single reviewer mode - only send based on current status
        if item['status'] in ('Assigned', 'In Review') and item['reviewer_email'] and not item['reviewer_response_at']:
            # Status indicates reviewer's turn
            result = send_single_reviewer_reminder_email(dict(item), 'reviewer', today, 'manual')
            results.append({
                'recipient': item['reviewer_email'],
                'role': 'reviewer',
                'success': result.get('success', False),
                'error': result.get('error')
            })
        elif item['status'] == 'In QC' and item['qcr_email'] and not item['qcr_response_at']:
            # Status indicates QCR's turn
            result = send_single_reviewer_reminder_email(dict(item), 'qcr', today, 'manual')
            results.append({
                'recipient': item['qcr_email'],
                'role': 'qcr',
                'success': result.get('success', False),
                'error': result.get('error')
            })
    
    if not results:
        return jsonify({
            'success': False, 
            'error': f'No reminders needed - item status is "{item["status"]}" which does not require a reminder'
        })
    
    return jsonify({
        'success': True,
        'results': results
    })

# =============================================================================
# STATIC FILE ROUTES
# =============================================================================

@app.route('/')
def index():
    """Serve the main dashboard."""
    return send_from_directory('static', 'index.html')

@app.route('/<path:filename>')
def serve_static(filename):
    """Serve static files."""
    return send_from_directory('static', filename)

# =============================================================================
# MAIN ENTRY POINT
# =============================================================================

def main():
    """Main entry point."""
    print("=" * 60)
    print("LEB RFI/Submittal Tracker")
    print("=" * 60)
    
    # Initialize database
    print("Initializing database...")
    init_db()
    
    # Create base folder
    base_folder = Path(CONFIG['base_folder_path'])
    base_folder.mkdir(parents=True, exist_ok=True)
    print(f"Base folder: {base_folder}")
    
    # Sync Airtable responses if configured
    if HAS_AIRTABLE:
        print("Checking for Airtable responses...")
        try:
            config = load_airtable_config()
            if config.get('api_key') and config.get('base_id'):
                result = sync_airtable_responses()
                if result['synced_count'] > 0:
                    print(f"  Synced {result['synced_count']} responses from Airtable")
                else:
                    print("  No new Airtable responses to sync")
                if result.get('errors'):
                    for err in result['errors']:
                        print(f"  Warning: {err}")
            else:
                print("  Airtable not configured (add credentials to config.json)")
        except Exception as e:
            print(f"  Airtable sync error: {e}")
    
    # Scan for file-based responses on startup
    print("Scanning for file-based responses...")
    try:
        results = scan_folders_for_responses()
        total_imported = len(results['reviewer_responses']) + len(results['qcr_responses'])
        if total_imported > 0:
            print(f"  Imported {total_imported} response(s) from folder files")
        else:
            print("  No pending file responses found")
        for err in results.get('errors', []):
            print(f"  Warning: {err}")
    except Exception as e:
        print(f"  Folder scan error: {e}")
    
    # Start email poller
    print("Starting email poller...")
    email_poller.start()
    
    # Start folder response watcher
    print("Starting folder response watcher...")
    folder_watcher.start()
    
    # Start reminder scheduler
    print("Starting reminder scheduler...")
    reminder_scheduler.start()
    
    # Process any pending reminders on startup (in case server was off at 8 AM)
    print("Checking for pending reminders...")
    try:
        results = process_all_reminders()
        total_sent = results.get('single_reviewer_sent', 0) + results.get('multi_reviewer_sent', 0) + results.get('multi_reviewer_qcr_sent', 0)
        if total_sent > 0:
            print(f"  Sent {total_sent} pending reminder(s)")
        elif results.get('processed') == False:
            print(f"  {results.get('reason', 'No reminders needed')}")
        else:
            print("  No reminders needed")
    except Exception as e:
        print(f"  Reminder check error: {e}")
    
    # Start web server
    port = CONFIG.get('server_port', 5000)
    print(f"\nStarting web server on http://localhost:{port}")
    print("Press Ctrl+C to stop\n")
    print("=" * 60)
    
    try:
        app.run(host='127.0.0.1', port=port, debug=False, threaded=True)
    except KeyboardInterrupt:
        print("\nShutting down...")
        email_poller.stop()
        folder_watcher.stop()
        reminder_scheduler.stop()

if __name__ == '__main__':
    main()
